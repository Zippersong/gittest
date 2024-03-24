# -*- coding: utf-8 -*-
"""
Created on Mon May 13 09:49:34 2019

@author: lenovo
"""

#%%------------调整excel的列宽

from openpyxl import load_workbook
from openpyxl.styles import Alignment
import os

os.chdir(r"D:\委外回款周报\线下近三外包\新建文件夹\债券转让合同明细2")
alist=os.listdir(r"D:\委外回款周报\线下近三外包\新建文件夹\债券转让合同明细2") 


alignment_center = Alignment(horizontal='center', vertical='center')  
for i in range(len(alist)):
# path = '样稿.xlsx'
    # print(alist[i])
    wb = load_workbook(alist[i])
    ws = wb.active

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 11
    
# 指定区域单元格居中

    # ws_area = ws["A1:F2000"]#F后面的数字利用sas那边查看单个客户还款记录的上限（1560）

    for ii in ws:#ws自动获ws取篇幅大小了
        for j in ii:
            j.alignment = alignment_center;
    #为了给下标+2的A列单元格内添加序号并居中       
    a2=ws.dimensions.split(':')[1]
    a3=a2[1-len(a2):]
    temp_roa=str(int(a2[1-len(a2):])+2)
    ws['A'+temp_roa].value=i+1
    ws['A'+temp_roa].alignment=alignment_center

    wb.save(alist[i])


#下面的都是草稿
os.chdir(r"D:\委外回款周报\线下近三外包\新建文件夹\test")
alist=os.listdir(r"D:\委外回款周报\线下近三外包\新建文件夹\test") 
for i in range(len(alist)):
    print(i)
    a=str(i+1)
    wb = load_workbook(alist[i])
    b=alist[i].split('.')[0]+'_'+a+'.xlsx'
    wb.save(b)

a=alist[0].split('.')[0]
# dir(Alignment(horizontal='center', vertical='center'))

wb = load_workbook('C151323518112403000000384.xlsx')

alignment_center = Alignment(horizontal='center', vertical='center')  
ws = wb.active
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 6
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 11
print(ws.dimensions)
for ii in ws:
    for j in ii:
        j.alignment = alignment_center;
a2=ws.dimensions.split(':')[1]
a3=a2[1-len(a2):]
temp_roa=str(int(a2[1-len(a2):])+2)

ws['A'+temp_roa].value=1
ws['A'+temp_roa].alignment=alignment_center
wb.save('C151323518112403000000384.xlsx')

#%%------------网上评分卡案例





#%%------------jupyter开发
os.chdir(r"F:\菲律宾日报")
# appr_mart = pd.read_excel("appr_mart.xlsx",dtype={'phone_no':str})
# appr_s1 = appr_mart.query("apply_time<='2020-06-15' and apply_time>='2020-05-12'")
appr_st = appr_mart.query("apply_time>='2021-01-01'").copy()
# repay_mart = pd.read_excel("repay_mart.xlsx",dtype={'phone_no':str})
repay_st = repay_mart.query("loan_date>='2021-01-01'").copy()

first_clear_date = repay_mart.loc[repay_mart.account_status=='ACCOUNT_SETTLE',['phone_no','clear_date']]
# first_clear_date.head(5)
first_clear_date.sort_values(by=['phone_no','clear_date'],inplace=True)
first_clear_date.drop_duplicates(subset='phone_no',keep='first',inplace=True)
first_clear_date['first_clear_time'] = pd.to_datetime(first_clear_date.clear_date)
first_clear_date.drop(columns='clear_date',inplace=True)
repay_st = pd.merge(repay_st,first_clear_date,how='left',on='phone_no')
repay_st['newOld'] = repay_st.apply(lambda x: 'new customer' if pd.isnull(x.first_clear_time) or x.loan_date<x.first_clear_time else 'old customer',axis=1)
appr_st = pd.merge(appr_st,first_clear_date,how='left',on='phone_no')
appr_st['newOld'] = appr_st.apply(lambda x: 'new customer' if pd.isnull(x.first_clear_time) or x.apply_time<x.first_clear_time else 'old customer',axis=1)



#白名单
repay_mart_w=repay_st[repay_st.customer_source_sys!='PesoCash']
pagket=['SunCash','SuncashPautang','FlashLoan']

for (i,css) in enumerate(pagket):
    temp=repay_mart_w.loc[((repay_mart_w.customer_source_sys==css) & (repay_mart_w.loan_type=='first_apply')),['phone_no','loan_date','到期','自然逾期','customer_source_sys','last_repay_month','last_repay_date']]
    temp.sort_values(by=['phone_no','loan_date'],inplace=True)
    temp.drop_duplicates(subset='phone_no',keep='first',inplace=True)
    
    temp_repayment=repay_mart_w.loc[repay_mart_w.customer_source_sys!=css,['phone_no','clear_date']]
    temp_repayment=temp_repayment[temp_repayment.clear_date.notna()]
    temp_repayment['settled']=1
    temp_repayment.sort_values(by=['phone_no','clear_date'],inplace=True)
    temp_repayment.drop_duplicates(subset='phone_no',keep='last',inplace=True)
    
    temp_data=pd.merge(temp,temp_repayment,how='left',left_on='phone_no',right_on='phone_no')
    if i==0:
        data = temp_data
    else:
        data = pd.concat([data,temp_data])
data['loan_date'] = pd.to_datetime(data['loan_date']).dt.date
data['W']=data.apply(lambda x : 1 if (x.settled==1) and  (x.loan_date-x.clear_date).days>=0 else 0,axis=1)

#%%------------文字工程
os.chdir(r"F:\菲律宾日报\建模")
wz = pd.read_excel("文字工程.xlsx")
thisset=set()
for i in range(len(wz)):
    print(i)
    a=wz.iloc[i,0].split(',')
    thisset=thisset.union(a)
sss=pd.DataFrame({'code':list(thisset),'loaned':1})   
sss.to_excel(r"app分类单词.xlsx",index=False)



#%%------------减免清单

thisset = {"apple", "banana", "cherry"}
thissetaaaa={"orange","aa"}
thisset.add("a")
thisset.union(thisset1)

thisset1 = ["apple", "banana", "cherry","asdfasd"]

thisset1.append(['a','b'])


print(thisset)



import openpyxl
import json
import numpy as np
import pandas as pd
import sys
sys.path.append(r"C:\Users\lenovo\Anaconda3\Lib\site-packages")
import pymysql
import math
import datetime
import gc
#gc.collect() 垃圾回收，返回处理这些循环引用一共释放掉的对象个数
import os
import oss2
import zipfile
import saspy
from itertools import islice
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = [u'SimHei']
plt.rcParams['axes.unicode_minus'] = False

os.chdir(r"F:\菲律宾日报")
with open(r"db_config.json") as db_config:
    cnx_args = json.load(db_config)

cnx = pymysql.connect(**cnx_args)  


app_info=pd.read_sql("select * from approval.app_info  ",cnx)
lockup_period=pd.read_sql("select * from approval.lockup_period  ",cnx)
white_list_info=pd.read_sql("select * from approval.white_list_info  ",cnx)

user_f=pd.read_sql("select * from flash_loan.user  ",cnx)
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
sas.df2sd(user_f.astype(str),'user_f',"MY",encode_errors='replace')#appr_mart特殊，单独搞

sas.df2sd(white_list_info.astype(str),'white_list_info',"MY",encode_errors='replace')#appr_mart特殊，单独搞
sas.df2sd(app_info.astype(str),'app_info',"MY",encode_errors='replace')#appr_mart特殊，单独搞
sas.df2sd(lockup_period.astype(str),'lockup_period',"MY",encode_errors='replace')#appr_mart特殊，单独搞

my_case_info=pd.read_sql("select CREATED_TIME,USER_NAME,CONTRACT_NO from collection.my_case_info  ",cnx)
# LAST_REPAY_DATE='2022-12-15'
account_info=pd.read_sql("select SYSTEM_SOURCE,contract_no,BORROWER_TEL_ONE,name,LAST_REPAY_DATE,CUSTOMER_TYPE,CONTRACT_AMOUNT from account.account_info \
                         where \
                         LAST_REPAY_DATE>=DATE_SUB(curdate(),INTERVAL 30 DAY) \
                         and ACCOUNT_STATUS='ACCOUNT_OVERDUE' ",cnx)
account_info.to_excel(r"1215_od.xlsx",index=False)

account_info=pd.read_sql("select SYSTEM_SOURCE,contract_no,BORROWER_TEL_ONE,name,LAST_REPAY_DATE,CUSTOMER_TYPE,CONTRACT_AMOUNT from account.account_info \
                         where \
                         LAST_REPAY_DATE>='2022-12-16' \
                         and ACCOUNT_STATUS='ACCOUNT_NORMAL' ",cnx)
account_info.to_excel(r"1216及以后.xlsx",index=False)



my_case_info.sort_values(by=['CONTRACT_NO','CREATED_TIME'],inplace=True)
my_case_info.drop_duplicates(subset='CONTRACT_NO',keep='last',inplace=True)
lista = pd.read_excel(r"F:\菲律宾日报\减免清单.xlsx",sheet_name=r'SheetJS')


lista=pd.merge(lista,my_case_info,how='left',left_on='contract_no',right_on='CONTRACT_NO')
account_info=pd.read_sql("select CONTRACT_NO,BORROWER_TEL_ONE from account.account_info  ",cnx)
lista=pd.merge(lista,account_info,how='left',left_on='contract_no',right_on='CONTRACT_NO')


lista.to_excel(r'F:\菲律宾日报\减免清单1.xlsx',index=False)

paymengt_info=pd.read_sql("select * from approval.payment_info  ",cnx)

image_info=pd.read_sql("select apply_code,file_desc from approval.image_info  ",cnx)

collection_log_30=pd.read_sql("select COLLECTION_DATE,CREATED_TIME,CONTRACT_NO,COLLECTION_RESULT_CODE,RELATIONSHIP,USER_NAME,CONTACT_PHONE,remark from collection.collection_log_info where DATE(CREATED_TIME)>=DATE_SUB(curdate(),INTERVAL 30 DAY) ",cnx)

s_credit_limit=pd.read_sql("select * from suncash_lend.credit_limit  ",cnx)
f_credit_limit=pd.read_sql("select * from flash_loan.credit_limit  ",cnx)


loan_info=pd.read_sql("select * from approval.loan_info  ",cnx)

repay_plan=pd.read_sql("select * from account.repay_plan  ",cnx)
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
sas.df2sd(repay_plan.astype(str),'repay_plan',"MY",encode_errors='replace')#appr_mart特殊，单独搞
sas.df2sd(s_credit_limit.astype(str),'s_credit_limit',"MY",encode_errors='replace')#appr_mart特殊，单独搞
sas.df2sd(f_credit_limit.astype(str),'f_credit_limit',"MY",encode_errors='replace')#appr_mart特殊，单独搞

sas.df2sd(paymengt_info.astype(str),'payment_info',"MY",encode_errors='replace')#appr_mart特殊，单独搞

sas.df2sd(image_info.astype(str),'image_info',"MY",encode_errors='replace')#appr_mart特殊，单独搞

sas.df2sd(collection_log_30.astype(str),'collection_log_30',"MY",encode_errors='replace')#appr_mart特殊，单独搞



# 尝试弄全表
importfiles = [
# 'approval.apply_info',
# 'approval.auto_audit_result',
# 'approval.manual_audit_result',
# 'approval.borrower_info',
# 'approval.employment_info',
# 'approval.loan_info',
'approval.contract_info'
]

os.chdir(r"F:\菲律宾日报\json") 
for fa in importfiles:
    query = f"SELECT * FROM {fa};"
    temp_table=pd.read_sql(query,cnx)
    temp_table.to_csv(fa+".csv",index=False)
    #去重


df1 = pd.read_sql(query, cnx)
os.chdir(r"F:\菲律宾日报\2022-10-04") 
temp = pd.read_csv('approval.borrower_info.csv',dtype=dtype,low_memory=False)


collection_log_latest=pd.read_sql("select COLLECTION_DATE,CONTRACT_NO,COLLECTION_RESULT_CODE,CONTACT_PHONE from collection.collection_log_info  ",cnx)
sas.df2sd(collection_log_latest.astype(str),'collection_log_latest',"MY",encode_errors='replace')#appr_mart特殊，单独搞

collection_log_latest=pd.read_sql("select contract_no,name  from account.account_info  ",cnx)
sas.df2sd(collection_log_latest.astype(str),'collection_log_latest',"MY",encode_errors='replace')#appr_mart特殊，单独搞



#%%------------模型开发







#%%------------组长近一个月的调件记录

# my_case_info：如果有多个M就会有多条，每一条存的是每个M的最新数据		
# case_flow_info：人工转件记录	没存M值发生变动时的记录，因为0点M值发生变化，案件原催收员关闭，到凌晨5点才开始自动分件，只有自动分件后才有M值变动后的催收员信息	
# 	每个M状态的第一条都是自动分配，也不会存进case_flow_info表	
# 	其实flow表里A转给B，C转给D，根据时间排序下，就知道中间的M值变动时B转给C的	



import openpyxl
import json
import numpy as np
import pandas as pd
import sys
sys.path.append(r"C:\Users\lenovo\Anaconda3\Lib\site-packages")
import pymysql
import math
import datetime
import gc
#gc.collect() 垃圾回收，返回处理这些循环引用一共释放掉的对象个数
import os
import oss2
import zipfile
import saspy
from itertools import islice
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = [u'SimHei']
plt.rcParams['axes.unicode_minus'] = False

os.chdir(r"F:\菲律宾日报")
with open(r"db_config.json") as db_config:
    cnx_args = json.load(db_config)

cnx = pymysql.connect(**cnx_args)  


# 抹去了，好像是用sql的in（组长名字）
my_case_info=pd.read_sql("select CREATED_TIME,USER_NAME,CONTRACT_NO from collection.my_case_info  ",cnx)

my_case_info.sort_values(by=['CONTRACT_NO','CREATED_TIME'],inplace=True)
my_case_info.drop_duplicates(subset='CONTRACT_NO',keep='last',inplace=True)
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')

sas.df2sd(my_case_info.astype(str),'my_case_info_ls',"MY",encode_errors='replace')#appr_mart特殊，单独搞






# case_flow_info.info()
case_flow_info=case_flow_info[['CREATED_USER_NAME','CREATED_TIME','OLD_USER_NAME','USER_NAME','REASON_FOR_TRANSFER_NAME','CONTRACT_NO']]
# Edgar Rey Jasmin Cheska  Girlie

sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
repay_mart=sas.sd2df('repay_mart','MY')
repay_martt=repay_mart[['contract_amount','customer_source_sys','loan_type','contract_no']]
data_end=pd.merge(case_flow_info,repay_martt,how='left',left_on='CONTRACT_NO',right_on='contract_no')

data_end.to_excel(r"F:\菲律宾日报\各组长20220707调件记录.xlsx",index=False)
#vintage.to_excel(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\vintage.xlsx",index=False)

# aaa=pd.read_sql("select *  from collection.case_flow_info where  DATE(created_time) >= DATE_SUB(curdate(),INTERVAL 160 DAY) ",cnx)#只需要近1个月，随便写个35
# aaa.CREATED_USER_NAME.value_counts()


#%%------------PSI
import pandas as pd
import numpy as np
import random
import math
from scipy.stats import chi2
import scipy


import pandas as pd
import numpy as np 

hangye= ['农业','农业','林业','林业','养殖业','养殖业']
region = ['农村','城市','城市','农村','城市','农村']
id = [1,2,3,4,5,6]
price =[234,123,34,2343,45,54]

df = pd.DataFrame({'id':id,'hangye':hangye,'price':price,'region':region})
df1=pd.crosstab(index=df['hangye'],
            columns=df['region']
            )
freq_array = df1.values
b=freq_array.cumsum(axis=0)
d=freq_array.sum(axis=0)
e=np.ones(freq_array.shape) * freq_array.sum(axis=0)
f=np.ones(freq_array.shape) * freq_array.sum(axis=0).T

# 测试数据构造，其中target为Y，1代表坏人，0代表好人。  
df = pd.read_csv('./data/autocut_testdata.csv')
print(len(df))
print(df.target.value_counts()/len(df))
print(df.head())

def get_maxks_split_point(data, var, target, min_sample=0.05):
    """ 计算KS值
    Args:
        data: DataFrame，待计算卡方分箱最优切分点列表的数据集
        var: 待计算的连续型变量名称
        target: 待计算的目标列Y的名称
        min_sample: int，分箱的最小数据样本，也就是数据量至少达到多少才需要去分箱，一般作用在开头或者结尾处的分箱点
    Returns:
        ks_v: KS值，float
        BestSplit_Point: 返回本次迭代的最优划分点，float
        BestSplit_Position: 返回最优划分点的位置，最左边为0，最右边为1，float
    """
    if len(data) < min_sample:
        ks_v, BestSplit_Point, BestSplit_Position = 0, -9999, 0.0
    else:
        freq_df = pd.crosstab(index=data[var], columns=data[target])
        freq_array = freq_df.values
        if freq_array.shape[1] == 1: # 如果某一组只有一个枚举值，如0或1，则数组形状会有问题，跳出本次计算
            # tt = np.zeros(freq_array.shape).T
            # freq_array = np.insert(freq_array, 0, values=tt, axis=1)
            ks_v, BestSplit_Point, BestSplit_Position = 0, -99999, 0.0
        else:
            bincut = freq_df.index.values
            tmp = freq_array.cumsum(axis=0)/(np.ones(freq_array.shape) * freq_array.sum(axis=0).T)
            tmp_abs = abs(tmp.T[0] - tmp.T[1])
            ks_v = tmp_abs.max()
            BestSplit_Point = bincut[tmp_abs.tolist().index(ks_v)]
            BestSplit_Position = tmp_abs.tolist().index(ks_v)/max(len(bincut) - 1, 1)
        
    return ks_v, BestSplit_Point, BestSplit_Position


def get_bestks_bincut(data, var, target, leaf_stop_percent=0.05):
    """ 计算最优分箱切分点
    Args:
        data: DataFrame，拟操作的数据集
        var: String，拟分箱的连续型变量名称
        target: String，Y列名称
        leaf_stop_percent: 叶子节点占比，作为停止条件，默认5%
    
    Returns:
        best_bincut: 最优的切分点列表，List
    """
    min_sample = len(data) * leaf_stop_percent
    best_bincut = []
    
    def cutting_data(data, var, target, min_sample, best_bincut):
        ks, split_point, position = get_maxks_split_point(data, var, target, min_sample)
        
        if split_point != -99999:
            best_bincut.append(split_point)
        
        # 根据最优切分点切分数据集，并对切分后的数据集递归计算切分点，直到满足停止条件
        # print("本次分箱的值域范围为{0} ~ {1}".format(data[var].min(), data[var].max()))
        left = data[data[var] < split_point]
        right = data[data[var] > split_point]
        
        # 当切分后的数据集仍大于最小数据样本要求，则继续切分
        if len(left) >= min_sample and position not in [0.0, 1.0]:
            cutting_data(left, var, target, min_sample, best_bincut)
        else:
            pass
        if len(right) >= min_sample and position not in [0.0, 1.0]:
            cutting_data(right, var, target, min_sample, best_bincut)
        else:
            pass
        return best_bincut
    best_bincut = cutting_data(data, var, target, min_sample, best_bincut)
    
    # 把切分点补上头尾
    best_bincut.append(data[var].min())
    best_bincut.append(data[var].max())
    best_bincut_set = set(best_bincut)
    best_bincut = list(best_bincut_set)
    
    best_bincut.remove(data[var].min())
    best_bincut.append(data[var].min()-1)
    # 排序切分点
    best_bincut.sort()
    
    return best_bincut
    

age_bins=get_bestks_bincut(df, 'age', 'target')
df['age_bins'] = pd.cut(df['age'], bins=age_bins)
print("age的最优分箱切分点：", age_bins)
print("age的最优分箱结果：\n", df['age_bins'].value_counts())
df.head()

# 导入测试数据集
df_test = pd.read_csv('./data/psi_testdata.csv')

def cal_psi(df_train, df_test, var, target):
    train_bins = get_bestks_bincut(df_train, var, target)
    train_cut_nums = pd.cut(df_train[var], bins=train_bins).value_counts().sort_index().values
    
    # 根据训练集分箱切分点对测试集进行切分
    def cut_test_data(data, var, bincut):
        # 扩大两端边界线
        if bincut[0] > data[var].min()-1:
            bincut.remove(bincut[0])
            bincut.append(data[var].min()-1)
        if bincut[-1] < data[var].max():
            bincut.remove(bincut[-1])
            bincut.append(data[var].max())
    
        # 排序切分点
        bincut.sort()
        return bincut
    
    test_cut_nums = pd.cut(df_test[var], 
                            bins=cut_test_data(df_test, var, train_bins)).value_counts().sort_index().values
    
    tt = pd.DataFrame(np.vstack((train_cut_nums, test_cut_nums)).T, columns=['train', 'test'])
    
    # 计算PSI
    E = tt['train'].values/tt['train'].values.sum(axis=0)
    A = tt['test'].values/tt['test'].values.sum(axis=0)
    A_sub_E = A-E
    A_divide_E = A/E
    ln_A_divide_E = np.log(A_divide_E) # numpy里的log其实指的是ln
    PSI_i = A_sub_E * ln_A_divide_E
    psi = PSI_i.sum()
    return tt, psi
    
tt, psi = cal_psi(df, df_test, 'age', 'target')
print("PSI: ", psi)
tt





#%%------------卡方分箱
def chi2(arr):
    
    assert(arr.ndim==2)
    #计算每行总频数
    R_N = arr.sum(axis=1)
    #每列总频数
    C_N = arr.sum(axis=0)
    #总频数
    N = arr.sum()
    # 计算期望频数 C_i * R_j / N
    E = np.ones(arr.shape)* C_N / N
    E = (E.T * R_N).T
    square = (arr-E)**2/ E
    #期望频数为0时，做除数没有意义，不计入卡方值
    square[E==0] = 0
    #卡方值
    v = square.sum()
    return v
def chiMerge(df,col,target,max_groups=None,threshold=None): 
    '''
    卡方分箱
    df: pandas dataframe数据集
    col: 需要分箱的变量名（数值型）
    target: 类标签
    max_groups: 最大分组数。
    threshold: 卡方阈值，如果未指定max_groups，默认使用置信度95%设置threshold。
    return: 包括各组的起始值的列表.
    '''
    freq_tab = pd.crosstab(df[col],df[target])
#转成numpy数组用于计算。
    freq = freq_tab.values
#初始分组切分点，每个变量值都是切分点。每组中只包含一个变量值.
#分组区间是左闭右开的，如cutoffs = [1,2,3]，则表示区间 [1,2) , [2,3) ,[3,3+)。
    cutoffs = freq_tab.index.values
#如果没有指定最大分组
    if max_groups is None :        
#如果没有指定卡方阈值，就以95%的置信度（自由度为类数目-1）设定阈值。     
        if threshold is None:         
#类数目
            cls_num = freq.shape[-1]
            threshold = chi2.isf(0.05,df= cls_num - 1)
    while True :
        minvalue = None
        minidx = None       
#从第1组开始，依次取两组计算卡方值，并判断是否小于当前最小的卡方       
        for i in range(len(freq) - 1):
            v = chi2(freq[i:i+2])
            if minvalue is None or minvalue > v: 
    #小于当前最小卡方，更新最小值
                minvalue = v
                minidx = i     
#如果最小卡方值小于阈值，则合并最小卡方值的相邻两组，并继续循环  
        if (max_groups is not None and max_groups< len(freq) ) or (threshold is not None and minvalue < threshold):   
        # 
        #minidx后一行合并到minidx
            tmp  = freq[minidx] + freq[minidx+1]
            freq[minidx] = tmp       
        #删除minidx后一行
            freq = np.delete(freq,minidx+1,0)       
        #删除对应的切分点
            cutoffs = np.delete(cutoffs,minidx+1,0)
        else: 
#最小卡方值不小于阈值，停止合并。
            break
    return cutoffs

def value2group(x,cutoffs):
    '''
    将变量的值转换成相应的组。
    x: 需要转换到分组的值
    cutoffs: 各组的起始值。
    return: x对应的组，如group1。从group1开始。
    '''
#切分点从小到大排序。
    cutoffs = sorted(cutoffs)
    num_groups = len(cutoffs)
#异常情况：小于第一组的起始值。这里直接放到第一组。
#异常值建议在分组之前先处理妥善。
    if x < cutoffs[0]:
        return 'group1'
    for i in range(1,num_groups):     
        if cutoffs[i-1] <= x < cutoffs[i]:
            return 'group{}'.format(i)
#最后一组，也可能会包括一些非常大的异常值。
    return 'group{}'.format(num_groups)


# 应用
cutoffs = chiMerge(sample_set,'in7d_rate','target',max_groups=5)
cutoffs
sample_set['total_acc_chi2_group']=sample_set['in7d_rate'].apply(value2group,args=(cutoffs,))
sample_set
b=get_bestsplit_list(sample_set,'in7d_rate')
#%%------------连续型变量最优分箱基于CART算法


import pandas as pd
import numpy as np
os.chdir(r"F:\菲律宾日报\建模\分箱\卡方")
#读取数据集，至少包含变量和target两列
sample_set = pd.read_excel('数据样本.xlsx')

def calc_score_median(sample_set, var):
    '''
    计算相邻评分的中位数，以便进行决策树二元切分
    param sample_set: 待切分样本
    param var: 分割变量名称
    '''
    var_list = list(np.unique(sample_set[var]))
    var_median_list = []
    for i in range(len(var_list) -1):
        var_median = (var_list[i] + var_list[i+1]) / 2
        var_median_list.append(var_median)
    return var_median_list




sample_set.columns



score_median_list = calc_score_median(sample_set, 'in7d_rate')


def choose_best_split(sample_set, var, min_sample):
    '''
    使用CART分类决策树选择最好的样本切分点
    返回切分点
    param sample_set: 待切分样本
    param var: 分割变量名称
    param min_sample: 待切分样本的最小样本量(限制条件)
    '''
    # 根据样本评分计算相邻不同分数的中间值
    score_median_list = calc_score_median(sample_set, var)
    median_len = len(score_median_list)
    sample_cnt = sample_set.shape[0]
    sample1_cnt = sum(sample_set['target'])
    sample0_cnt =  sample_cnt- sample1_cnt
    Gini = 1 - np.square(sample1_cnt / sample_cnt) - np.square(sample0_cnt / sample_cnt)
    
    bestGini = 0.0; bestSplit_point = 0.0; bestSplit_position = 0.0
    for i in range(median_len):
        left = sample_set[sample_set[var] < score_median_list[i]]
        right = sample_set[sample_set[var] > score_median_list[i]]
        
        left_cnt = left.shape[0]; right_cnt = right.shape[0]
        left1_cnt = sum(left['target']); right1_cnt = sum(right['target'])
        left0_cnt =  left_cnt - left1_cnt; right0_cnt =  right_cnt - right1_cnt
        left_ratio = left_cnt / sample_cnt; right_ratio = right_cnt / sample_cnt
        
        if left_cnt < min_sample or right_cnt < min_sample:
            continue
        
        Gini_left = 1 - np.square(left1_cnt / left_cnt) - np.square(left0_cnt / left_cnt)
        Gini_right = 1 - np.square(right1_cnt / right_cnt) - np.square(right0_cnt / right_cnt)
        Gini_temp = Gini - (left_ratio * Gini_left + right_ratio * Gini_right)
        if Gini_temp > bestGini:
            bestGini = Gini_temp; bestSplit_point = score_median_list[i]
            if median_len > 1:
                bestSplit_position = i / (median_len - 1)
            else:
                bestSplit_position = i / median_len
        else:
            continue
               
    Gini = Gini - bestGini
    return bestSplit_point, bestSplit_position




def bining_data_split(sample_set, var, min_sample, split_list):
    '''
    划分数据找到最优分割点list
    param sample_set: 待切分样本
    param var: 分割变量名称
    param min_sample: 待切分样本的最小样本量(限制条件)
    param split_list: 最优分割点list
    '''
    split, position = choose_best_split(sample_set, var, min_sample)
    if split != 0.0:
        split_list.append(split)
    # 根据分割点划分数据集，继续进行划分
    sample_set_left = sample_set[sample_set[var] < split]
    sample_set_right = sample_set[sample_set[var] > split]
    # 如果左子树样本量超过2倍最小样本量，且分割点不是第一个分割点，则切分左子树
    if len(sample_set_left) >= min_sample * 2 and position not in [0.0, 1.0]:
        bining_data_split(sample_set_left, var, min_sample, split_list)
    else:
        None
    # 如果右子树样本量超过2倍最小样本量，且分割点不是最后一个分割点，则切分右子树
    if len(sample_set_right) >= min_sample * 2 and position not in [0.0, 1.0]:
        bining_data_split(sample_set_right, var, min_sample, split_list)
    else:
        None
def get_bestsplit_list(sample_set, var):
    '''
    根据分箱得到最优分割点list
    param sample_set: 待切分样本
    param var: 分割变量名称
    '''
    # 计算最小样本阈值（终止条件）
    min_df = sample_set.shape[0] * 0.05
    split_list = []
    # 计算第一个和最后一个分割点
    bining_data_split(sample_set, var, min_df, split_list)
    return split_list
b=get_bestsplit_list(sample_set,'in7d_rate')


#%%------------技巧

"""
t0=time.perf_counter()
t1=time.perf_counter()
print('时间为：%.3f'%(t1-t0))
 这一套可以看到运行时间
"""


df.rename(index={'一':'one'},columns={'col1':'new_col1'}) #原来rename可以修改索引，怪不得平时修改列名的时候一定要参数columns
df.values #由很多行值(本身是list) 构成的大list
df.dtypes #查看每列的类型
df.shape  #返回行列
df.mean() #返回各个列名下的均值，估计其他聚合函数也能使用
# 索引对齐特性:两个df进行数学运算时会按照索引来计算，而不是物理层面上的A表第一行加B表第一行
df.drop(index='五',columns='col1')  #原来drop可以修改索引，怪不得平时删除列名的时候一定要参数columns
df.select_dtypes(include=['number']).head()  #这个技巧是好东西
s.to_frame()  #将s这个series转为dataframe
df['Physics'].nunique() #nunique显示有多少个唯一值
df['Physics'].unique() #unique显示所有的唯一值
df['Physics'].count()  #count返回非缺失值元素个数 
df['Physics'].value_counts() #value_counts返回每个元素有多少个
df.info()   #info函数返回有哪些列、有多少非缺失值、每列的类型
df.describe()  #describe默认统计数值型数据的各个统计量
df.describe(percentiles=[.05, .25, .75, .95])  #可以自行选择分位数
df['Physics'].describe()  #对于非数值型也可以用describe函数

df.loc #里面如果涉及到多个条件：应该使用&而不是and在每个条件周围加上括号
df.loc[df['Address'].isin(['street_7','street_4'])].head()  #布尔索引，本质上说，loc中能传入的只有布尔列表和索引子集构成的列表
df.loc[[True if i[-1]=='4' or i[-1]=='7' else False for i in df['Address'].values]].head() #花里胡哨的布尔索引
df.iloc[(df['School']=='S_1').values].head() #iloc中接收的参数只能为整数或整数列表或布尔列表，不能使用布尔Series，如果要用就必须如下把values拿出来
#布尔符号：'&','|','~'：分别代表和and，或or，取反not
df[df['Address'].isin(['street_1','street_4'])&df['Physics'].isin(['A','A+'])]#isin
df_using_mul.sort_index().loc['C_2','street_5'] #多层索引要先排序才能用切片
df_using_mul.sort_index().loc[('C_2','street_6'):('C_3','street_4')] #多层索引要先排序才能用切片
df_using_mul.sort_index().loc[('C_2','street_7'):'C_3'].head() #多层索引要先排序才能用切片
df_using_mul.sort_index().loc[[('C_2','street_7'),('C_3','street_2')]] #表示选出某几个元素，精确到最内层索引
df_using_mul.sort_index().loc[(['C_2','C_3'],['street_4','street_7']),:] ##选出第一层在‘C_2’和'C_3'中且第二层在'street_4'和'street_7'中的行
# 多层索引中的slice对象??  xs  idx
df.reindex(index=[1101,1203,1206,2402])  #reindex是指重新索引，它的重要特性在于索引对齐，很多时候用于重新排序,这里排序了行索引
df.reindex(columns=['Height','Gender','Average']).head()#reindex是指重新索引，它的重要特性在于索引对齐，很多时候用于重新排序，这里重新排序了列
# ping -t 192.168.1.1  结束的时候用Ctrl+C
df.reindex(index=[1101,1203,1206,2402],method='bfill')#df必须索引经过排序，否则报错，bfill表示用所在索引1206的后一个有效行填充，ffill为前一个有效行，nearest是指最近的
df.set_index('Class',append=True).head() #利用append参数可以将当前索引维持不变
df.set_index(pd.Series(range(df.shape[0]))).head()#当使用与表长相同的列作为索引（需要先转化为Series，否则报错）
df.set_index([pd.Series(range(df.shape[0])),pd.Series(np.ones(df.shape[0]))]).head()#可以直接添加多级索引
df_temp.rename_axis(index={'Lower':'LowerLower'},columns={'Big':'BigBig'})#rename_axis是针对多级索引的方法，作用是修改某一层的索引名，而不是索引标签，索引名是不常用的，比如set_index(class),这个就是行索引名，但列索引名可能在多重索引里才会出现吧
df_temp.rename(index={'A':'T'},columns={'e':'changed_e'}).head()#rename方法用于修改列或者行索引标签，而不是索引名,索引标签是列名或者行索引的值
f.where(df['Gender']=='M').dropna().head() #当对条件为False的单元进行填充,默认填充NaN,通过这种方法筛选结果和[]操作符的结果完全一致
df.mask(df['Gender']=='M').dropna().head() #mask函数与where功能上相反，其余完全一致，即对条件为True的单元进行填充
df.query('(Address in ["street_6","street_7"])&(Weight>(70+10))&(ID in [1303,2304,2402])') #query函数中的布尔表达式中，下面的符号都是合法的：行列索引名、字符串、and/not/or/&/|/~/not in/in/==/!=、四则运算符
df.sample(frac=0.05) #frac为抽样比
df.sample(n=3,axis=1).head() #抽行
print([attr for attr in dir(grouped_single) if not attr.startswith('_')]) #好东西
grouped_single.head(2) #对分组对象使用head函数，返回的是每个组的前几行，而不是数据集前几行
df.groupby(np.random.choice(['a','b','c'],df.shape[0])).get_group('a').head()#对于groupby函数而言，分组的依据是非常自由的，只要是与数据框长度相同的列表即可，同时支持函数型分组

mean/sum/size/count/std/var/sem/describe/first/last/nth/min/max  #常用聚合函数
group_m.agg([('rename_sum','sum'),('rename_mean','mean')]) #利用元组进行重命名
grouped_mul.agg({'Math':['mean','max'],'Height':'var'}) #指定哪些函数作用哪些列
pd.pivot_table(df,index='School',columns='Gender',values='Height',aggfunc=['mean','sum'],margins=True).head()#margins是汇总的功能
loan.columns.set_levels(['合同量','合同金额'],level=0,inplace=True)#重命名第一层（最上面那一层)索引

df_append.append(pd.Series({'Gender':'F','Height':188},name='new_row')) 
df_append.append(pd.DataFrame({'Gender':['F','M'],'Height':[188,176]},index=['new_1','new_2']))
df_append.assign(col1=lambda x:x['Gender']*2, col2=s) #该方法主要用于添加列，列名直接由参数指定
df['Physics'].isna().head() #对Series使用会返回布尔列表
df['Physics'].notna().head()
df.isna().head() #对DataFrame使用会返回布尔表
df.isna().sum() #对于DataFrame我们更关心到底每列有多少缺失值,也可以用info函数查看缺失信息
df[df['Physics'].isna()]  #挑出该列缺失值的行
df[df.notna().all(1)] #使用all就是全部非缺失值，如果是any就是至少有一个不是缺失值
    # np.nan是一个麻烦的东西，首先它不等与任何东西，甚至不等于自己，在用equals函数比较时，自动略过两侧全是np.nan的单元格，因此结果不会影响
    # 它在numpy中的类型为浮点，由此导致数据集读入时，即使原来是整数的列，只要有缺失值就会变为浮点型
    # 布尔类型的列表，如果是np.nan填充，那么它的值会自动变为True而不是False，但当修改一个布尔列表时，会改变列表类型，而不是赋值为True
    # 在所有的表格读取后，无论列是存放什么类型的数据，默认的缺失值全为np.nan类型
    # 因此整型列转为浮点；而字符由于无法转化为浮点，因此只能归并为object类型（'O'），原来是浮点型的则类型不变
        # None比前者稍微好些，至少它会等于自身，它的布尔值为False，修改布尔列表不会改变数据类型，在传入数值类型后，会自动变为np.nan
        # 只有当传入object类型是保持不动，几乎可以认为，除非人工命名None，它基本不会自动出现在Pandas中
        # 在使用equals函数时不会被略过，因此下面的情况下返回False

# NaT是针对时间序列的缺失值，是Pandas的内置类型，可以完全看做时序版本的np.nan，与自己不等，且使用equals是也会被跳过

s_new = pd.Series([1, 2], dtype="Int64")#Nullable整形，它与原来标记int上的符号区别在于首字母大写：'Int'，它的好处就在于，其中前面提到的三种缺失值都会被替换为统一的NA符号，且不改变数据类型
s_new = pd.Series([0, 1], dtype="boolean") #作用与上面的类似
s = pd.Series(['dog','cat'],dtype='string')#它本质上也属于Nullable类型，因为并不会因为含有缺失而改变类型，此外，和object类型的一点重要区别就在于，在调用字符方法后，string类型返回的是Nullable类型，object则会根据缺失类型和数据类型而改变
pd.read_csv('data/table_missing.csv').convert_dtypes().dtypes #这个函数的功能往往就是在读取数据时，就把数据列转为Nullable类型，是1.0的新函数

# 使用加法时，缺失值为0；使用乘法时，缺失值为1；使用累计函数时，缺失值自动略过；groupby方法自动忽略为缺失值的组
df_d.dropna(axis=1,how='all') #how参数（可以选all或者any，表示全为缺失去除和存在缺失去除）
df_d.dropna(axis=0,subset=['B','C']) #subset参数（即在某一组列范围中搜索缺失值）
month_loan = repay_st.groupby(['loan_month','loan_type']).agg({'contract_amount':[np.size,np.sum],
                                                              'pay_amount':np.sum}).unstack()#unstack讲行索引转到列索引
round_r = settle_apply.loc[settle_apply.number_of_loans==r,:].groupby('first_loan_month')['next_apply'].agg(clear=np.size,next_apply=np.sum)

string类型和object:
    字符存取方法（string accessor methods，如str.count）会返回相应数据的Nullable类型，而object会随缺失值的存在而改变返回类型
    某些Series方法不能在string上使用，例如： Series.str.decode()，因为存储的是字符串而不是字节
    string类型在缺失值存储或运算时，类型会广播为pd.NA，而不是浮点型np.nan

pd.Series([1,'1.']).astype('str').astype('string') #将一个其他类型的容器直接转换string类型,分两部转换，先转为str型object，在转为string类型
pd.Series(['a_b_c', 'c_d_e', np.nan, 'f_g_h'], dtype="string").str.split('_') #注意split后的类型是object，因为现在Series中的元素已经不是string，而包含了list，且string类型只能含有字符串
s.str.split('_').str[1] #str方法可以进行元素的选择，如果该单元格元素是列表，那么str[i]表示取出第i个元素
pd.Series(['a_b_c', ['a','b','c']], dtype="object").str[1] #如果是单个元素，则先把元素转为列表在取出,第一个元素先转为['a','_','b','_','c']
s.str.split('_',expand=True,n=1) #expand参数控制了是否将列拆开，n参数代表最多分割多少次
pd.Series(['ab',None,'d'],dtype='string').str.cat(sep=',',na_rep='*') #对于单个Series而言，就是指所有的元素进行字符合并为一个字符串,可选sep分隔符参数，和缺失值替代字符na_rep参数
s.str.cat(pd.Series(['24',None,None],dtype='string'),,sep=',',na_rep='*') #对于两个Series合并而言，是对应索引的元素进行合并,需要注意的是两个缺失值会被同时替换
pd.Series(list('abc'),index=[' space1  ','space2  ','  space3'],dtype="string").index.str.strip() #常用于过滤空格 
pd.Series('A',dtype="string").str.lower() 
pd.Series('abCD',dtype="string").str.swapcase()#交换字母大小写
pd.Series('abCD',dtype="string").str.capitalize() #大写首字母
pd.Series(['1.2','1','-0.3','a',np.nan],dtype="string").str.isnumeric() #检查每一位是否都是数字


分类变量的创建:元素值（values）、分类类别（categories）、是否有序（order:s.cat.ordered）值排序和索引排序都是适用的；分类变量的比较
    pd.Series(["a", "b", "c", "a"], dtype="category")
    temp_df = pd.DataFrame({'A':pd.Series(["a", "b", "c", "a"], dtype="category"),'B':list('abcd')})
    cat = pd.Categorical(["a", "b", "c", "a"], categories=['a','b','c']) &&&&  pd.Series(cat)  
    pd.cut(np.random.randint(0,60,5), [0,10,30,60],right=False, labels=['0-10','10-30','30-60'])
        s.cat.set_categories(['new_a','c']) #修改分类，但本身值不会变化  
        s.cat.rename_categories(['new_%s'%i for i in s.cat.categories]) #把值和分类同时修改
        s.cat.rename_categories({'a':'new_a','b':'new_b'})
        s.cat.add_categories(['e']) #添加
        s.cat.remove_categories(['d'])#移除
        pd.Series(["a", "d", "c", "a"]).astype('category').cat.as_ordered()#将一个序列转为有序变量
        s.cat.as_unordered() #退化为无序变量
        pd.Series(["a", "d", "c", "a"]).astype('category').cat.set_categories(['a','c','d'],ordered=True) #利用set_categories方法中的order参数

名称	描述	元素类型	创建方式
① Date times（时间点/时刻）	描述特定日期或时间点	Timestamp	to_datetime或date_range
② Time spans（时间段/时期）	由时间点定义的一段时期	Period	Period或period_range
③ Date offsets（相对时间差）	一段时间的相对大小（与夏/冬令时无关）	DateOffset	DateOffset
④ Time deltas（绝对时间差）	一段时间的绝对大小（与夏/冬令时有关）	Timedelta	to_timedelta或timedelta_range

pd.to_datetime('2020\\1\\1',format='%Y\\%m\\%d') #利用format参数强制匹配
pd.to_datetime(pd.DataFrame({'year': [2020, 2020],'month': [1, 1], 'day': [1, 2]}))#对于DataFrame而言，如果列已经按照时间顺序排好，则利用to_datetime可自动转换
pd.Timestamp.min  pd.Timestamp.max #Timestamp('1677-09-21 00:12:43.145225')  Timestamp('2262-04-11 23:47:16.854775807')
ts['2020-7'].head() #子集索引
ts['2011-1':'20200726'].head()#支持混合形态索引
pd.Series(ts.index).dt.strftime('%Y-间隔1-%m-间隔2-%d').head()#利用strftime可重新修改时间格式

g_profit = pd.DataFrame(data={'start_date':[st],
                                  'end_date':[et],
#                                   'first_pay_amount':rev_exp.loc[rev_exp.loan_type1=='first_apply','pay_amount'].values,
#                                   'first_profit':rev_exp.loc[rev_exp.loan_type1=='first_apply','in7d毛利率'].values,
#                                   'first_到期回收率':rev_exp.loc[rev_exp.loan_type1=='first_apply','到期回收率'].values,
#                                   'first_7天内催回率':rev_exp.loc[rev_exp.loan_type1=='first_apply','7天内催回率'].values,
                                  'ext_first_pay_amount':rev_exp.loc[rev_exp.loan_type1=='ext_first_apply','pay_amount'].values,
                                  'ext_first_profit':rev_exp.loc[rev_exp.loan_type1=='ext_first_apply','in7d毛利率'].values,
                                  'ext_first_到期回收率':rev_exp.loc[rev_exp.loan_type1=='ext_first_apply','到期回收率'].values,
                                  'ext_first_7天内催回率':rev_exp.loc[rev_exp.loan_type1=='ext_first_apply','7天内催回率'].values,
                                  're_pay_amount':rev_exp.loc[rev_exp.loan_type1=='re_apply','pay_amount'].values,
                                  're_profit':rev_exp.loc[rev_exp.loan_type1=='re_apply','in7d毛利率'].values,
                                  're_到期回收率':rev_exp.loc[rev_exp.loan_type1=='re_apply','到期回收率'].values,
                                  're_7天内催回率':rev_exp.loc[rev_exp.loan_type1=='re_apply','7天内催回率'].values,
                                 },
                            index=[cut_date]
                           )



df_result = pd.DataFrame(
    [
        ['A', 0],
        ['A', 1],
        ['A', 1],
        ['B', 1],
        ['B', 0],
        ['A', 0],
        ['A', 1],
        ['B', 1],
        ['B', 0]
    ], columns = ['key', 'cond'])

.astype(int)
df_result['new'] = df_result.groupby('key').cond.apply(
        lambda x: x.shift()
)
df_result['new1'] = df_result.groupby('key').cond.apply(
        lambda x: x.shift().fillna(1).cumsum()
)

# 独立性卡方检验
import scipy
from scipy.stats import chi2_contingency
import numpy as np
import pandas as pd
data=[[25,21,10],[82,88,30],[223,16,5]]
df=pd.DataFrame(data,index=['美式咖啡','拿铁咖啡','卡布奇诺'],columns=['IT','行政','工程'])
kt=chi2_contingency(df)
print('卡方值=%.4f,p值=%.4f,自由度=%i expected_frep=%s' %kt)
# 一致性卡方检验
from scipy import stats 
df=[7.68,6.63,7.16,5.98,6.34,
    6.92,8.27,8.19,8.6,6.82,
    7.52,7.96,9.23,6.04,8.08,
    6.4,5.63,5.8,6.18,7.4,
    8.52,6.34,5.62,7.29,7.33,
    6.44,6.03,8.18,8.32,8.52,
    6.24,6.23,8.2,7.31,8.5,
    7.51,6.01,7.25,7.44,7.86]
ks_test = stats.kstest(df,'norm')
shapiron_test = stats.shapiro(df)
print('ks_test: ',ks_test)
print('shapiro_test: ',shapiron_test)



#%%------repay_mart&appr_mart
import json
import numpy as np
import pandas as pd
import sys
sys.path.append(r"C:\Users\lenovo\Anaconda3\Lib\site-packages")
import pymysql
import math
import datetime
import gc
#gc.collect() 垃圾回收，返回处理这些循环引用一共释放掉的对象个数
import os
import oss2
import zipfile
from itertools import islice
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = [u'SimHei']
plt.rcParams['axes.unicode_minus'] = False


#根据数据库连接信息文件所在路径做修改
try:
    os.chdir(r"E:\guan\菲律宾")
except:
    os.chdir(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart")

with open("oss_access_key.json") as oss_access_key:
    oss_args = json.load(oss_access_key)
with open(r"db_config.json") as db_config:
    cnx_args = json.load(db_config)

#从oss下载数据
os.chdir(r"E:\guan\原数据\flb_csv")
auth = oss2.Auth(oss_args['access_key_id'], oss_args['access_key_secret'])
bucket = oss2.Bucket(auth, oss_args['endpoint'], 'spark99-prd')
#for b in islice(oss2.ObjectIterator(bucket), 10):
#    print(b.key)
#bucket.get_object_to_file("suncash-2020-03-23-dbdata-full.zip","suncash-2020-03-23-dbdata-full.zip")    
if datetime.date.today().isoweekday()==1:
    for i in range(1,4): #周一下载周五、六、日三天数据
        datestr = (datetime.date.today() - datetime.timedelta(days=i)).strftime('%Y-%m-%d')
        key = "suncash-" + datestr + "-dbdata.zip"
        bucket.get_object_to_file(key,key)
        zipfile.ZipFile(key).extractall()
else:    
    yesterday = datetime.date.today() - datetime.timedelta(days=1)
    key = "suncash-" + yesterday.strftime('%Y-%m-%d') + "-dbdata.zip"
    bucket.get_object_to_file(key,key)
    zipfile.ZipFile(key).extractall()
#for i in range(3):
#    datestr = (datetime.date(2020,5,22) + datetime.timedelta(days=i)).strftime('%Y-%m-%d')
#    key = "suncash-" + datestr + "-dbdata.zip"
#    bucket.get_object_to_file(key,key)
#    zipfile.ZipFile(key).extractall()
    
    
dtype ={'borrower_tel_one':str,'phone_no':str,'payment_type':str,
        'ad_id':str,'adset_id':str,'campaign_id':str,'status':str}    

importfiles = [
#'approval.apply_info.csv',
#'approval.auto_audit_result.csv',
#'approval.manual_audit_result.csv',
#'approval.borrower_info.csv',
#'approval.employment_info.csv',
#'approval.loan_info.csv',
#'approval.contract_info.csv',
'account.account_info.csv',
#'account.bill_main.csv',
#'account.trans_journal_detail.csv',
#'account.realtime_draw.csv',
#'suncash_lend.user.csv',
#'suncash_lend.af_event_info_detail.csv',
#'suncash_lend.user_verification_info.csv',
#'suncash_lend.apply_info.csv',
#'suncash_lend.user_bind_recommend_info.csv'
'collection.case_flow_info.csv',
#'collection.collection_log_info.csv',
'collection.my_case_info.csv'
]

#增量数据的导入与合并
#for i in range(8):
#    sub_dir = (datetime.date(2020,5,10) + datetime.timedelta(days=i)).strftime('%Y-%m-%d')
for (i,d) in enumerate(pd.date_range(datetime.date(2020,5,10),
                                     datetime.date.today(),
                                     closed='left')):
    sub_dir = d.strftime('%Y-%m-%d')
    os.chdir(os.path.join(r"E:\guan\原数据\flb_csv",sub_dir))
    for j in importfiles:
        tb = j.split(".")[1]
        try:
            if i == 0:
                locals()[tb] = pd.read_csv(j,dtype=dtype)
#                locals()[tb] = pd.read_csv(j,dtype=str)
            else:
                temp = pd.read_csv(j,dtype=dtype)
#                temp = pd.read_csv(j,dtype=str)
                locals()[tb] = pd.concat([locals()[tb],temp])
        except FileNotFoundError:
            print("file {0} not found in {1}".format(j,sub_dir))
        except Exception as e:
            print("Error:",e)          

#全量+增量数据的导入与合并
#从某一天的增量数据获取表头
os.chdir(r"E:\guan\原数据\flb_csv\2020-07-27")
for f in importfiles:
    locals()[f.split(".")[1]+'_cols'] = pd.read_csv(f,dtype=dtype).columns
#导入全量数据
os.chdir(r"E:\guan\原数据\flb_csv\2020-08-01-full")
for f in importfiles:
    col_names = locals()[f.split(".")[1]+'_cols']
    locals()[f.split(".")[1]] = pd.read_csv(f,dtype=dtype,names=col_names,low_memory=False)
#导入增量数据并与全量合并
for (i,d) in enumerate(pd.date_range(datetime.date(2020,8,1),
                                     datetime.date.today(),
                                     closed='left')):
    sub_dir = d.strftime('%Y-%m-%d')
    os.chdir(os.path.join(r"E:\guan\原数据\flb_csv",sub_dir))
    for j in importfiles:
        tb = j.split(".")[1]
        try:
            temp = pd.read_csv(j,dtype=dtype,low_memory=False)
            locals()[tb] = pd.concat([locals()[tb],temp],sort=False)
        except FileNotFoundError:
            print("file {0} not found in {1}".format(j,sub_dir))
        except Exception as e:
            print("Error:",e)      

#去重
os.chdir(r"E:\guan\原数据\flb_csv\处理好")           
for f in importfiles:
    tb = f.split('.')[1]
    try:
        locals()[tb].columns = locals()[tb].columns.str.lower()
        locals()[tb].sort_values(by=['id','updated_time'],inplace=True)
        locals()[tb].drop_duplicates(subset='id',keep='last',inplace=True)
        locals()[tb].to_csv(f,index=False)
    except Exception as e:
        print("Error:",e)
                
            

#给oss下载的数据加上表头。后面的数据已经有表头，可以不需要了
cnx = pymysql.connect(**cnx_args)
#filelist = os.listdir(r"E:\guan\原数据\flb_csv")
#filelist = os.listdir()
filelist = [
 'account.account_info.csv',
 'account.bill_fee_dtl.csv',
 'account.bill_main.csv',
 'account.offset_info.csv',
 'account.realtime_draw.csv',
 'account.trans_journal_detail.csv',
 'account.trans_journal_main.csv',
 'approval.act_hi_procinst.csv',
 'approval.act_hi_taskinst.csv',
 'approval.apply_info.csv',
 'approval.auto_audit_result.csv',
# 'approval.base_derived_variable.csv',
 'approval.borrower_info.csv',
 'approval.contract_info.csv',
# 'approval.device_info_derived_variable.csv',
 'approval.employment_info.csv',
 'approval.loan_info.csv',
 'approval.manual_audit_result.csv',
 'collection.case_flow_info.csv',
 'collection.collection_log_info.csv',
 'collection.my_case_info.csv',
# 'suncash_data.td_live_image_compare_info.csv',
# 'suncash_data.td_task_info.csv',
 'suncash_lend.af_event_info_detail.csv',
 'suncash_lend.apply_info.csv',
# 'suncash_lend.geo_point_distance.csv',
 'suncash_lend.user_bind_recommend_info.csv',
 'suncash_lend.user_recommend_code_info.csv']
for f in filelist:
    tb = os.path.splitext(f)[0]
    temp = pd.read_sql("select * from "+tb+" limit 10",cnx)
    locals()[tb.split(".")[1]] = pd.read_csv(f,names=temp.columns,dtype=dtype)
#    locals()[tb.split(".")[1]].to_excel(r"处理好\\" + tb.split(".")[1] + ".xlsx",index=False)
    locals()[tb.split(".")[1]].to_csv(r"处理好\\" + f,index=False)
    print("dealt with %s" % tb)

#处理好的数据+新增
#f = 'account.account_info.csv'
#filelist = [
# 'account.account_info.csv',
# 'account.bill_main.csv',
# 'approval.apply_info.csv',
# 'approval.borrower_info.csv',
# 'approval.contract_info.csv',
# 'approval.employment_info.csv',
# 'approval.loan_info.csv'
#]
filelist = [
 'account.account_info.csv',
 'account.bill_main.csv',
 'approval.apply_info.csv',
 'approval.borrower_info.csv',
 'approval.contract_info.csv',
 'approval.employment_info.csv',
 'approval.loan_info.csv',
 'account.bill_fee_dtl.csv',
 'account.offset_info.csv',
 'account.realtime_draw.csv',
 'account.trans_journal_detail.csv',
 'account.trans_journal_main.csv',
# 'approval.act_hi_procinst.csv',
# 'approval.act_hi_taskinst.csv',
 'approval.auto_audit_result.csv',
 'approval.base_derived_variable.csv',
 'approval.device_info_derived_variable.csv',
 'approval.manual_audit_result.csv',
 'collection.case_flow_info.csv',
 'collection.collection_log_info.csv',
 'collection.my_case_info.csv',
 'suncash_data.td_live_image_compare_info.csv',
 'suncash_data.td_task_info.csv',
 'suncash_lend.af_event_info_detail.csv',
 'suncash_lend.apply_info.csv',
 'suncash_lend.geo_point_distance.csv',
 'suncash_lend.user_bind_recommend_info.csv',
 'suncash_lend.user_recommend_code_info.csv']

for f in filelist:
    base = pd.read_csv(r"处理好\\" + f,dtype=dtype)
    base.columns = base.columns.str.lower()
    try:
#        updated = pd.read_csv(f,names=base.columns,dtype=dtype)
        updated = pd.read_csv(f,dtype=dtype)
        updated.columns = updated.columns.str.lower()
        new = pd.concat([base,updated])
        new.sort_values(by=['id','updated_time'],inplace=True)
        new.drop_duplicates(subset='id',keep='last',inplace=True)
        new.to_csv(r"处理好\\" + f,index=False)
        print("dealt with %s" % f)   
    except FileNotFoundError as e:
        print("except:",e)
    except Exception as e:
        print("Error:",e)   
     

#导入处理好的数据
os.chdir(r"E:\guan\原数据\flb_csv\处理好")
importfiles = [
 'account.account_info.csv',
 'account.bill_main.csv',
 'approval.apply_info.csv',
 'approval.borrower_info.csv',
 'approval.contract_info.csv',
 'approval.employment_info.csv',
 'approval.loan_info.csv'
]
for f in importfiles:
    try:
        locals()[f.split('.')[1]] = pd.read_csv(f,dtype=dtype)
        print("Import %s success" % f)
    except Exception as e:
        print("Error:",e)


#本地导入数据的处理
apply_info_cols = ['apply_code','user_code','apply_loan_amount','apply_time',
                   'duration','loan_type','case_state','usage_of_loan_text',
                   'number_of_loans','approval_status','customer_source_sys',
                   'pay_type_name']
apply_info = apply_info[apply_info_cols]
apply_info['apply_time'] = pd.to_datetime(apply_info['apply_time'])

#申请序数
number_of_apply = apply_info.loc[:,['apply_code','user_code','apply_time']]
number_of_apply['number_of_apply'] = number_of_apply['apply_time'].groupby(number_of_apply['user_code']).rank(method='dense')
number_of_apply.sort_values(['user_code','apply_time'],inplace=True)
number_of_apply.drop(columns=['user_code','apply_time'],inplace=True)

loan_info_cols = ['apply_code','contract_no','status','pay_channel',
                  'pay_time','payable_store_name']
loan_status = loan_info.loc[:,loan_info_cols]
loan_status['pay_day'] = pd.to_datetime(loan_status.pay_time).dt.strftime('%Y%m%d')
loan_status['pay_month'] = pd.to_datetime(loan_status.pay_time).dt.strftime('%Y%m')
loan_status.rename(columns={'status':'loan_status'},inplace=True)
loan_status['loan_status'].replace({'6':'放款成功'},inplace=True)

service_fee_rate = contract_info.loc[:,['apply_code', 'service_fee_rate']]

auto_audit_cols = ['apply_code','refuse_info_2']
auto_audit_result = auto_audit_result.loc[:,auto_audit_cols]
auto_audit_result.rename(columns={'refuse_info_2':'auto_refuse_reason'},inplace=True)

manual_audit_cols = ['apply_code','refuse_info_1','refuse_info_2','cancle_info_1']
manual_audit_result = manual_audit_result.loc[:,manual_audit_cols]

account_info_cols = ['contract_no', 'account_status', 'loan_date',
                     'clear_date', 'loan_term', 'last_repay_date',
                     'contract_amount', 'loan_amount', 'borrower_tel_one',
                     'loan_type','updated_time','customer_type']
account_info.columns = account_info.columns.str.lower()
account_info = account_info[account_info_cols]
account_info.rename(columns={'loan_type':'loan_channel'},inplace=True)
account_info.sort_values(by=['contract_no','updated_time'],inplace=True)
account_info.drop_duplicates(subset='contract_no',keep='last',inplace=True)
account_info['loan_date'] = pd.to_datetime(account_info['loan_date'])
account_info['loan_month'] = account_info['loan_date'].dt.strftime('%Y%m')
account_info['loan_day'] = account_info['loan_date'].dt.strftime('%Y%m%d')   
account_info['last_repay_month'] = pd.to_datetime(account_info['last_repay_date']).dt.strftime('%Y%m')
account_info['last_repay_date'] = pd.to_datetime(account_info['last_repay_date']).dt.date
account_info['last_repay_week'] = pd.to_datetime(account_info['last_repay_date']).dt.strftime('%W')
account_info['clear_date'] = pd.to_datetime(account_info['clear_date']).dt.date
account_info['loan_month'] = account_info['loan_date'].dt.strftime('%Y%m')
account_info['loan_day'] = account_info['loan_date'].dt.strftime('%Y%m%d')
account_info['extend_date'] = pd.to_datetime(account_info['updated_time']).dt.date
account_info = account_info.query("contract_no not in ('C202008072054334758664')") #剔除取消展期的合同

bill_main.columns = bill_main.columns.str.lower()
bill_main.sort_values(by=['contract_no','updated_time'],inplace=True)
bill_main.drop_duplicates(subset='contract_no',keep='last',inplace=True)
bill_main['od_days'] = bill_main.apply(lambda x: 0 if x.bill_status=='BILL_CLEAR' else x.overdue_days,axis=1)
od_days_ever = bill_main[['contract_no','overdue_days','od_days']].rename(columns={'overdue_days':'od_days_ever'})

borrower_info_cols = ['apply_code', 'id_no', 'id_type', 'age',
                      'children_text', 'education', 'gender', 'marriage',
                      'province_addr_text', 'city_addr_text', 'district_addr_text',
                      'length_of_residence', 'device_id', 'phone_no','birthday'
                      ]
borrower_info = borrower_info[borrower_info_cols]
borrower_info.rename(columns={'province_addr_text':'home_province',
                              'city_addr_text':'home_city',
                              'district_addr_text':'home_district'},
                     inplace=True)
borrower_info['age_grp'] = pd.cut(borrower_info.age,
                                 bins=[-np.inf,21,28,35,49,np.inf],
                                 labels=['  - 21','22 - 28','29 - 35','36 - 49','50 -  '])

employment_info_cols = ['apply_code', 'job_type', 'monthly_net_income_text',
                        'on_the_job_time', 'province_addr_text',
                        'city_addr_text','district_addr_text'
                        ]
employment_info = employment_info[employment_info_cols]
employment_info.rename(columns={'province_addr_text':'job_province',
                                'city_addr_text':'job_city',
                                'district_addr_text':'job_district'},
                       inplace=True)


#生成数据集appr_mart & repay_mart-----------------------------------------------------------
try:
    os.chdir(r"E:\guan\菲律宾")
except:
    os.chdir(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart")

apply_info['apply_month'] = apply_info['apply_time'].dt.strftime('%Y%m')
apply_info['apply_day'] = apply_info['apply_time'].dt.strftime('%Y%m%d')
 
loan_status = pd.merge(loan_status,service_fee_rate,how='left',on='apply_code')

apply_appr_loan_info = pd.merge(apply_info,loan_status,how='left',on='apply_code')
apply_appr_loan_info['approval_status'] = apply_appr_loan_info['approval_status'].astype('category')
apply_appr_loan_info['approval_status'].cat.set_categories(
                                                    ['AUTO_IN_REVIEW',
                                                    'AUTO_REJECTED',
                                                    'NO_MANUAL_REVIEW',
                                                    'MANUAL_REJECTED',
                                                    'MANUAL_APPROVED',
                                                    'MANUAL_CANCEL',
                                                    'MANUAL_IN_REVIEW'],inplace=True)
apply_appr_loan_info['loan_status'] = apply_appr_loan_info['loan_status'].astype('category')
apply_appr_loan_info['loan_status'].cat.set_categories(['放款成功',5,7,12],inplace=True)

user_apply_code = apply_info.loc[:,['apply_code','user_code','loan_type','number_of_loans','customer_source_sys']]
apply_contract_code = loan_status.loc[:,['apply_code','contract_no','service_fee_rate']]
code = pd.merge(user_apply_code,apply_contract_code,on='apply_code')
code.rename(columns={'contract_no':'source_contract_no'},inplace=True)
ext_info = pd.read_sql("select source_contract_no, act_contract_no \
                       from account.extension_info \
                       where extension_status='SUCCESS'",cnx)
ext_info['展期后的合同'] = 1
account_info = pd.merge(account_info,ext_info,how='left',
                        left_on='contract_no',right_on='act_contract_no')
account_info['source_contract_no'] = account_info.apply(lambda x: x.contract_no if pd.isnull(x.source_contract_no) else x.source_contract_no,axis=1)
account_info = pd.merge(account_info,code,how='left',on='source_contract_no')

#缺失clear_date的处理
# account_info.loc[(account_info.account_status=='ACCOUNT_SETTLE') & (account_info.clear_date.isnull()),:]
account_info.loc[account_info.contract_no=='C201906120207382040078','clear_date'] = datetime.date(2019,6,28)
account_info.loc[account_info.contract_no=='C201906261851549490077','clear_date'] = datetime.date(2019,7,11)
#修正number_of_loans有误的合同
account_info.loc[account_info.apply_code=='PL201905271454230580090','number_of_loans'] = 3
account_info.loc[account_info.apply_code=='PL201905291324034000015','number_of_loans'] = 3
account_info.loc[account_info.apply_code=='PL201905271244097550033','number_of_loans'] = 3

account_info = pd.merge(account_info,od_days_ever,how='left',on='contract_no')

first_loan = account_info.loc[account_info.loan_type=='first_apply',
                              ['user_code','loan_month','loan_day','service_fee_rate']]
first_loan.sort_values(by=['user_code','loan_day'],inplace=True)
first_loan.drop_duplicates(subset='user_code',inplace=True)
first_loan.rename(columns={'loan_month':'first_loan_month',
                           'loan_day':'first_loan_day',
                           'service_fee_rate':'first_service_fee_rate'},
                            inplace=True)
account_info = pd.merge(account_info,first_loan,how='left',on='user_code')

last_loan = account_info.loc[:,['user_code','contract_no','last_repay_date']]
last_loan = last_loan.sort_values(by=['user_code','last_repay_date'])
last_loan.drop_duplicates(subset='user_code',keep='last',inplace=True)
last_loan.drop(columns=['user_code','last_repay_date'],inplace=True)
last_loan['last_loan'] = 'Y'
account_info = pd.merge(account_info,last_loan,how='left',on='contract_no')

account_info['自然逾期'] = account_info.od_days_ever.apply(lambda x: 1 if x>0 else 0)
account_info['到期'] = account_info.last_repay_date.apply(lambda x: 1 if (datetime.date.today()-x).days>0 else 0)
account_info['曾经逾期7天以上'] = account_info.od_days_ever.apply(lambda x: 1 if x>7 else 0)
account_info['到期7天以上'] = account_info.last_repay_date.apply(lambda x: 1 if (datetime.date.today()-x).days>7 else 0)
account_info['催回'] = account_info.apply(lambda x: 1 if x.od_days_ever>0 and x.od_days==0 else 0,axis=1)
account_info['7天内催回'] = account_info.apply(lambda x: 1 if 7>=x.od_days_ever>0 and x.od_days==0 else 0,axis=1)
account_info['当前逾期'] = account_info.account_status.apply(lambda x: 1 if x=='ACCOUNT_OVERDUE' else 0)
account_info['展期'] = account_info.account_status.apply(lambda x: 1 if x=='ACCOUNT_CLOSED' else 0)
account_info['逾期后展期'] = account_info.apply(lambda x: 1 if x.account_status=='ACCOUNT_CLOSED' and x.od_days_ever>0 else 0,axis=1)
account_info['逾期后7天内展期'] = account_info.apply(lambda x: 1 if x.account_status=='ACCOUNT_CLOSED' and 7>=x.od_days_ever>0 else 0,axis=1)
account_info['结清'] = account_info.account_status.apply(lambda x: 1 if x=='ACCOUNT_SETTLE' else 0)
account_info['放款'] = 1
account_info['interest'] = account_info.apply(lambda x: x.contract_amount*x.loan_term*0.01, axis=1)
trans_journal_detail.columns = trans_journal_detail.columns.str.lower()
transRepayAmt = trans_journal_detail.groupby('contract_no',as_index=False)['trans_amount'].agg({'transRepayAmt':sum})
publicRepayAmt = realtime_draw.query("busi_type=='PUBLIC_TRANSFER'").groupby('contract_no',as_index=False)['real_amount'].agg({'publicRepayAmt':sum})
account_info = pd.merge(account_info,transRepayAmt,on='contract_no',how='left')
account_info = pd.merge(account_info,publicRepayAmt,on='contract_no',how='left')
account_info.transRepayAmt.fillna(0,inplace=True)
account_info.publicRepayAmt.fillna(0,inplace=True)
account_info['actualRepayAmt'] = account_info['transRepayAmt'] + account_info['publicRepayAmt']
account_info['已还够到手'] = account_info.apply(lambda x: 1 if x.actualRepayAmt>=x.loan_amount else 0, axis=1)
account_info['已还够合同'] = account_info.apply(lambda x: 1 if x.actualRepayAmt>=x.contract_amount else 0, axis=1)
account_info['已还够合同与利息'] = account_info.apply(lambda x: 1 if x.actualRepayAmt>=x.contract_amount+x.interest else 0, axis=1)

#统计客户还款行为特征:提前还款次数、逾期还款次数
def repay_var(times):
    cols = ['user_code','自然逾期','od_days_ever']
    for i in range(2,times):
        tmp = account_info.loc[account_info.number_of_loans<i,cols]
        var = tmp.groupby('user_code').agg({'自然逾期':'sum','od_days_ever':'max'})
        var.reset_index(inplace=True)
        var['number_of_loans'] = i
        if i == 2:
            ret = var
        else:
            ret = pd.concat([ret,var])
    ret.rename(columns={'od_days_ever':'曾经最大逾期天数','自然逾期':'逾期还款次数'},inplace=True)
    return ret

repay_var = repay_var(15)
account_info = pd.merge(account_info,repay_var,how='left',
                        on=['user_code','number_of_loans'])        

#还款类型是提前还款、正常还款、逾期还款的判断
def payment_type(x):
    if x.account_status=='ACCOUNT_SETTLE':
        if x.clear_date<x.last_repay_date:
            return 'early_settle'
        elif x.clear_date==x.last_repay_date:
            return 'normal_settle'
        else:
            return 'overdue_settle'
    else:
        return x.account_status
    
account_info['payment_type'] = account_info.apply(payment_type,axis=1)

#截止cut_date（通常是yesterday)状态的判断
def status_at_cutdate(x):
    if x.last_repay_date>x.cut_date:
        if x.clear_date<=x.cut_date:
            return 'Early_settle'
        elif x.account_status=='ACCOUNT_CLOSED' and x.extend_date<=x.cut_date:
            return 'Extension'
        else:
            return 'Normal'
    elif x.clear_date>x.cut_date:
        return 'Overdue'
    elif x.clear_date==None or pd.isnull(x.clear_date):
        if x.account_status=='ACCOUNT_CLOSED' and x.extend_date<=x.cut_date:
            return 'Extension'
        else:
            return 'Overdue'
    elif x.clear_date==x.last_repay_date:
        return 'Normal_settle'
    elif x.clear_date<x.last_repay_date:
        return 'Early_settle'
    else:
        return 'Overdue_settle'
    
account_info['cut_date'] = (datetime.date.today()-datetime.timedelta(days=1))
account_info['status_at_cutdate'] = account_info.apply(status_at_cutdate,axis=1)

#截止cut_date逾期天数的判断
def od_days_at_cutdate(x):
    if x.loan_date.date()>x.cut_date:
        return -3   #未放款
    elif x.last_repay_date>x.cut_date:
        return -2   #放款未到期
    elif x.clear_date>=x.cut_date:
        return (x.cut_date - x.last_repay_date).days    #逾期
    elif x.clear_date==None or pd.isnull(x.clear_date):
        if x.account_status=='ACCOUNT_CLOSED' and x.extend_date<x.cut_date:
            return -4  #展期
        else:
            return (x.cut_date - x.last_repay_date).days    #逾期
    else:
        return -1   #结清

##每月月底的逾期天数
#import calendar
#
#year = 2019
#tmp = account_info.copy()
#vintage = pd.DataFrame()
#for m in range(5,12):
#    d = calendar.monthrange(year,m)
#    tmp['cut_date'] = datetime.date(year,m,d[1])
#    tmp['od_days_at_cutdate'] = tmp.apply(od_days_at_cutdate,axis=1)
#    tmp['status_at_cutdate'] = tmp.apply(status_at_cutdate,axis=1)
#    if m == 5:
#        vintage = tmp.query("od_days_at_cutdate!=-3")
#    else:
#        vintage = pd.concat([vintage,tmp.query("od_days_at_cutdate!=-3")])
#
#vintage['od_days_at_cutdate_grp'] = pd.cut(vintage['od_days_at_cutdate'],
#                                           bins=[-np.inf,-1,3,10,30,60,np.inf],
#                                           labels=['nod','od_0-3','od_4-10','od_11-30','od_31-60','od_60+'])
#vintage.to_excel(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\vintage.xlsx",index=False)
 
repay_mart = account_info
repay_mart = pd.merge(repay_mart,borrower_info,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,employment_info,how='left',on='apply_code')
white1 = pd.read_excel(r"E:\guan\菲律宾\分析\邀请名单\邀请名单数据.xlsx",sheet_name=r'白名单1224',dtype={'phone_no':str})
white1['白名单1'] = 1
repay_mart = pd.merge(repay_mart,white1,how='left',on='phone_no')
repay_mart.to_excel(r"repay_mart.xlsx",index=False)


appr_mart = apply_appr_loan_info
appr_mart = pd.merge(appr_mart,number_of_apply,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,borrower_info,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,employment_info,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,auto_audit_result,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,manual_audit_result,how='left',on='apply_code')
appr_mart['自动拒绝'] = appr_mart.approval_status.apply(
                        lambda x: 1 if x=='AUTO_REJECTED' else 0)
appr_mart['自动通过'] = appr_mart.approval_status.apply(
                        lambda x: 1 if x=='NO_MANUAL_REVIEW' else 0)
appr_mart['取消'] = appr_mart.approval_status.apply(
                        lambda x: 1 if x=='MANUAL_CANCEL' else 0)
appr_mart['锁定期'] = appr_mart.auto_refuse_reason.apply(
                        lambda x: 1 if x=='Lock periodic rejection' else 0)
appr_mart['通过'] = appr_mart.approval_status.apply(
                    lambda x: 1 if x in ['MANUAL_APPROVED','NO_MANUAL_REVIEW'] else 0)
appr_mart['处理'] = appr_mart.approval_status.apply(
                    lambda x: 1 if x in ['MANUAL_APPROVED','MANUAL_REJECTED','AUTO_REJECTED','NO_MANUAL_REVIEW'] else 0)
appr_mart['人工通过'] = appr_mart.approval_status.apply(lambda x: 1 if x=='MANUAL_APPROVED' else 0)
appr_mart['人工处理'] = appr_mart.approval_status.apply(
                        lambda x: 1 if x in ['MANUAL_APPROVED','MANUAL_REJECTED'] else 0)
appr_mart = pd.merge(appr_mart,white1,how='left',on='phone_no')
appr_mart.to_excel(r"appr_mart.xlsx",index=False)

#%%---------柱形图
#每日申请量柱形图
(pd.pivot_table(appr_mart,values='apply_code',index='apply_day',
               columns='loan_type',aggfunc='count',fill_value=0)
    .plot(kind='bar',use_index=False,title='每日申请量',table=True)
)
#每日放款量柱形图
(pd.pivot_table(appr_mart,values='contract_no',index='pay_day',
                columns='loan_type',aggfunc=np.size,fill_value=0)
    .plot(kind='bar',use_index=False,title='每日放款量',table=True)
#    .plot(kind='bar',rot=45,title='每日放款量')
)



trans_j_d = trans_journal_detail.loc[:,['contract_no','trans_amount','settle_date']]
trans_j_d['settle_day'] = pd.to_datetime(trans_j_d['settle_date']).dt.date
realtime_d = realtime_draw.query("busi_type=='PUBLIC_TRANSFER'").loc[:,['contract_no','amount','created_time']]
realtime_d['settle_day'] = pd.to_datetime(realtime_d['created_time']).dt.date
realtime_d.rename(columns={'amount':'trans_amount'},inplace=True)
repayamt_d = pd.concat([trans_j_d,realtime_d],sort=True)
repayamt = repayamt_d.groupby(['contract_no','settle_day'])['trans_amount'].agg(sum).reset_index()


od_col = pd.DataFrame()
tmp = repay_mart.query("loan_date>='2020-07-01'").loc[:,['contract_no','loan_date','clear_date',
                          'last_repay_date','loan_type','job_province','account_status','extend_date']]
for (i,cut_date) in enumerate(pd.date_range(datetime.date(2020,8,1),
                                            datetime.date.today(),
                                            closed='left')):
    tmp['cut_date'] = cut_date.date()
    tmp['od_days_at_cutdate'] = tmp.apply(od_days_at_cutdate,axis=1)
    tmp['status_at_cutdate'] = tmp.apply(status_at_cutdate,axis=1)
    tmp['到期'] = tmp.apply(lambda x: 1 if x.last_repay_date==x.cut_date else 0, axis=1 )
    tmp['自然逾期'] = tmp.apply(lambda x: 1 if x.last_repay_date==x.cut_date and x.status_at_cutdate=='Overdue' else 0, axis=1)
    tmp['到期日展期'] = tmp.apply(lambda x: 1 if x.last_repay_date==x.cut_date and x.status_at_cutdate=='Extension' else 0, axis=1)  
    tmp['展期'] = tmp.apply(lambda x: 1 if x.extend_date==x.cut_date and x.status_at_cutdate=='Extension' else 0, axis=1)  
    tmp['催回'] = tmp.apply(lambda x: 1 if x.clear_date==x.cut_date and x.status_at_cutdate=='Overdue_settle' else 0, axis=1)
    tmp['提前结清'] = tmp.apply(lambda x: 1 if x.clear_date==x.cut_date and x.status_at_cutdate=='Early_settle' else 0, axis=1)
    rp_tmp = repayamt.loc[repayamt.settle_day==cut_date.date(),:]
    tmp2 = pd.merge(tmp,rp_tmp,on='contract_no',how='left')
    tmp2['有还款'] = tmp2.apply(lambda x: 1 if x.trans_amount>0 else 0, axis=1)
    tmp2['结清'] = tmp2.apply(lambda x: 1 if x.clear_date==x.cut_date else 0,axis=1)
    tmp2['结清金额'] = tmp2.apply(lambda x: x.trans_amount if x.clear_date==x.cut_date else 0, axis=1)
    if i == 0:
        od_col = tmp2.query("od_days_at_cutdate!=-3")
    else:
        od_col = pd.concat([od_col,tmp2.query("od_days_at_cutdate!=-3")])

od_col['od_days_at_cutdate_grp'] = pd.cut(od_col['od_days_at_cutdate'],
                                          bins=[-np.inf,0,3,7,15,30,np.inf],
                                          labels=['0','od1_3','od4_7','od8_15','od16_30','od30+'])
od_col.trans_amount.fillna(0,inplace=True)
od_col.rename(columns={'trans_amount':'还款金额'},inplace=True)
od_col['队列'] = 1
od_col.to_excel("od_col.xlsx",index=False)

fod = od_col.groupby('cut_date')['到期','自然逾期'].agg(sum)
fod['自然逾期率'] = fod['自然逾期']/fod['到期']
od1_3 = od_col.query("od_days_at_cutdate_grp=='od1_3'").groupby('cut_date')['催回'].agg(队列=np.size,催回=sum)
od1_3['od1_3催回率'] = od1_3['催回']/od1_3['队列']
od4_7 = od_col.query("od_days_at_cutdate_grp=='od4_7'").groupby('cut_date')['催回'].agg(队列=np.size,催回=sum)
od4_7['od4_7催回率'] = od4_7['催回']/od4_7['队列']
od8_15 = od_col.query("od_days_at_cutdate_grp=='od8_15'").groupby('cut_date')['催回'].agg(队列=np.size,催回=sum)
od8_15['od8_15催回率'] = od8_15['催回']/od8_15['队列']
od16_30 = od_col.query("od_days_at_cutdate_grp=='od16_30'").groupby('cut_date')['催回'].agg(队列=np.size,催回=sum)
od16_30['od16_30催回率'] = od16_30['催回']/od16_30['队列']
od30p = od_col.query("od_days_at_cutdate_grp=='od30+'").groupby('cut_date')['催回'].agg(队列=np.size,催回=sum)
od30p['od30+催回率'] = od30p['催回']/od30p['队列']



#%%---------数据库导入数据及生成数据集appr_mart & repay_mart的逻辑--------------------------
#apply_info = pd.read_sql("select * from approval.apply_info",cnx)
apply_info = pd.read_sql("select apply_code, user_code, apply_loan_amount, \
                                 apply_time, duration, loan_type, case_state, \
                                 usage_of_loan_text, number_of_loans, \
                                 approval_status, payment_type \
                        from approval.apply_info",cnx)

loan_status = pd.read_sql("select apply_code, contract_no, status as loan_status \
                            from approval.loan_info",cnx)
loan_status['loan_status'].replace({6:'放款成功'},inplace=True)

service_fee_rate = pd.read_sql("select apply_code, service_fee_rate \
                               from approval.contract_info",cnx)

auto_audit_result = pd.read_sql("select apply_code, refuse_info_1 as auto_refuse_reason \
                                  from approval.auto_audit_result \
                                  where refuse_info_1 is not null",cnx)
manual_audit_result = pd.read_sql("select apply_code, result_type as manual_result_type, \
                                    refuse_info_1 as manual_refuse_reason_1, \
                                    refuse_info_2 as manual_refuse_reason_2 \
                                    from approval.manual_audit_result",cnx)

account_info = pd.read_sql("select contract_no, account_status, loan_date, \
                              date_format(loan_date,'%Y%m') as loan_month, \
                              date_format(loan_date,'%Y%m%d') as loan_day, \
                              clear_date, loan_term, last_repay_date, \
                              contract_amount, loan_amount, borrower_tel_one, \
                              loan_channel \
                              from account.account_info",cnx)

od_days_ever = pd.read_sql("select contract_no, overdue_days as od_days_ever, \
                      case when bill_status='BILL_CLEAR' then 0 else overdue_days end as od_days \
                      from account.bill_main", cnx)

#基本信息
borrower_info = pd.read_sql("select apply_code, id_no, id_type, age, \
                            children_text, education, gender, marriage, \
                            province_addr_text as home_province, \
                            city_addr_text as home_city, \
                            district_addr_text as home_district, \
                            length_of_residence, device_id, phone_no \
                            from approval.borrower_info",cnx)
borrower_info['age_grp'] = pd.cut(borrower_info.age,
                                 bins=[-np.inf,21,28,35,49,np.inf],
                                 labels=['  - 21','22 - 28','29 - 35','36 - 49','50 -  '])
#工作信息
employment_info = pd.read_sql("select apply_code, job_type, \
                              monthly_net_income_text, on_the_job_time, \
                              province_addr_text as job_province, \
                              city_addr_text as job_city, \
                              district_addr_text as job_district\
                              from approval.employment_info",cnx)
#申请序数
number_of_apply = apply_info.loc[:,['apply_code','user_code','apply_time']]
number_of_apply['number_of_apply'] = number_of_apply['apply_time'].groupby(number_of_apply['user_code']).rank(method='dense')
number_of_apply.sort_values(['user_code','apply_time'],inplace=True)
number_of_apply.drop(columns=['user_code','apply_time'],inplace=True)
#关联信息
base_derived_variable = pd.read_sql("select * from approval.base_derived_variable",cnx)
df = base_derived_variable.copy()
#20190829上线的关联用到的字段
ir_var = [
'apply_code','user_code',

'ir_id_x_cell_cnt',
'ir_name_x_cell_cnt',
'ir_m12_home_addr_x_cell_cnt',
'ir_m12_linkman_cell_x_cell_cnt',
'ir_m12_tel_company_x_cell_cnt',

'ir_cell_x_id_cnt',
'ir_name_x_id_cnt',
'ir_m12_home_addr_x_id_cnt',
'ir_m12_linkman_cell_x_id_cnt',
'ir_m12_tel_company_x_id_cnt',

'ir_m12_cell_x_tel_company_cnt',
'ir_m12_id_x_tel_company_cnt',
'ir_m12_linkman_cell_x_tel_company_cnt',
'ir_m12_home_addr_x_company_home_cnt',

'ir_m12_cell_x_biz_addr_cnt',
'ir_m12_id_x_biz_addr_cnt',

'ir_m12_cell_x_home_addr_cnt',
'ir_m12_id_x_home_addr_cnt',
'ir_m12_tel_company_x_home_addr_cnt',

'ir_cell_x_name_cnt',
'ir_id_x_name_cnt'
]
var_df = df.loc[:,ir_var]
#来源渠道
event_info = pd.read_sql("select apps_flyer_id, device_type, install_time, \
                         operator, os_version, media_source \
                         from suncash_lend.af_event_info_detail",cnx)
apps_flyer_id = pd.read_sql("select apply_code, apps_flyer_id \
                            from suncash_lend.apply_info",cnx)
source = pd.merge(event_info,apps_flyer_id,on='apps_flyer_id')
#通讯录联系人数量   设备指纹关联信息里面有，这里可以注释掉
#contact_num = pd.read_sql("select apply_code, max(contact_num) as contact_num \
#                          from approval.address_book_info \
#                          group by apply_code",cnx)
#contact_num['contact_num_grp'] = pd.cut(contact_num['contact_num'],
#                                        bins=[0,10,20,50,100,10000],
#                                        labels=['1-10','11-20','21-50','51-100','101-10000'])
#设备指纹关联信息
device_ir_info = pd.read_sql("SELECT * FROM approval.device_info_derived_variable",cnx)
device_ir_info.drop(columns=['id','user_code','created_time','updated_time'],inplace=True)
device_ir_info['contact_num_grp'] = pd.cut(device_ir_info['eqc_direc_cell_num'].astype(int),
                                        bins=[0,10,20,50,100,np.inf],
                                        labels=['1-10','11-20','21-50','51-100','101-'])
#开始申请和申请提交所在时点及用时
apply_time = apply_info.loc[:,['apply_code','apply_time']]
apply_time['apply_hour'] = apply_time['apply_time'].dt.strftime('%H')
apply_start_time = pd.read_sql("select apply_code, \
                                       created_time as apply_start_time, \
                                       date_format(created_time,'%H') as apply_start_hour \
                               from suncash_lend.apply_info",cnx)
apply_time = pd.merge(apply_time,apply_start_time,how='left',on='apply_code')
apply_time['apply_interval'] = (apply_time['apply_time'] - apply_time['apply_start_time']).dt.seconds
apply_time['apply_hour_grp'] = pd.cut(apply_time.apply_hour.astype(int),
                                     bins=[-1,5,8,17,22,100],
                                     labels=['0-5','6-8','9-17','18-22','23-0'])
apply_time['apply_start_hour_grp'] = pd.cut(apply_time.apply_start_hour.astype(int),
                                     bins=[-1,5,8,17,22,100],
                                     labels=['0-5','6-8','9-17','18-22','23-0'])
apply_time['apply_interval_grp'] = pd.cut(apply_time.apply_interval.astype(int),
                                     bins=[-1,0,30,60,120,300,600,np.inf],
                                     labels=['0','1-30','31-60','61-120','121-300','301-600','601-'])
apply_time.drop(columns=['apply_time','apply_start_time'],inplace=True)
#申请过程的位置信息
geo_point_distance = pd.read_sql("select apply_code, event_date, event_name, \
                                 home_address_distance, work_address_distance \
                                 from suncash_lend.geo_point_distance",cnx)
home_distance = geo_point_distance.groupby('apply_code',as_index=False)['home_address_distance'].agg({'home_distance_max':max,'home_distance_min':min})
work_distance = geo_point_distance.groupby('apply_code',as_index=False)['work_address_distance'].agg({'work_distance_max':max,'work_distance_min':min})
move_interval = geo_point_distance.groupby('apply_code',as_index=False)['event_date'].agg({'event_date_last':max,'event_date_first':min})
geo_distance = pd.merge(home_distance,work_distance,how='left',on='apply_code')
geo_distance = pd.merge(geo_distance,move_interval,how='left',on='apply_code')
geo_distance['home_distance_move'] = geo_distance['home_distance_max'] - geo_distance['home_distance_min']
geo_distance['work_distance_move'] = geo_distance['work_distance_max'] - geo_distance['work_distance_min']
geo_distance['distance_move_interval'] = (geo_distance['event_date_last'] - geo_distance['event_date_first']).dt.seconds
geo_distance.drop(columns=['event_date_last','event_date_first'],inplace=True)
geo_distance['home_distance_min_grp'] = pd.cut(geo_distance['home_distance_min'],
                                            bins=[-1,100,500,1000,2000,3000,5000,10000,20000,50000,np.inf])
geo_distance['work_distance_min_grp'] = pd.cut(geo_distance['work_distance_min'],
                                            bins=[-1,100,500,1000,2000,3000,5000,10000,20000,50000,np.inf])
geo_distance['home_distance_move_grp'] = pd.cut(geo_distance['home_distance_move'],
                                            bins=[-1,100,500,1000,2000,3000,5000,10000,20000,50000,np.inf])
geo_distance['distance_move_interval_grp'] = pd.cut(geo_distance['distance_move_interval'],
                                            bins=[-1,600,1800,3600,7200,np.inf],
                                            labels=['(0-10min]','(10min,30min]','(30min,1h]','(1h,2h]','(2h,)'])
#各信息项填写时间
info_time = geo_point_distance.loc[:,['apply_code', 'event_date', 'event_name']]
info_time.sort_values(by=['apply_code','event_date'],inplace=True)
info_time['last_event_date'] = info_time['event_date'].shift(1)
info_time = info_time.loc[info_time.event_name!='SAVE_APPLY_INFO',:]
info_time['event_interval'] = (info_time['event_date'] - info_time['last_event_date']).dt.seconds
identification = info_time.loc[info_time.event_name=='SAVE_IDENTIFICATION',:].groupby('apply_code',as_index=False)['event_interval'].agg({'IDENTIFICATION_save_cnt':np.size,'IDENTIFICATION_use_time':np.sum})
identification['IDENTIFICATION_use_time_grp'] = pd.cut(identification.IDENTIFICATION_use_time,
                                                  bins=[0,60,120,300,600,np.inf],
                                                  labels=['0-60','61-120','121-300','301-600','601-'])
personal = info_time.loc[info_time.event_name=='SAVE_PERSONAL_INFO',:].groupby('apply_code',as_index=False)['event_interval'].agg({'PERSONAL_save_cnt':np.size,'PERSONAL_use_time':np.sum})
personal['PERSONAL_use_time_grp'] = pd.cut(personal.PERSONAL_use_time,
                                      bins=[0,60,120,300,600,np.inf],
                                      labels=['0-60','61-120','121-300','301-600','601-'])
contact = info_time.loc[info_time.event_name=='SAVE_CONTACT_INFO',:].groupby('apply_code',as_index=False)['event_interval'].agg({'CONTACT_save_cnt':np.size,'CONTACT_use_time':np.sum})
contact['CONTACT_use_time_grp'] = pd.cut(contact.CONTACT_use_time,
                                      bins=[0,60,120,300,600,np.inf],
                                      labels=['0-60','61-120','121-300','301-600','601-'])
employment = info_time.loc[info_time.event_name=='SAVE_EMPLOYMENT_INFO',:].groupby('apply_code',as_index=False)['event_interval'].agg({'EMPLOYMENT_save_cnt':np.size,'EMPLOYMENT_use_time':np.sum})
employment['EMPLOYMENT_use_time_grp'] = pd.cut(employment.EMPLOYMENT_use_time,
                                          bins=[0,60,120,300,600,np.inf],
                                          labels=['0-60','61-120','121-300','301-600','601-'])
facedetection = info_time.loc[info_time.event_name=='SAVE_FACEDETECTION_INFO',:].groupby('apply_code',as_index=False)['event_interval'].agg({'FACEDETECTION_save_cnt':np.size,'FACEDETECTION_use_time':np.sum})
facedetection['FACEDETECTION_use_time_grp'] = pd.cut(facedetection.FACEDETECTION_use_time,
                                                  bins=[0,60,120,300,600,np.inf],
                                                  labels=['0-60','61-120','121-300','301-600','601-'])
loan = info_time.loc[info_time.event_name=='SAVE_LOAN_INFO',:].groupby('apply_code',as_index=False)['event_interval'].agg({'LOAN_save_cnt':np.size,'LOAN_use_time':np.sum})
loan['LOAN_use_time_grp'] = pd.cut(loan.LOAN_use_time,
                              bins=[0,60,120,300,600,np.inf],
                              labels=['0-60','61-120','121-300','301-600','601-'])
#活体与证件人脸比对相似度
similarity = pd.read_sql("select apply_code, similarity \
                         from suncash_data.td_task_info \
                         where similarity<>''",cnx)


repay_mart = pd.merge(repay_mart,number_of_apply,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,var_df,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,source,how='left',on='apply_code')
#repay_mart = pd.merge(repay_mart,contact_num,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,device_ir_info,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,apply_time,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,geo_distance,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,identification,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,personal,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,contact,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,employment,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,facedetection,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,loan,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,similarity,how='left',on='apply_code')

appr_mart = pd.merge(appr_mart,number_of_apply,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,source,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,var_df,how='left',on='apply_code')
#appr_mart = pd.merge(appr_mart,contact_num,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,auto_audit_result,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,manual_audit_result,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,device_ir_info,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,apply_time,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,geo_distance,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,identification,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,personal,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,contact,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,employment,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,facedetection,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,loan,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,similarity,how='left',on='apply_code')



#-------------------------------------------------------------------------------------------------


#%%---下面注释为历史代码和开发中的代码
#cursor = cnx.cursor()
#cursor.execute("show databases")
#databases = cursor.fetchall()
#
##database:suncash_lend -----------------------------------------------------
#cursor.execute("use suncash_lend")
#cursor.execute("show tables")
#lend_tables = cursor.fetchall()
#for tb in lend_tables:
#    locals()[tb[0]] = pd.read_sql("select * from "+tb[0]+" limit 10",cnx)
##table:apply_info
#cursor.execute("show create table apply_info")
#create_apply_info = cursor.fetchall()
#sql_stmt = "select user_code, apply_code, status, apply_amount, issue_amount,\
#            apply_day_en_name,\
#            date_format(created_time,'%Y-%m') as apply_month,\
#            date_format(created_time,'%Y-%m-%d') as apply_date,\
#            date_format(approval_start_time,'%Y-%m-%d') as approval_start_date,\
#            date_format(approval_end_time,'%Y-%m-%d') as approval_end_date \
#            from suncash_lend.apply_info"
#apply_info = pd.read_sql(sql_stmt,cnx)
#apply_info['status'].replace({1:'初始化',10:'审批中',12:'审批通过',20:'提现',
#                              125:'案件拒绝',126:'案件结束',127:'案件取消'},
#                            inplace=True)
#pd.pivot_table(apply_info,
#               index=['approval_end_date','apply_day_en_name'],
#               columns=['status'],
#               values='apply_code',
#               aggfunc=np.size,
#               fill_value=0,
#               margins=True)
#a = pd.pivot_table(apply_info,
#               index='approval_end_date',
#               columns='status',
#               values='issue_amount',
#               aggfunc=[np.size,np.sum],
#               fill_value='')
#a = pd.crosstab(index=apply_info.approval_end_date,
#                columns=apply_info.status,
#                values=apply_info.issue_amount,
#                aggfunc=[np.size,np.sum])
#a.fillna('',inplace=True)
##----------------------------------------------------------------------------
#
##database:approval-----------------------------------------------------------
#cursor.execute("use approval")
#cursor.execute("show tables")
#approval_tables = cursor.fetchall()
##table:apply_info
#cursor.execute("show create table apply_info")
#create_apply_info = cursor.fetchall()
#apply_info = pd.read_sql("select * from approval.apply_info",cnx)
#apply_info['apply_month'] = apply_info['apply_time'].dt.strftime('%Y%m')
##apply_info['apply_month'] = apply_info['apply_time'].apply(lambda x:datetime.datetime.strftime(x,'%Y%m'))
#%统计每月申请的订单量
#a = apply_info.groupby('apply_month').size()
#apply_Num = pd.DataFrame(a,columns=['申请的订单量'])
##统计每月申请的客户量(有点活跃用户量的意思)
#user = apply_info.ix[:,['user_code','apply_month','apply_time']]
#user.sort_values(by=['user_code','apply_time'],inplace=True)
#user.drop_duplicates(subset=['user_code','apply_month'],keep='first',inplace=True)
#a = user.groupby('apply_month').size()
#user_Num = pd.DataFrame(a,columns=['申请的客户量'])
##统计每月新增的客户量
#new_user = apply_info.ix[:,['user_code','apply_month','apply_time']]
#new_user.sort_values(by=['user_code','apply_time'],inplace=True)
#new_user.drop_duplicates(subset='user_code',keep='first',inplace=True)
#a = new_user.groupby('apply_month').size()
#new_user_Num = pd.DataFrame(a,columns=['新增的客户量'])
##统计每月复贷客户申请的订单量
#re_apply = apply_info.ix[apply_info.loan_type=='re_apply',['user_code','apply_month','apply_time']]
##re_apply = apply_info[apply_info.loan_type=='re_apply'].loc[:,['user_code','apply_month','apply_time']]
#a = re_apply.groupby('apply_month').size()
#re_apply_Num = pd.DataFrame(a,columns=['复贷客户申请的订单量'])
##统计每月申请的复贷客户量
#re_apply.sort_values(['user_code','apply_time'],inplace=True)
#re_user = re_apply.drop_duplicates(['user_code','apply_month'])
#a = re_user.groupby('apply_month').size()
#re_user_Num = pd.DataFrame(a,columns=['申请的复贷客户量'])
##统计每月拒绝后再次申请的客户量
#first_apply = apply_info.ix[apply_info.loan_type=='first_apply',['user_code','apply_month','apply_time']]
#first_apply.sort_values(by=['user_code','apply_time'],inplace=True)
#dup_apply_tag = first_apply.duplicated('user_code')
#refuse_first_apply = first_apply.loc[dup_apply_tag,:]
#refuse_user = refuse_first_apply.drop_duplicates(['user_code','apply_month'])
#a = refuse_user.groupby('apply_month').size()
#refuse_user_Num = pd.DataFrame(a,columns=['拒绝后再次申请的客户量'])
#
##table:loan_info
#cursor.execute("show create table loan_info")
#create_loan_info = cursor.fetchall()
#loan_info = pd.read_sql("select apply_code, pay_time, status, \
#                          date_format(pay_time,'%Y%m') as loan_month \
#                          from approval.loan_info \
#                          where status=6",cnx)
#user_apply_code = apply_info.loc[:,['user_code','apply_code','loan_type']]
#loan_info = pd.merge(loan_info,user_apply_code,on='apply_code')
##统计每月的放款笔数
#a = loan_info.groupby('loan_month').size()
#loan_Num = pd.DataFrame(a,columns=['放款笔数'])
##统计每月的放款客户数
#loan_info.sort_values(by=['user_code','pay_time'],inplace=True)
#loan_user = loan_info.drop_duplicates(subset=['user_code','loan_month'])
#a = loan_user.groupby('loan_month').size()
#loan_user_Num = pd.DataFrame(a,columns=['放款客户数'])
##统计每月新增的放款客户数
#new_loan_user = loan_info.drop_duplicates(subset=['user_code'])
#a = new_loan_user.groupby('loan_month').size()
#new_loan_user_Num = pd.DataFrame(a,columns=['新增放款客户数'])
##统计每月复贷的放款客户数
#re_loan_info = loan_info.ix[loan_info.loan_type=='re_apply',:]
#re_loan_user = re_loan_info.drop_duplicates(['user_code','loan_month'])
#a = re_loan_user.groupby('loan_month').size()
#re_loan_user_Num = pd.DataFrame(a,columns=['复贷放款客户数'])
##统计每月复贷的放款笔数
#a = re_loan_info.groupby('loan_month').size()
#re_loan_Num = pd.DataFrame(a,columns=['复贷放款笔数'])
#
##拼接
#Num = pd.concat([user_Num,new_user_Num,re_user_Num,refuse_user_Num,
#                 apply_Num,re_apply_Num,
#                 loan_Num,loan_user_Num,new_loan_user_Num,re_loan_user_Num,
#                 re_loan_Num],axis=1)
#
##首次申请时间
#first_apply_month = first_apply.loc[:,['user_code','apply_month']]
#first_apply_month.drop_duplicates(subset='user_code',inplace=True)
##审批通过
#approved = apply_info.loc[apply_info.approval_status=='MANUAL_APPROVED',['user_code','approval_status']]
#approved.drop_duplicates(subset='user_code',inplace=True)
#first_apply_month = pd.merge(first_apply_month,approved,how='left',on='user_code')
##放款成功
#loan = loan_info.loc[loan_info.status==6,['user_code','status']]
#loan.drop_duplicates(subset='user_code',inplace=True)
#first_apply_month = pd.merge(first_apply_month,loan,how='left',on='user_code')
#first_apply_month.rename(columns={'apply_month':'first_apply_month','status':'loan_status'},
#                         inplace=True)
#first_apply_month.fillna('',inplace=True)
#a = pd.pivot_table(first_apply_month,
#                   values='user_code',
#                   index=['approval_status','loan_status'],
#                   columns='first_apply_month',
#                   aggfunc=len,
#                   margins=True,
#                   fill_value=0)
#a.to_excel(r"C:\Users\lenovo\Desktop\first_apply_month.xlsx")
#
##number_of_loans
#number_of_loans = apply_info.drop_duplicates(subset=['user_code','number_of_loans'])
#number_of_loans.groupby('number_of_loans').size()
##-----------------------------------------------------------------------------
#
##审核结果----------------------------------------------------------------------
#auto_audit_result = pd.read_sql("select apply_code, refuse_info_1 as auto_refuse_reason \
#                                  from approval.auto_audit_result \
#                                  where refuse_info_1 is not null",cnx)
##auto_audit_result.sort_values(by='apply_code',inplace=True)
#manual_audit_result = pd.read_sql("select apply_code, result_type as manual_result_type, \
#                                    refuse_info_1 as manual_refuse_reason_1, \
#                                    refuse_info_2 as manual_refuse_reason_2, opition \
#                                    from approval.manual_audit_result",cnx)
##manual_audit_result.sort_values(by='apply_code',inplace=True)
#
##新客户申请通过率
#first_apply_info = pd.read_sql("select user_code, apply_time, \
#                                date_format(apply_time,'%Y%m') as apply_month, \
#                                number_of_loans, approval_status \
#                                from approval.apply_info \
#                                where loan_type='first_apply'",cnx)
##因无拒绝锁定期，拒绝后重复申请得到的通过率
#a = pd.pivot_table(first_apply_info,
#                    index='apply_month',
#                    columns='approval_status',
#                    values='user_code',
#                    aggfunc=np.size,
#                    margins=True,
#                    fill_value=0)
##客户第一次申请的通过率
#first_apply_info.sort_values(by=['user_code','apply_time'],inplace=True)
#first_apply_info.drop_duplicates(subset=['user_code'],inplace=True)
#b = pd.pivot_table(first_apply_info,
#                    index='apply_month',
#                    columns='approval_status',
#                    values='user_code',
#                    aggfunc=np.size,
#                    margins=True,
#                    fill_value=0)
#c = pd.concat([a,b],keys=['新客户所有申请','新客户第一次申请'])
#c.to_excel(r"C:\Users\lenovo\Desktop\new_user_approval_rate.xlsx")
#
##复贷客户申请通过率
#re_apply_info = pd.read_sql("select user_code, apply_code, apply_time, \
#                              date_format(apply_time,'%Y%m') as apply_month, \
#                              number_of_loans, approval_status\
#                              from approval.apply_info \
#                              where loan_type='re_apply'",cnx)
##re_apply_info.sort_values('apply_code',inplace=True)
#re_apply_info = pd.merge(re_apply_info,auto_audit_result,how='left',on='apply_code')
#re_apply_info = pd.merge(re_apply_info,manual_audit_result,how='left',on='apply_code')
#re_apply_info.to_excel(r"C:\Users\lenovo\Desktop\re_apply_info.xlsx",index=False)
#
##database:account-------------------------------------------------------------
#cursor.execute("use account")
#cursor.execute("show tables")
#account_tables = cursor.fetchall()
##table:account_info
#cursor.execute("show create table account_info")
#create_account_info = cursor.fetchall()
#account_info = pd.read_sql("select contract_no, account_status, loan_date, \
#                              date_format(loan_date,'%Y%m') as loan_month, \
#                              date_format(loan_date,'%Y%m%d') as loan_day, \
#                              clear_date, loan_term, last_repay_date, \
#                              contract_amount, loan_amount, borrower_tel_one, \
#                              loan_channel \
#                              from account.account_info",cnx)
#user_apply_code = pd.read_sql("select apply_code, user_code, loan_type, number_of_loans \
#                                from approval.apply_info",cnx)
#apply_contract_code = pd.read_sql("select apply_code, contract_no \
#                                    from approval.loan_info",cnx)
#code = pd.merge(user_apply_code,apply_contract_code,on='apply_code')
#account_info = pd.merge(account_info,code,on='contract_no')
##缺失clear_date的处理
#account_info.loc[(account_info.account_status=='ACCOUNT_SETTLE') & (account_info.clear_date.isnull()),:]
#account_info.loc[account_info.contract_no=='C201906120207382040078','clear_date'] = datetime.date(2019,6,28)
#account_info.loc[account_info.contract_no=='C201906261851549490077','clear_date'] = datetime.date(2019,7,11)
##修正number_of_loans有误的合同
#account_info.loc[account_info.apply_code=='PL201905271454230580090','number_of_loans'] = 3
#account_info.loc[account_info.apply_code=='PL201905291324034000015','number_of_loans'] = 3
#account_info.loc[account_info.apply_code=='PL201905271244097550033','number_of_loans'] = 3
#
##还款类型是提前还款、正常还款、逾期还款的判断
#def payment_type(x):
#    if x.account_status=='ACCOUNT_SETTLE':
#        if x.clear_date<x.last_repay_date:
#            return 'early_settle'
#        elif x.clear_date==x.last_repay_date:
#            return 'normal_settle'
#        else:
#            return 'overdue_settle'
#    else:
#        return x.account_status
#    
#account_info['payment_type'] = account_info.apply(payment_type,axis=1)
#a = pd.pivot_table(account_info,
#                   values='contract_no',
#                   index=['loan_type','loan_month'],
#                   columns='payment_type',
#                   aggfunc=len,
#                   fill_value=0,
#                   margins=True)
##截止cut_date（通常是yesterday)状态的判断
#def status_at_cutdate(x):
#    if x.last_repay_date>x.cut_date:
#        if x.clear_date==None or x.clear_date>x.cut_date:
#            return 'Normal'
#        elif x.clear_date<=x.cut_date:
#            return 'Early_settle'
#    elif x.clear_date==None:
#        return 'Overdue'
#    elif x.clear_date==x.last_repay_date:
#        return 'Normal_settle'
#    elif x.clear_date<x.last_repay_date:
#        return 'Early_settle'
#    else:
#        return 'Overdue_settle'
#    
##account_info['cut_date'] = (datetime.datetime.today()-datetime.timedelta(days=1)).date()
#account_info['cut_date'] = (datetime.date.today()-datetime.timedelta(days=1))
#account_info['status_at_cutdate'] = account_info.apply(status_at_cutdate,axis=1)
##曾经逾期天数的判断
##def od_days_ever(x):
##    if x.account_status=='ACCOUNT_NORMAL':
##        return 0
##    elif x.account_status=='ACCOUNT_SETTLE' and x.clear_date<x.last_repay_date:
##        return 0
##    elif x.account_status=='ACCOUNT_SETTLE':
##        return (x.clear_date-x.last_repay_date).days
##    else:
##        return (datetime.date.today()-x.last_repay_date).days
##account_info['od_days_ever'] = account_info.apply(od_days_ever,axis=1)
#od_days_ever = pd.read_sql("select contract_no, overdue_days as od_days_ever, \
#                      case when bill_status='BILL_CLEAR' then 0 else overdue_days end as od_days \
#                      from account.bill_main", cnx)
#account_info = pd.merge(account_info,od_days_ever,how='left',on='contract_no')
#
#first_loan_month = account_info.loc[account_info.loan_type=='first_apply',['user_code','loan_month','borrower_tel_one']]
#first_loan_month.rename(columns={'loan_month':'first_loan_month'},inplace=True)
#account_info = pd.merge(account_info,first_loan_month,how='left',on='user_code')
##已结清的合同
account_settle = account_info.ix[account_info.account_status=='ACCOUNT_SETTLE',:]
account_settle.rename(columns={'number_of_loans':'第几次放款'},inplace=True)
account_settle['下次是第几次申请'] = account_settle['第几次放款'] + 1
user_apply_info = apply_info.loc[:,['user_code','apply_time','number_of_loans']]
user_apply_info.sort_values(['user_code','number_of_loans','apply_time'],inplace=True)
user_apply_info.drop_duplicates(['user_code','number_of_loans'],inplace=True)
account_settle = pd.merge(account_settle,user_apply_info,how='left',
                          left_on=['user_code','下次是第几次申请'],
                          right_on=['user_code','number_of_loans'])
account_settle['复贷申请时间间隔'] = (account_settle['apply_time'].dt.date-account_settle['clear_date']).dt.days
account_settle['复贷申请时间间隔'].fillna('未申请',inplace=True)
a = pd.pivot_table(account_settle,
                    index='第几次放款',
                    columns='复贷申请时间间隔',
                    values='contract_no',
                    aggfunc=np.size,
                    fill_value='',
                    margins=True,
                    dropna=False)
#a.to_excel(r"C:\Users\lenovo\Desktop\复贷申请时间间隔.xlsx")
a = pd.pivot_table(account_settle,
                    index=['first_loan_month','第几次放款'],
                    columns='复贷申请时间间隔',
                    values='contract_no',
                    aggfunc=np.size,
                    fill_value='',
                    margins=True,
                    dropna=False)
#
##提前还款与后续逾期的关系
#account_info['提前结清天数'] = (account_info['last_repay_date'] - account_info['clear_date']).dt.days
#cols = ['borrower_tel_one','payment_type','提前结清天数','clear_date','loan_date','loan_term','loan_amount']
##first_loan_201905_1st = account_info.loc[(account_info.first_loan_month=='201905') & (account_info.number_of_loans==1),cols]
##first_loan_201905_1st.rename(columns={'payment_type':'payment_type_1st','提前结清天数':'提前结清天数_1st'},inplace=True)
##first_loan_201905_2nd = account_info.loc[(account_info.first_loan_month=='201905') & (account_info.number_of_loans==2),cols]
##first_loan_201905_2nd.rename(columns={'payment_type':'payment_type_2nd','提前结清天数':'提前结清天数_2nd'},inplace=True)
##first_loan_201905_3rd = account_info.loc[(account_info.first_loan_month=='201905') & (account_info.number_of_loans==3),cols]
##first_loan_201905_3rd.rename(columns={'payment_type':'payment_type_3rd','提前结清天数':'提前结清天数_3rd'},inplace=True)
##把客户每次借款拼接到一行
##first_loan_month:首笔贷款发放月份 times:复贷放款次数
#def merge_month(first_loan_month,times):
#    for i in range(1,times):
#        tmp = account_info.loc[(account_info.first_loan_month==first_loan_month) & (account_info.number_of_loans==i),cols]
#        tmp.rename(columns={'payment_type':'payment_type_'+str(i),
#                            '提前结清天数':'提前结清天数_'+str(i),
#                            'loan_term':'loan_term_'+str(i),
#                            'loan_amount':'loan_amount_'+str(i)},inplace=True)
#        if i == 1:
#            reloan = tmp
#        else:
#            reloan = pd.merge(reloan,tmp,how='left',on='borrower_tel_one')
#            #reloan['下笔放款间隔天数'] = (reloan['loan_date'].dt.date-reloan['last_clear_date']).dt.days
#            reloan['下笔放款间隔天数'] = (reloan['loan_date'].dt.date-reloan['last_clear_date'])
#            reloan.drop(['last_clear_date'],axis=1,inplace=True)
#        reloan.rename(columns={'clear_date':'last_clear_date',
#                               '下笔放款间隔天数':'下笔放款间隔天数_'+str(i)},
#                        inplace=True)
#        reloan.drop(['loan_date'],axis=1,inplace=True)
#    return reloan   
#
#first_loan_201905 = merge_month('201905',9)
#first_loan_201905.to_excel(r"C:\Users\lenovo\Desktop\reloan_1st_loan_at_201905.xlsx",index=False)
#
#def merge(times):
#    for i in range(1,times):
#        tmp = account_info.loc[account_info.number_of_loans==i,cols]
#        tmp.rename(columns={'payment_type':'payment_type_'+str(i),
#                            '提前结清天数':'提前结清天数_'+str(i),
#                            'loan_term':'loan_term_'+str(i),
#                            'loan_amount':'loan_amount_'+str(i)},inplace=True)
#        if i == 1:
#            reloan = tmp
#        else:
#            reloan = pd.merge(reloan,tmp,how='left',on='borrower_tel_one')
#            #reloan['下笔放款间隔天数'] = (reloan['loan_date'].dt.date-reloan['last_clear_date']).dt.days
#            reloan['下笔放款间隔天数'] = (reloan['loan_date'].dt.date-reloan['last_clear_date'])
#            reloan.drop(['last_clear_date'],axis=1,inplace=True)
#        reloan.rename(columns={'clear_date':'last_clear_date',
#                               '下笔放款间隔天数':'下笔放款间隔天数_'+str(i)},
#                        inplace=True)
#        reloan.drop(['loan_date'],axis=1,inplace=True)
#    return reloan   
#
#reloan = merge(9)
#reloan = pd.merge(reloan,first_loan_month,on='borrower_tel_one')
#reloan.to_excel(r"C:\Users\lenovo\Desktop\reloan.xlsx",index=False)
#
##统计客户还款行为特征:提前还款次数、逾期还款次数
#account_info['提前还款'] = account_info['提前结清天数'].apply(lambda x: 1 if x>1 else 0)
#account_info['逾期还款'] = account_info['提前结清天数'].apply(lambda x: 1 if x<0 else 0)
#def repay_var(times):
#    cols = ['user_code','提前还款','逾期还款']
#    for i in range(2,times):
#        tmp = account_info.loc[account_info.number_of_loans<i,cols]
#        #var = tmp.groupby('user_code').agg({'提前还款':'sum','逾期还款':'sum'})
#        var = tmp.groupby('user_code').agg('sum')
#        var.reset_index(inplace=True)
#        var['number_of_loans'] = i
#        if i == 2:
#            ret = var
#        else:
#            ret = pd.concat([ret,var])
#    ret.rename(columns={'提前还款':'提前还款次数','逾期还款':'逾期还款次数'},inplace=True)
#    return ret
#
#repay_var = repay_var(9)
#account_info = pd.merge(account_info,repay_var,how='left',
#                        on=['user_code','number_of_loans'])        
#account_info.to_excel(r"C:\Users\lenovo\Desktop\account_info.xlsx",index=False)
#
##统计贷款次数与还款表现的关系--------------------------------------------------
#account_info.groupby('number_of_loans').size()
#pd.pivot_table(account_info,
#               values='contract_no',
#               index='number_of_loans',
#               columns='account_status',
#               aggfunc=np.size,
#               margins=True)
##----------------------------------------------------------------------------
##存量放款客户的状态分布，可以用于预测复贷量
#ac = account_info.sort_values(by=['borrower_tel_one','loan_date'])
#ac.drop_duplicates(subset='borrower_tel_one',keep='last',inplace=True)
#ac.groupby('account_status').size()
##结清客户流失情况
#ac['结清距今天数'] = (datetime.date.today()-ac['clear_date']).dt.days
#ac.groupby('结清距今天数').size()
#ac['od_days_ever'] = ac['clear_date'] - ac['last_repay_date']
#
##客户转化漏斗:进件-审批-放款---------------------------------------------------
#apply_info = pd.read_sql("select * from approval.apply_info",cnx)
#apply_info['apply_month'] = apply_info['apply_time'].dt.strftime('%Y%m')
#apply_info['apply_day'] = apply_info['apply_time'].dt.strftime('%Y%m%d')
#loan_status = pd.read_sql("select apply_code, status as loan_status \
#                            from approval.loan_info",cnx)
#loan_status['loan_status'].replace({6:'放款成功'},inplace=True)
#apply_appr_loan_info = pd.merge(apply_info,loan_status,how='left',on='apply_code')
#cols = ['user_code','apply_code','apply_time','apply_month',
#        'approval_status','loan_status','loan_type']
#apply_appr_loan_info = apply_appr_loan_info.loc[:,cols]
#apply_appr_loan_info['approval_status'] = apply_appr_loan_info['approval_status'].astype('category')
#apply_appr_loan_info['approval_status'].cat.set_categories(
#                                                    ['AUTO_IN_REVIEW',
#                                                    'AUTO_REJECTED',
#                                                    'MANUAL_REJECTED',
#                                                    'MANUAL_APPROVED',
#                                                    'MANUAL_CANCEL',
#                                                    'MANUAL_IN_REVIEW'],inplace=True)
#apply_appr_loan_info['loan_status'] = apply_appr_loan_info['loan_status'].astype('category')
#apply_appr_loan_info['loan_status'].cat.set_categories(['放款成功',5,7,12],inplace=True)
#auto_audit_result = pd.read_sql("select apply_code, refuse_info_1 as auto_refuse_reason \
#                                  from approval.auto_audit_result \
#                                  where refuse_info_1 is not null",cnx)
#manual_audit_result = pd.read_sql("select apply_code, result_type as manual_result_type, \
#                                    refuse_info_1 as manual_refuse_reason_1, \
#                                    refuse_info_2 as manual_refuse_reason_2, opition \
#                                    from approval.manual_audit_result",cnx)
#
##总订单的审批放款状态
#total_appr_status = pd.pivot_table(apply_appr_loan_info,
#                    values='apply_code',
#                    columns='apply_month',
#                    index='approval_status',
#                    aggfunc=np.size,
#                    fill_value=0,
#                    margins=True)
#total_loan_status = pd.pivot_table(apply_appr_loan_info,
#                    values='apply_code',
#                    columns='apply_month',
#                    index='loan_status',
#                    aggfunc=np.size,
#                    fill_value=0,
#                    margins=True)
#total_status = pd.concat([total_appr_status,total_loan_status],
#                         keys=['审批状态','放款状态'])
##新增申请用户订单的审批放款状态
#first_apply = apply_appr_loan_info.sort_values(by=['user_code','apply_time'])
#true_first_apply = first_apply.drop_duplicates(subset='user_code')
#first_apply_appr_status = pd.pivot_table(true_first_apply,
#                    values='apply_code',
#                    columns='apply_month',
#                    index='approval_status',
#                    aggfunc=np.size,
#                    fill_value=0,
#                    margins=True)
#first_apply_loan_status = pd.pivot_table(true_first_apply,
#                    values='apply_code',
#                    columns='apply_month',
#                    index='loan_status',
#                    aggfunc=np.size,
#                    fill_value=0,
#                    margins=True)
#first_apply_status = pd.concat([first_apply_appr_status,first_apply_loan_status],
#                         keys=['审批状态','放款状态'])
##复贷申请用户订单的审批放款状态
#re_apply = apply_appr_loan_info.loc[apply_appr_loan_info.loan_type=='re_apply',:]
#re_apply_appr_status = pd.pivot_table(re_apply,
#                    values='apply_code',
#                    columns='apply_month',
#                    index='approval_status',
#                    aggfunc=np.size,
#                    fill_value=0,
#                    margins=True)
#re_apply_loan_status = pd.pivot_table(re_apply,
#                    values='apply_code',
#                    columns='apply_month',
#                    index='loan_status',
#                    aggfunc=np.size,
#                    fill_value=0,
#                    margins=True)
#re_apply_status = pd.concat([re_apply_appr_status,re_apply_loan_status],
#                         keys=['审批状态','放款状态'])
##拒绝申请用户订单的审批放款状态
#first_apply = apply_appr_loan_info.loc[apply_appr_loan_info.loan_type=='first_apply',:]
#first_apply = first_apply.sort_values(by=['user_code','apply_time'])
#dup_apply_tag = first_apply.duplicated('user_code')
#refuse_first_apply = first_apply.loc[dup_apply_tag,:]
#refuse_apply_appr_status = pd.pivot_table(refuse_first_apply,
#                    values='apply_code',
#                    columns='apply_month',
#                    index='approval_status',
#                    aggfunc=np.size,
#                    fill_value=0,
#                    margins=True)
#refuse_apply_loan_status = pd.pivot_table(refuse_first_apply,
#                    values='apply_code',
#                    columns='apply_month',
#                    index='loan_status',
#                    aggfunc=np.size,
#                    fill_value=0,
#                    margins=True)
#refuse_apply_status = pd.concat([refuse_apply_appr_status,refuse_apply_loan_status],
#                         keys=['审批状态','放款状态'])
#
#a = pd.concat([total_status,first_apply_status,re_apply_status,refuse_apply_status],
#          keys=['总订单','新增申请用户订单','复贷申请用户订单','拒绝申请用户订单'])
#a.to_excel(r"C:\Users\lenovo\Desktop\审批放款状态.xlsx")
#
##系统原因导致的拒绝（如复贷规则没上线）
#refuse_info_1 = pd.read_sql("select apply_code, refuse_info_1 \
#                     from approval.manual_audit_result",cnx)
#re_apply_actual = pd.merge(re_apply,refuse_info_1,how='left',on='apply_code')
#re_apply_actual = re_apply_actual.loc[re_apply_actual.refuse_info_1!='system',:]
#re_apply_actual_appr_status = pd.pivot_table(re_apply_actual,
#                    values='apply_code',
#                    columns='apply_month',
#                    index='approval_status',
#                    aggfunc=np.size,
#                    fill_value=0,
#                    margins=True)
#re_apply_actual_appr_status.to_excel(r"C:\Users\lenovo\Desktop\剔除系统原因的复贷审批状态.xlsx")
#
##新增放款用户
#new_loan_apply = apply_appr_loan_info.loc[apply_appr_loan_info.loan_status=='放款成功',:]
#new_loan_apply.sort_values(['user_code','apply_time'],inplace=True)
#new_loan_apply.drop_duplicates('user_code',inplace=True)
#new_loan_user = pd.pivot_table(new_loan_apply,
#                               values='user_code',
#                               columns='apply_month',
#                               index='loan_status',
#                               aggfunc=np.size,
#                               fill_value=0,
#                               margins=True)
##-----------------------------------------------------------------------------
##用于关联的主键
#ir_no = pd.read_sql("select apply_code, user_code, id_no, id_type, phone_no \
#                from approval.borrower_info where apply_code='PL201908301606239050010'",cnx)
#linkman_cell = pd.read_sql("select user_code, contacts_tel \
#                           from approval.contact_info \
#                           where apply_code='PL201907120845470260090'",cnx)
##手机号关联的订单
#ir_cell_x = pd.read_sql("select apply_code, user_code, id_no, id_type, phone_no \
#                from approval.borrower_info where phone_no='09994546249'",cnx)
#ir_cell_x_apply = pd.read_sql("select apply_code, apply_time, approval_status, \
#                              approval_status_time \
#                              from approval.apply_info \
#                              where apply_code in \
#                                  (select apply_code \
#                                  from approval.borrower_info \
#                                  where phone_no='09959360354') \
#                             order by apply_time",cnx)
##身份证号关联的订单
#ir_id_x = pd.read_sql("select apply_code, user_code, id_no, id_type, phone_no \
#                from approval.borrower_info where id_no='N0395180100'",cnx)
#ir_id_x_apply = pd.read_sql("select a.apply_code, a.user_code, a.id_no, \
#                            a.id_type, a.phone_no, b.apply_time, \
#                            b.approval_status, b.approval_status_time \
#                            from approval.borrower_info as a \
#                            join approval.apply_info as b \
#                            on a.id_no='N0395180100' and a.apply_code=b.apply_code \
#                            order by b.apply_time",cnx)
##身份证号关联手机号
#ir_id_x_cell = pd.read_sql("select user_code, phone_no, id_no \
#                           from approval.borrower_info where id_no in \
#                               (select id_no \
#                               from approval.borrower_info \
#                               where apply_code='PL201907120845470260090')",cnx)
##相同联系人的用户ir_linkman_x_cell
#pd.read_sql("select distinct user_code from approval.contact_info \
#            where contacts_tel in ('09103883470','09366031255')",cnx)
#pd.read_sql("select distinct user_code from approval.contact_info \
#            where contacts_tel in \
#            (select contacts_tel \
#            from approval.contact_info \
#            where apply_code='PL201908301606239050010')",cnx)
##单位电话关联手机号（手机号与user_code一一对应）
#pd.read_sql("select distinct user_code from approval.employment_info \
#            where office_tel in \
#                (select office_tel from approval.employment_info \
#                 where apply_code='PL201907120845470260090')",cnx)
#%%-----urule用户特征
cnx = pymysql.connect(**cnx_args)
base_derived_variable = pd.read_sql("select * from approval.base_derived_variable",cnx)
#df = base_derived_variable.copy()
#----内部黑名单类特征
#inter_blk_x  = [
#'apply_code','user_code',
#'ir_cell_black_status',
#'ir_device_id_black_status',
#'ir_id_black_status',
#'ir_linkman_cell_black_status',
#'ir_linkman_x_device_id_black_status',
#'ir_linkman_x_id_black_status'
#]
#----手机号、设备号、身份证号在内部黑名单
#inter_blk_df = df.loc[(df.ir_cell_black_status=='1') |
#                      (df.ir_device_id_black_status=='1') |
#                      (df.ir_id_black_status=='1'),
#                      ['apply_code','user_code']]
#----被拒绝时长特征
#refuse_days_x = [
#'apply_code','user_code',
#'ir_cell_x_account_refuse_days',
#'ir_id_x_account_refuse_days',
#'ir_linkman_cell_x_account_refuse_days',
#'ir_name_x_account_refuse_days'
#]
#refuse_days_df = df.loc[:,refuse_days_x]
#----账户状态特征
#account_status_x = [
#'apply_code','user_code',
#'ir_cell_x_account_status',
#'ir_id_x_account_status',
#'ir_linkman_cell_x_account_status',
#'ir_m12_cell_x_account_status',
#'ir_m12_id_x_account_status',
#'ir_m12_linkman_cell_x_account_status',
#'ir_m1_cell_x_account_status',
#'ir_m1_id_x_account_status',
#'ir_m1_linkman_cell_x_account_status',
#'ir_m3_cell_x_account_status',
#'ir_m3_id_x_account_status',
#'ir_m3_linkman_cell_x_account_status',
#'ir_m6_cell_x_account_status',
#'ir_m6_id_x_account_status',
#'ir_m6_linkman_cell_x_account_status'
#]
#account_status_df = df.loc[:,account_status_x]
#----手机号个数特征
#cell_cnt_x = [
#'apply_code','user_code',
#'ir_id_x_cell_cnt',
#'ir_m12_home_addr_x_cell_cnt',
#'ir_m12_id_x_cell_cnt',
#'ir_m12_linkman_cell_x_cell_cnt',
#'ir_m12_name_x_cell_cnt',
#'ir_m12_tel_company_x_cell_cnt',
#'ir_m1_home_addr_x_cell_cnt',
#'ir_m1_id_x_cell_cnt',
#'ir_m1_linkman_cell_x_cell_cnt',
#'ir_m1_name_x_cell_cnt',
#'ir_m1_tel_company_x_cell_cnt',
#'ir_m3_home_addr_x_cell_cnt',
#'ir_m3_id_x_cell_cnt',
#'ir_m3_linkman_cell_x_cell_cnt',
#'ir_m3_name_x_cell_cnt',
#'ir_m3_tel_company_x_cell_cnt',
#'ir_m6_home_addr_x_cell_cnt',
#'ir_m6_id_x_cell_cnt',
#'ir_m6_linkman_cell_x_cell_cnt',
#'ir_m6_name_x_cell_cnt',
#'ir_m6_tel_company_x_cell_cnt',
#'ir_name_x_cell_cnt'        
#]
#cell_cnt_df = df.loc[:,cell_cnt_x]
#----身份证号个数特征
#id_cnt_x = [
#'apply_code','user_code',
#'ir_cell_x_id_cnt',
#'ir_m12_cell_x_id_cnt',
#'ir_m12_home_addr_x_id_cnt',
#'ir_m12_linkman_cell_x_id_cnt',
#'ir_m12_name_x_id_cnt',
#'ir_m12_tel_company_x_id_cnt',
#'ir_m1_cell_x_id_cnt',
#'ir_m1_home_addr_x_id_cnt',
#'ir_m1_linkman_cell_x_id_cnt',
#'ir_m1_name_x_id_cnt',
#'ir_m1_tel_company_x_id_cnt',
#'ir_m3_cell_x_id_cnt',
#'ir_m3_home_addr_x_id_cnt',
#'ir_m3_linkman_cell_x_id_cnt',
#'ir_m3_name_x_id_cnt',
#'ir_m3_tel_company_x_id_cnt',
#'ir_m6_cell_x_id_cnt',
#'ir_m6_home_addr_x_id_cnt',
#'ir_m6_linkman_cell_x_id_cnt',
#'ir_m6_name_x_id_cnt',
#'ir_m6_tel_company_x_id_cnt',
#'ir_name_x_id_cnt'
#]
#id_cnt_df = df.loc[:,id_cnt_x]
#----姓名个数特征
#name_cnt_x = [
#'apply_code','user_code',
#'ir_cell_x_name_cnt',
#'ir_id_x_name_cnt',
#'ir_m12_cell_x_name_cnt',
#'ir_m12_id_x_name_cnt',
#'ir_m1_cell_x_name_cnt',
#'ir_m1_id_x_name_cnt',
#'ir_m3_cell_x_name_cnt',
#'ir_m3_id_x_name_cnt',
#'ir_m6_cell_x_name_cnt',
#'ir_m6_id_x_name_cnt'        
#]
#name_cnt_df = df.loc[:,name_cnt_x]
#----邮箱个数特征
#mail_cnt_x = [
#'apply_code','user_code',
#'ir_name_x_mail_cnt'
#]
#mail_cnt_df = df.loc[:,name_cnt_x]
#----家庭地址个数特征
#home_addr_cnt_x = [
#'apply_code','user_code',
#'ir_m12_cell_x_home_addr_cnt',
#'ir_m12_id_x_home_addr_cnt',
#'ir_m12_tel_company_x_home_addr_cnt',
#'ir_m1_cell_x_home_addr_cnt',
#'ir_m1_id_x_home_addr_cnt',
#'ir_m1_tel_company_x_home_addr_cnt',
#'ir_m3_cell_x_home_addr_cnt',
#'ir_m3_id_x_home_addr_cnt',
#'ir_m3_tel_company_x_home_addr_cnt',
#'ir_m6_cell_x_home_addr_cnt',
#'ir_m6_id_x_home_addr_cnt',
#'ir_m6_tel_company_x_home_addr_cnt'
#]
#home_addr_cnt_df = df.loc[:,home_addr_cnt_x]
#----单位地址个数特征
#biz_addr_cnt_x = [
#'apply_code','user_code',
#'ir_m12_cell_x_biz_addr_cnt',
#'ir_m12_id_x_biz_addr_cnt',
#'ir_m1_cell_x_biz_addr_cnt',
#'ir_m1_id_x_biz_addr_cnt',
#'ir_m3_cell_x_biz_addr_cnt',
#'ir_m3_id_x_biz_addr_cnt',
#'ir_m6_cell_x_biz_addr_cnt',
#'ir_m6_id_x_biz_addr_cnt'
#]
#biz_addr_cnt_df = df.loc[:,biz_addr_cnt_x]
#----单位电话个数特征
#biz_tel_cnt_x = [
#'apply_code','user_code',
#'ir_m12_cell_x_tel_company_cnt',
#'ir_m12_id_x_tel_company_cnt',
#'ir_m12_linkman_cell_x_tel_company_cnt',
#'ir_m1_cell_x_tel_company_cnt',
#'ir_m1_id_x_tel_company_cnt',
#'ir_m1_linkman_cell_x_tel_company_cnt',
#'ir_m3_cell_x_tel_company_cnt',
#'ir_m3_id_x_tel_company_cnt',
#'ir_m3_linkman_cell_x_tel_company_cnt',
#'ir_m6_cell_x_tel_company_cnt',
#'ir_m6_id_x_tel_company_cnt',
#'ir_m6_linkman_cell_x_tel_company_cnt',
#'ir_m12_home_addr_x_company_home_cnt',
#'ir_m1_home_addr_x_company_home_cnt',
#'ir_m3_home_addr_x_company_home_cnt',
#'ir_m6_home_addr_x_company_home_cnt'
#]
#biz_tel_cnt_df = df.loc[:,biz_tel_cnt_x]
#
#----20190829上线的关联用到的字段
#ir_var = [
#'apply_code','user_code',
#
#'ir_id_x_cell_cnt',
#'ir_name_x_cell_cnt',
#'ir_m12_home_addr_x_cell_cnt',
#'ir_m12_linkman_cell_x_cell_cnt',
#'ir_m12_tel_company_x_cell_cnt',
#
#'ir_cell_x_id_cnt',
#'ir_name_x_id_cnt',
#'ir_m12_home_addr_x_id_cnt',
#'ir_m12_linkman_cell_x_id_cnt',
#'ir_m12_tel_company_x_id_cnt',
#
#'ir_m12_cell_x_tel_company_cnt',
#'ir_m12_id_x_tel_company_cnt',
#'ir_m12_linkman_cell_x_tel_company_cnt',
#'ir_m12_home_addr_x_company_home_cnt',
#
#'ir_m12_cell_x_biz_addr_cnt',
#'ir_m12_id_x_biz_addr_cnt',
#
#'ir_m12_cell_x_home_addr_cnt',
#'ir_m12_id_x_home_addr_cnt',
#'ir_m12_tel_company_x_home_addr_cnt',
#
#'ir_cell_x_name_cnt',
#'ir_id_x_name_cnt'
#]
#var_df = df.loc[:,ir_var]
#var_df.drop(columns=['apply_code','user_code'],inplace=True)
#----统计各变量的分布
#for i, col in enumerate(var_df.columns):
#    a =var_df.groupby(col).size()
#    a_df = pd.DataFrame(a,columns=[col])
#    if i == 0:
#        stat_df = a_df
#    else:
#        stat_df = pd.concat([stat_df,a_df],axis=1)
#stat_df.fillna("",inplace=True)
##-----------------------------------------------------------------------------
#%%----客户特征demo
#borrower_info = pd.read_sql("select apply_code, id_type, age, children_text, \
#                            province_addr_text, education, gender, \
#                            length_of_residence, marriage \
#                            from approval.borrower_info",cnx)
#employment_info = pd.read_sql("select apply_code, job_type, \
#                              monthly_net_income_text, on_the_job_time, \
#                              province_addr_text as job_province \
#                              from approval.employment_info",cnx)
#
#number_of_apply = apply_info.loc[:,['apply_code','user_code','apply_time']]
#number_of_apply['number_of_apply'] = number_of_apply['apply_time'].groupby(number_of_apply['user_code']).rank(method='dense')
#number_of_apply.sort_values(['user_code','apply_time'],inplace=True)
#number_of_apply.drop(columns=['user_code','apply_time'],inplace=True)
#
#event_info = pd.read_sql("select apps_flyer_id, device_type, install_time, \
#                         operator, os_version, media_source \
#                         from suncash_lend.af_event_info_detail",cnx)
#apps_flyer_id = pd.read_sql("select apply_code, apps_flyer_id \
#                            from suncash_lend.apply_info",cnx)
#source = pd.merge(event_info,apps_flyer_id,on='apps_flyer_id')
#
#contact_num = pd.read_sql("select apply_code, max(contact_num) as contact_num \
#                          from approval.address_book_info \
#                          group by apply_code",cnx)
#contact_num['contact_num_grp'] = pd.cut(contact_num['contact_num'],
#                                        bins=[0,10,20,50,100,10000],
#                                        labels=['1-10','11-20','21-50','51-100','101-10000'])
#
#device_info = pd.read_sql("SELECT * FROM approval.device_info_derived_variable",cnx)
#device_info.drop(columns=['id','user_code','created_time','updated_time'],inplace=True)
#
#repay_mart = pd.merge(account_info,borrower_info,how='left',on='apply_code')
#repay_mart = pd.merge(repay_mart,employment_info,how='left',on='apply_code')
#repay_mart = pd.merge(repay_mart,number_of_apply,how='left',on='apply_code')
#repay_mart = pd.merge(repay_mart,var_df,how='left',on='apply_code')
#repay_mart = pd.merge(repay_mart,source,how='left',on='apply_code')
#repay_mart = pd.merge(repay_mart,contact_num,how='left',on='apply_code')
#repay_mart.to_excel(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\repay_mart.xlsx",index=False)
#
#appr_mart = pd.merge(apply_appr_loan_info,borrower_info,how='left',on='apply_code')
#appr_mart = pd.merge(appr_mart,employment_info,how='left',on='apply_code')
#appr_mart = pd.merge(appr_mart,number_of_apply,how='left',on='apply_code')
#appr_mart = pd.merge(appr_mart,source,how='left',on='apply_code')
#appr_mart = pd.merge(appr_mart,var_df,how='left',on='apply_code')
#appr_mart = pd.merge(appr_mart,contact_num,how='left',on='apply_code')
#appr_mart = pd.merge(appr_mart,auto_audit_result,how='left',on='apply_code')
##appr_mart = pd.merge(appr_mart,manual_audit_result,how='left',on='apply_code')
#appr_mart.to_excel(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\appr_mart.xlsx",index=False)
#
#

import pandas as pd
 
data = {'地址1':['广东深圳','广东汕头','重庆'],'地址2':['揭阳','深圳','北京']}
df = pd.DataFrame(data)
a=['深圳','北京']
df = df[df.apply(lambda x: x.地址2 not in a, axis=1)]

#%%----001补充变量--------------------------------------------------------------------


try:
    os.chdir(r"F:\菲律宾日报")
except:
    os.chdir(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart")
# urule
data_it=pd.read_csv('data_urule.csv')

with open("oss_access_key_txl.json") as oss_access_key:
    oss_args = json.load(oss_access_key)
auth = oss2.Auth(oss_args['access_key_id'], oss_args['access_key_secret'])
bucket = oss2.Bucket(auth, oss_args['endpoint'], 'suncash-prd')
uruleone=pd.read_sql("select * from approval.urule_flow_context   ",cnx)
uruleone=uruleone.loc[(uruleone.apply_code.notna()) & (uruleone.package_id=='SunCash/001'),:]#005预授信还没有订单号，001有，盲猜是先走预授信再走自动拒绝,没有重复

apply_info['apply_timeaaaa']=pd.to_datetime(apply_info['apply_time']).dt.date
data=apply_info.loc[(apply_info.apply_timeaaaa>=datetime.date(2021,5,1)) & (apply_info.loan_type=='first_apply'),['apply_code']]#随便定一个日期
data['test']=1
# a=data_it.head(20)
data=pd.merge(data,data_it,how='left',left_on='apply_code',right_on='applyCode')
data=data[data.applyCode.isna()  ]
data=pd.merge(data,uruleone,how='left',on='apply_code')
data=data.loc[:,['apply_code','flow_context_key']]
data=data[data.flow_context_key.notna()  ]
os.chdir(r"F:\菲律宾日报\json")
for i in range(len(data)):
    oss_key = data.iloc[i,1]
    
    temp_file=r"F:\菲律宾日报\json\temp.json"
    bucket.get_object_to_file(oss_key,temp_file)
    with open('temp.json','r',encoding='UTF-8') as temp:
        temp_json = json.load(temp)
    try:
        # print(i)
        # temp_objects1 = pd.DataFrame(temp_json['objects'][0],index=[0])
        temp_objects2 = pd.DataFrame(temp_json['objects'][1],index=[0])
        temp_objects3 = pd.DataFrame(temp_json['objects'][2],index=[0])
        temp_objects3['applyCode']=data.iloc[i,0]
        # temp_objects=pd.merge(temp_objects1,temp_objects2,how='left',on='applyCode')
        temp_objects=pd.merge(temp_objects2,temp_objects3,how='left',on='applyCode')
    except:
            continue
    if i == 0:
        urule_v = temp_objects
    else:
        urule_v = pd.concat([urule_v,temp_objects],axis=0,sort=True)  
# urule_v.to_csv(r'F:\菲律宾日报\20220106_urule.csv')


try:
    b=datetime.date.today().strftime("%Y%m%d")
    data_download=pd.concat([data_it,urule_v],axis=0,sort=True)  
    data_download.to_csv(r'F:\菲律宾日报\\' +b+ '_urule.csv')
    data_download.to_csv(r'F:\菲律宾日报\data_urule.csv')
    data_it=data_download.loc[:,['applyCode','totalNumberOfLoans','nowHour','eqcDirecValidCellNum','gender','idType','isWhiteList','jobType','paymentType',
                       'applyCount1M', 'limitCount1M', 'firstLimitIntvl','age','education','eqcDirecCellNum']]
    data_it.rename(columns={'gender':'gender_u','idType':'idType_u','isWhiteList':'isWhiteList_u','jobType':'jobType_u','paymentType':'paymentType_u',
                            'age':'age_u','education':'education_u'},inplace=True)
    data_it.drop_duplicates(subset='applyCode',keep='last',inplace=True)
except:

    data_it=data_it.loc[:,['applyCode','totalNumberOfLoans','nowHour','eqcDirecValidCellNum','gender','idType','isWhiteList','jobType','paymentType',
                           'applyCount1M', 'limitCount1M', 'firstLimitIntvl','age','education','eqcDirecCellNum']]
    data_it.rename(columns={'gender':'gender_u','idType':'idType_u','isWhiteList':'isWhiteList_u','jobType':'jobType_u','paymentType':'paymentType_u',
                            'age':'age_u','education':'education_u'},inplace=True)
    data_it.drop_duplicates(subset='applyCode',keep='last',inplace=True)


data_it.eqcDirecValidCellNum.fillna(0,inplace=True)
data_it['eqcDirecValidCellNum'] = data_it.eqcDirecValidCellNum.map(lambda x:int(x))
data_it['vc_num'] = pd.cut(data_it.eqcDirecValidCellNum,
                                 bins=[-np.inf,9,19,29,39,49,np.inf],
                                 labels=['A1','A2','A3','A4','A5','A6'])



#%%----通讯录数据（志培版）--------------------------------------------------------------------

import json
import numpy as np
import pandas as pd
import sys
sys.path.append(r"C:\Users\lenovo\Anaconda3\Lib\site-packages")
import pymysql
#import math
import datetime
#import gc
#gc.collect() 垃圾回收，返回处理这些循环引用一共释放掉的对象个数
import os
import oss2
import zipfile
from itertools import islice
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = [u'SimHei']
plt.rcParams['axes.unicode_minus'] = False
import saspy
os.chdir(r"F:\菲律宾日报")
with open("oss_access_key_txl.json") as oss_access_key:
    oss_args = json.load(oss_access_key)
with open(r"db_config.json") as db_config:
    cnx_args = json.load(db_config)
auth = oss2.Auth(oss_args['access_key_id'], oss_args['access_key_secret'])
bucket = oss2.Bucket(auth, oss_args['endpoint'], 'suncash-prd')

cnx = pymysql.connect(**cnx_args)

#
s_contact_book_key = pd.read_sql("select user_code,substring_index(contact_book_key,';',1) as pag, \
                               substring_index(contact_book_key,';',-1) as oss_key \
                               from suncash_lend.user_contact_book where status=1",cnx)
p_contact_book_key = pd.read_sql("select user_code,substring_index(contact_book_key,';',1) as pag, \
                               substring_index(contact_book_key,';',-1) as oss_key \
                               from suncash_pautang.user_contact_book where status=1",cnx)
f_contact_book_key = pd.read_sql("select user_code,substring_index(contact_book_key,';',1) as pag, \
                               substring_index(contact_book_key,';',-1) as oss_key \
                               from flash_loan.user_contact_book where status=1",cnx)   
contact_book_key=pd.concat([s_contact_book_key,p_contact_book_key],axis=0,sort=True)
contact_book_key=pd.concat([contact_book_key,f_contact_book_key],axis=0,sort=True)
account = pd.read_sql("select user_code,contract_no   from account.account_info ",cnx)
                               
# 全量F包(之前的放在移动硬盘弄家里了，家里又用不了vpn)


# 某一次                        
os.chdir(r"H:\搬家\F\TS\PreWork\PhoneList\check")
lista = pd.read_excel('list.xlsx')
lista=lista.loc[lista.contract_no.notna(),:]
lista=pd.merge(lista,account,how='left',on='contract_no')
lista=pd.merge(lista,contact_book_key,how='left',on='user_code')


# pautang评分
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
repay_mart=sas.sd2df('repay_mart','MY')
repay_mart['last_repay_date'] = pd.to_datetime(repay_mart['last_repay_date'] ).dt.date

fa=repay_mart.loc[(repay_mart.customer_source_sys=='SuncashPautang') & (repay_mart.last_repay_date>=datetime.date(2022,2,1)) ,['user_code','last_repay_date']]
fa.sort_values(['user_code','last_repay_date'],inplace=True)#这里就决定了是新贷
fa.drop_duplicates('user_code',keep='last',inplace=True)
lista=pd.merge(fa,s_contact_book_key,how='left',on='user_code')

lista=lista[lista.oss_key.notna()  ]
os.chdir(r"H:\搬家\F\TS\PreWork\PhoneList\raw_s_0216\json")
for i in range(len(lista)):
    print(i)
    oss_key = lista.iloc[i,3]
    temp_file=lista.iloc[i,0]+'_'+str(i)+'.json'
    bucket.get_object_to_file(oss_key,temp_file)

# 抽查个人
os.chdir(r"H:\搬家\F\TS\PreWork\PhoneList\check\json_personal")

oss_key = r'contactBook/suncash-lend_202203231338282770801/2022/03/23/13/42/42ae50e4e4-4940-49cc-8aa1-ed4b75e54fba.json'
temp_file=r'eight.json'
bucket.get_object_to_file(oss_key,temp_file)
with open('eight.json','r',encoding='UTF-8') as temp:

    temp_json = json.load(temp)
temp_df = pd.DataFrame(temp_json)





alist=os.listdir(r"H:\搬家\F\TS\PreWork\PhoneList\check\json_personal")   
os.chdir(r"H:\搬家\F\TS\PreWork\PhoneList\check\json_personal")
for i in range(len(alist)):#89989那里崩了，重新用range(89990,len(alist))
    with open(alist[i],'r',encoding='UTF-8') as temp:
        print(alist[i])
        temp_json = json.load(temp)
    temp_df = pd.DataFrame(temp_json)
    try:
        temp_df=temp_df.loc[temp_df.phoneList.notna(),:]
        
        # s=pd.DataFrame({'phone':np.concatenate(temp_df.phoneList.values),'name':temp_df.name.repeat(temp_df.phoneList.str.len())})
        s=pd.DataFrame({'phone':np.concatenate(temp_df.phoneList.values)})#出现因为名字不规范导致后面导出sas有问题
        s['numb'] = i
    except:
            continue
    if i == 0:
        contact_book = s
    else:
        contact_book = pd.concat([contact_book,s],axis=0,sort=True)


# flashloan查看
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
repay_mart=sas.sd2df('repay_mart','MY')
repay_mart['last_repay_date'] = pd.to_datetime(repay_mart['last_repay_date'] ).dt.date

#获取以前下载的，避免重复下载
alist=os.listdir(r"H:\搬家\F\TS\PreWork\PhoneList\raw_f_0211\json")#方便并且账单日是20220101之后截止到0211
# alist1=os.listdir(r"H:\搬家\F\TS\PreWork\PhoneList\flash_raw")#这部分数据应该是包含拒绝的
alist_t=[]
for ele in alist:
    alist_t.append(ele[0:ele.rfind('_')])
# for ele in alist1:
#     alist_t.append(ele[0:ele.rfind('_')])
# 使用字典方法来给list内元素去重
alist_dict={}
alist_dict = alist_dict.fromkeys(alist_t)
alist_t = list(alist_dict.keys())

fa=repay_mart.loc[(repay_mart.customer_source_sys=='FlashLoan'),['user_code','last_repay_date']]
fa.sort_values(['user_code','last_repay_date'],inplace=True)#这里就决定了是新贷
fa.drop_duplicates('user_code',keep='last',inplace=True)
fa=fa[fa.apply(lambda x: x.user_code not in alist_t,axis=1)]

lista=pd.merge(fa,f_contact_book_key,how='left',on='user_code')
lista=lista[lista.oss_key.notna()  ]
os.chdir(r"H:\搬家\F\TS\PreWork\PhoneList\raw_f_0215\json")
for i in range(len(lista)):
    print(i)
    oss_key = lista.iloc[i,3]
    temp_file=lista.iloc[i,0]+'_'+str(i)+'.json'
    bucket.get_object_to_file(oss_key,temp_file)



# suncash查看1月份
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
repay_mart=sas.sd2df('repay_mart','MY')
appr_mart=sas.sd2df('appr_mart','MY')
repay_mart['last_repay_date'] = pd.to_datetime(repay_mart['last_repay_date'] ).dt.date
appr_mart['apply_time'] = pd.to_datetime(appr_mart['apply_time'] ).dt.date
alist=os.listdir(r"H:\搬家\F\TS\PreWork\PhoneList\raw_s_0214\json")#方便并且账单日是20220101之后截止到0214,放款
alist1=os.listdir(r"H:\搬家\F\TS\PreWork\PhoneList\raw_s_0216\json")#方便并且账单日是20210601-20220101，放款
alist2=os.listdir(r"H:\搬家\F\TS\PreWork\PhoneList\raw_s_0218\json")#方便并且账单日是20210601-20220101,大部分申请
alist_t=[]
for ele in alist:
    alist_t.append(ele[0:ele.rfind('_')])
# for ele in alist1:
#     alist_t.append(ele[0:ele.rfind('_')])
# 使用字典方法来给list内元素去重
alist_dict={}
alist_dict = alist_dict.fromkeys(alist_t)
alist_t = list(alist_dict.keys())

# 第二次
fa=repay_mart.loc[(repay_mart.customer_source_sys=='SunCash') & (repay_mart.last_repay_date>=datetime.date(2021,6,1)) ,['user_code','last_repay_date']]
fa.sort_values(['user_code','last_repay_date'],inplace=True)#这里就决定了是新贷
fa.drop_duplicates('user_code',keep='last',inplace=True)
fa=fa[fa.apply(lambda x: x.user_code not in alist_t,axis=1)]
lista=pd.merge(fa,s_contact_book_key,how='left',on='user_code')

lista=lista[lista.oss_key.notna()  ]
os.chdir(r"H:\搬家\F\TS\PreWork\PhoneList\raw_s_0216\json")
for i in range(len(lista)):
    print(i)
    oss_key = lista.iloc[i,3]
    temp_file=lista.iloc[i,0]+'_'+str(i)+'.json'
    bucket.get_object_to_file(oss_key,temp_file)
    
# 第三次  第三次小部分
alist_t=[]
for ele in alist:
    alist_t.append(ele[0:ele.rfind('_')])
for ele in alist1:
    alist_t.append(ele[0:ele.rfind('_')])
for ele in alist2:
    alist_t.append(ele[0:ele.rfind('_')])
# 使用字典方法来给list内元素去重
alist_dict={}
alist_dict = alist_dict.fromkeys(alist_t)
alist_t = list(alist_dict.keys())
fa=appr_mart.loc[(appr_mart.customer_source_sys=='SunCash') & (appr_mart.apply_time>=datetime.date(2021,6,1)) ,['user_code','apply_time']]
fa.sort_values(['user_code','apply_time'],inplace=True)#这里就决定了是新贷，没毛病，暂时不考虑多次申请多个通讯录问题
fa.drop_duplicates('user_code',keep='last',inplace=True)
fa=fa[fa.apply(lambda x: x.user_code not in alist_t,axis=1)]

lista=pd.merge(fa,s_contact_book_key,how='left',on='user_code')

lista=lista[lista.oss_key.notna()  ]
os.chdir(r"H:\搬家\F\TS\PreWork\PhoneList\raw_s_0218\json")
for i in range(len(lista)):
    print(i)
    oss_key = lista.iloc[i,3]
    temp_file=lista.iloc[i,0]+'_'+str(i)+'.json'
    bucket.get_object_to_file(oss_key,temp_file)

# alist=os.listdir(r"H:\搬家\F\TS\PreWork\PhoneList\test")   
# os.chdir(r"H:\搬家\F\TS\PreWork\PhoneList\test")
# with open('C202110201829548550003_363.json','r',encoding='UTF-8') as temp:
#     temp_json = json.load(temp)
# temp_df = pd.DataFrame(temp_json)
# temp_df=temp_df.loc[temp_df.phoneList.notna(),:]
# s=pd.DataFrame({'phone':np.concatenate(temp_df.phoneList.values),'name':temp_df.name.repeat(temp_df.phoneList.str.len())})
# s['user_code']='FlashLoan_202107211552328366171_19754'
contact_book_b=contact_book.copy()
alist=os.listdir(r"H:\搬家\F\TS\PreWork\PhoneList\raw_s_0218\json")   
os.chdir(r"H:\搬家\F\TS\PreWork\PhoneList\raw_s_0218\json")
for i in range(89990,len(alist)):#89989那里崩了，重新用range(89990,len(alist))
    with open(alist[i],'r',encoding='UTF-8') as temp:
        print(alist[i])
        temp_json = json.load(temp)
    temp_df = pd.DataFrame(temp_json)
    try:
        temp_df=temp_df.loc[temp_df.phoneList.notna(),:]
        
        # s=pd.DataFrame({'phone':np.concatenate(temp_df.phoneList.values),'name':temp_df.name.repeat(temp_df.phoneList.str.len())})
        s=pd.DataFrame({'phone':np.concatenate(temp_df.phoneList.values)})#出现因为名字不规范导致后面导出sas有问题
        s['user_code'] = alist[i]
    except:
            continue
    if i == 0:
        contact_book = s
    else:
        contact_book = pd.concat([contact_book,s],axis=0,sort=True)
sas=saspy.SASsession()
sas.saslib('MY "H:\搬家\F\TS\PreWork\PhoneList\check"')
sas.df2sd(contact_book.astype(str),'s_contact20220118',"MY",encode_errors='replace')#appr_mart特殊，单独搞
# contact_book_test=contact_book.copy()
# contact_book_test.drop(columns=['name'],inplace=True)
# contact_book_h1=contact_book.iloc[0:1000000].copy()#太大了导不出sas
# contact_book_h2=contact_book.iloc[1000000:2000000].copy()#太大了导不出sas
# contact_book_h2.drop(columns=['name'],inplace=True)
# contact_book_h3=contact_book.iloc[2000000:3000000].copy()#太大了导不出sas
# contact_book_h4=contact_book.iloc[3000000:4191060].copy()#太大了导不出sas
# kan=contact_book_h2.iloc[367770:367820].copy()#太大了导不出sas
# sas.df2sd(contact_book_h1.astype(str),'s_contact20220115_h1',"MY",encode_errors='replace')#appr_mart特殊，单独搞
# sas.df2sd(contact_book_h2.astype(str),'s_contact20220115_h2',"MY",encode_errors='replace')#appr_mart特殊，单独搞
# sas.df2sd(contact_book_h3.astype(str),'s_contact20220115_h3',"MY",encode_errors='replace')#appr_mart特殊，单独搞
# sas.df2sd(contact_book_h4.astype(str),'s_contact20220115_h4',"MY",encode_errors='replace')#appr_mart特殊，单独搞

sas.df2sd(contact_book.astype(str),'s_contact20220118',"MY",encode_errors='replace')#appr_mart特殊，单独搞


#oss_key = 'contactBook/2019/09/08/15/35/57d23bbca6-9761-42fa-b6a4-86a24580a97e.json'
#bucket.get_object_to_file(oss_key,temp_file)
with open(temp_file,'r') as temp:
    temp_json = json.load(temp)
temp_df = pd.DataFrame(temp_json)
s=pd.DataFrame({'phone':np.concatenate(temp_df.phoneList.values)},index=temp_df.index.repeat(temp_df.phoneList.str.len()))
#ss = s.join(temp_df.drop('phoneList',1),how='left')
sss = temp_df.reindex(temp_df.index.repeat(temp_df.phoneList.str.len())).assign(phone=np.concatenate(temp_df.phoneList.values))



#%%----通讯录数据（登锋版）--------------------------------------------------------------------
#import json
#import oss2
#from itertools import islice
with open("oss_config.json") as oss_config:
   oss_args = json.load(oss_config)
auth = oss2.Auth(oss_args['suncash-prd']['access_key_id'], oss_args['suncash-prd']['access_key_secret'])
bucket = oss2.Bucket(auth, oss_args['suncash-prd']['endpoint'], 'suncash-prd')
for b in islice(oss2.ObjectIterator(bucket), 10):
    print(b.key)
#
contact_book_key = pd.read_sql("select user_code, \
                               substring_index(contact_book_key,';',-1) as oss_key \
                               from suncash_lend.user_contact_book limit 10",cnx)
temp_file = r'C:\Users\lenovo\Desktop\contact_book.json'
for i in range(len(contact_book_key)):
    oss_key = contact_book_key.iloc[i,1]
    bucket.get_object_to_file(oss_key,temp_file)
    with open(temp_file,'r') as temp:
        temp_json = json.load(temp)
    temp_df = pd.DataFrame(temp_json)
    temp_df['user_code'] = contact_book_key.iloc[i,0]
    if i == 0:
        contact_book = temp_df
    else:
        contact_book = pd.concat([contact_book,temp_df],axis=0,sort=True)
#
#oss_key = 'contactBook/2019/09/08/15/35/57d23bbca6-9761-42fa-b6a4-86a24580a97e.json'
#bucket.get_object_to_file(oss_key,temp_file)
with open(temp_file,'r') as temp:
    temp_json = json.load(temp)
temp_df = pd.DataFrame(temp_json)
s=pd.DataFrame({'phone':np.concatenate(temp_df.phoneList.values)},index=temp_df.index.repeat(temp_df.phoneList.str.len()))
#ss = s.join(temp_df.drop('phoneList',1),how='left')
sss = temp_df.reindex(temp_df.index.repeat(temp_df.phoneList.str.len())).assign(phone=np.concatenate(temp_df.phoneList.values))



#%%----applist数据（志培版） ——————————————————————————————————————————————————————————————————————————————————

import os
import oss2
import pymysql
import json
import pandas as pd
import time
from itertools import islice
import datetime
os.chdir(r"H:\搬家\e\guan\菲律宾")
with open("oss_access_key_txl.json") as oss_access_key:
    oss_args = json.load(oss_access_key)

auth = oss2.Auth(oss_args['access_key_id'], oss_args['access_key_secret'])
bucket = oss2.Bucket(auth, oss_args['endpoint'], 'suncash-prd')
with open(r"db_config.json") as db_config:
    cnx_args = json.load(db_config)
cnx = pymysql.connect(**cnx_args)   




                     
#就之前下载的数据基础上再下载                           
temp_list=os.listdir(r'H:\搬家\F\TS\PreWork\app\raw_data')
s=pd.DataFrame({'appList':temp_list})
s['appLista']=s['appList'].apply(lambda x:x[0:x.rfind('_')])
temp_file = r'H:\搬家\F\TS\PreWork\app\raw_data'
app_key=pd.merge(app_key,s,how='left',left_on='apply_code',right_on='appLista')
app_key=app_key[app_key['appList'].isna()]

code=app_key['oss_key']
# 下载
for (i,f) in enumerate(code):
    # oss_key = app_key.iloc[i,2]
    bucket.get_object_to_file(f,temp_file+'\\'+app_key.iloc[i,1]+'_'+str(i)+'.json')
    
# 解析——————————————————————————————————————————————————————————————————————————————————
os.chdir(r"F:\TS\PreWork\app\raw_data")
temp_list=os.listdir(r'F:\TS\PreWork\app\raw_data')
for i in range(len(temp_list)):
    with open(temp_list[i],'r',encoding='UTF-8') as temp:
        
        
        try:
            temp_json = json.load(temp)
            temp_df = pd.DataFrame(temp_json)
            # s=pd.DataFrame({'phone':np.concatenate(temp_df.phoneList.values)},index=temp_df.index.repeat(temp_df.phoneList.str.len()))
            temp_df['apply_code'] = temp_list[i].split(".")[0]
            if i == 0:
                app_cum = temp_df
            else:
                print(i)
                app_cum = pd.concat([app_cum,temp_df],axis=0,sort=True)
        except Exception as e:
            pass
        continue
import saspy
import numpy as np
sas=saspy.SASsession()

sas.saslib('MY "F:\TS\PreWork\\app"')

sas.df2sd(app_cum,"app_cum","MY",encode_errors='replace') 

app_cum.sort_values(by=['apply_code'],inplace=True)
app_cum.drop_duplicates(subset=['apply_code'],keep='first',inplace=True)


           
#个案
aa=data[data.code=='PL202205050904292010093']

os.chdir(r"F:\菲律宾日报\json")
oss_key=r'appList/suncash-lend_202205050854469272638//2022/05/05/08/54/488a1446b7-374e-431b-b182-0dd99d776422.json'
temp_file=r"F:\菲律宾日报\json\temp.json"
bucket.get_object_to_file(oss_key,temp_file)
with open('temp.json','r',encoding='UTF-8') as temp:
    temp_json = json.load(temp)
    temp_df = pd.DataFrame(temp_json)

    

# 批量下载suncash检视数据缺失问题20220415:确实缺失了
# app_cum01to03.csv：(不含有suncash-lend) & (app_key.created_day>=datetime.date(2022,1,1)) & (app_key.created_day<=datetime.date(2022,3,31)
import saspy
import numpy as np
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')

appr_mart=sas.sd2df('appr_mart','MY')
right_appr=appr_mart[['apply_code','apply_time','user_code']].rename(columns={'apply_code':'code'})
right_appr['apply_day']=pd.to_datetime(right_appr['apply_time']).dt.date
right_appr.sort_values(by=['user_code','apply_day','apply_time'],inplace=True)
right_appr.drop_duplicates(subset=['user_code','apply_day'],keep='first',inplace=True)


app_key = pd.read_sql("select user_code,apply_code, \
                               oss_key,created_time \
                               FROM approval.file_key where  TYPE =1",cnx)
app_key['test']=app_key.user_code.apply(lambda x:x.rfind('suncash-lend'))  #0表示找到的位置，-1是没找到
app_key['created_day']=pd.to_datetime(app_key.created_time).dt.date
data=app_key.loc[(app_key.test==0) & (app_key.created_day>=datetime.date(2022,5,1)) & (app_key.created_day<=datetime.date(2022,5,9)),['user_code','oss_key','apply_code','created_day','created_time']]
data.sort_values(by=['user_code','created_day','created_time'],inplace=True)
data.drop_duplicates(subset=['user_code','created_day'],keep='first',inplace=True)

data=pd.merge(data,right_appr,how='left',left_on=['user_code','created_day'],right_on=['user_code','apply_day'])
data.sort_values(by=['user_code','created_day'],inplace=True)


data=data[data.code.notna()  ]
data.drop_duplicates(subset='code',keep='first',inplace=True)


os.chdir(r"F:\菲律宾日报\json")
for i in range(len(data)):
# for i in range(45000,66908):#5000一个循环是阈值？
    print(i)

    oss_key = data.iloc[i,1]
    
    temp_file=r"F:\菲律宾日报\json\temp.json"
    bucket.get_object_to_file(oss_key,temp_file)
    with open('temp.json','r',encoding='UTF-8') as temp:
        try:
            temp_json = json.load(temp)
            temp_df = pd.DataFrame(temp_json)
            # s=pd.DataFrame({'phone':np.concatenate(temp_df.phoneList.values)},index=temp_df.index.repeat(temp_df.phoneList.str.len()))
            temp_df['user_code'] = data.iloc[i,0]
            temp_df['apply_code'] = data.iloc[i,5]
            temp_df['apply_day'] = data.iloc[i,7]
            temp_df['apply_time'] = data.iloc[i,6]
            if i == 0:
                app_cumr = temp_df
            else:

                app_cumr = pd.concat([app_cumr,temp_df],axis=0,sort=True)
        except Exception as e:
            pass
        continue
app_cum=app_cumr.copy(deep=True)
app_cum['aa']=app_cum['firstInstallTime'].apply(lambda x:int(round(int(x))) if x==x else 0)
app_cum['createdate']=app_cum['aa'].apply(lambda x:time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(x/1000)) if len(str(x))==13 else np.nan)
app_cum['aa']=app_cum['lastUpdateTime'].apply(lambda x:int(round(int(x))) if x==x else 0)
app_cum['updatedate']=app_cum['aa'].apply(lambda x:time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(x/1000)) if len(str(x))==13 else np.nan)
app_cum.drop(columns=['firstInstallTime','aa','lastUpdateTime'],inplace=True)
# app_cum['apply_day']=pd.to_datetime(app_cum['apply_day']).dt.date
app_cum.apply_day.fillna(datetime.date(2050,1,1),inplace=True)#随便给很远的未来一个日子
app_cum['createday']=pd.to_datetime(app_cum['createdate']).dt.date
app_cum.createday.fillna(datetime.date(2010,1,1),inplace=True)#随便给很远的以前一个日子
app_cum['in7d']=app_cum.apply(lambda x: 1 if (x.apply_day-x.createday).days<=7  else 0 ,axis=1)
app_cum['in10d']=app_cum.apply(lambda x: 1 if (x.apply_day-x.createday).days<=10  else 0 ,axis=1)
app_cum['in15d']=app_cum.apply(lambda x: 1 if (x.apply_day-x.createday).days<=15  else 0 ,axis=1)

app_cum.drop(columns=['packageName','versionName'],inplace=True)

app_cum.to_csv(r'F:\菲律宾日报\app_sn_cum50109.csv',index=False)


# 检视
app_cum['apply_code'].nunique()
a=app_cum['apply_code'].value_counts()

# 分析
app_cum['appName']=app_cum.appName.apply(lambda x:str(x).replace(' ', '').lower())
app_cum=app_cum[['appName','usercode','packageName']]
a=app_cum.appName.value_counts().reset_index()
a['cash']=a['index'].apply(lambda x:x.rfind('cash')) 
a['loan']=a['index'].apply(lambda x:x.rfind('loan')) 
b=a.loc[(a.cash>0) | (a.loan>0),:]
listb=b['index']
def func(x):
    if x in list(listb):
        return 1
    else:
        return 0
app_cum['aim']=app_cum.appName.apply(func)
app_cum['appcount']=1
# c=app_cum.loc[app_cum.aim==1,:]

appl = app_cum.groupby(by=['usercode']).agg({'aim':'sum','appcount':'count'}).reset_index()
appl['rate']=appl['aim']/appl['appcount']

sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
repay_mart=sas.sd2df('repay_mart','MY')

repay_mart['自然逾期']=repay_mart.自然逾期.apply(lambda x:int(x))
repay = repay_mart.groupby(by=['user_code']).agg({'自然逾期':'sum','contract_no':'count'})
appl=pd.merge(appl,repay,how='left',left_on='usercode',right_on='user_code')
appl.to_excel(r'F:\菲律宾日报\app分析.xlsx')

import toad
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
app_ana3=sas.sd2df('acc3','MY')
c = toad.transform.Combiner()
binsa=appl.copy(deep=True)
binsa['TARGET_rate']=binsa['自然逾期']/binsa['contract_no']
binsa=binsa.loc[(binsa.contract_no>=1) ,:]
binsa['TARGET']=binsa.TARGET_rate.apply(lambda x: 1 if x>=0.3 else 0) #权且当做第三笔不还：1/3



import toad
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
app_ana3=sas.sd2df('acc3','MY')


a=app_ana3[['target','numloan','appcount','in7d_rate','in10d_rate','in15d_rate','all_rate']]
# 使使用稳定的卡方分箱，规定每箱至少有5%数据, 空值将自动被归到最佳箱。
c.fit(a,  y = 'target',  method = 'chi',  min_samples = 0.05)
## 展示分箱结果
for col in a.columns.drop('target'):
    print(col + ' : ' , c.export()[col])



import toad
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
app_ana3=sas.sd2df('app_ana3','MY')
c = toad.transform.Combiner()

a=app_ana3[['target','numloan','numapp','rate']]
# 使使用稳定的卡方分箱，规定每箱至少有5%数据, 空值将自动被归到最佳箱。
c.fit(a,  y = 'target',  method = 'chi',  min_samples = 0.05)
## 展示分箱结果
for col in a.columns.drop('target'):
    print(col + ' : ' , c.export()[col])







js=app_cum.loc[app_cum.usercode=='FlashLoan_202108070049044819392',:]




data.to_excel(r'F:\菲律宾日报\json\test.xlsx')
#%%----applist数据（登锋版）------------------------------------------------------------------
applist_key = pd.read_sql("select apply_code,user_code,oss_key \
                          from approval.file_key where type = 1",cnx)
applist_key.drop_duplicates(subset='apply_code',inplace=True)
temp_file = r'applist.json'
for i in range(len(applist_key)):
    applist_k = applist_key.iloc[i,2]
    bucket.get_object_to_file(applist_k,temp_file)
    with open(temp_file,'r',encoding='utf8') as temp:
        temp_json = json.load(temp)
    temp_df = pd.DataFrame(temp_json)
    temp_df['apply_code'] = applist_key.iloc[i,0]
    if i == 0:
        applist = temp_df
    else:
        applist = pd.concat([applist,temp_df],axis=0,sort=True)


#%%---设备指纹数据--------------------------------------------------------------------
import json
import oss2

with open("oss_access_key.json") as oss_access_key:
   oss_args = json.load(oss_access_key)
auth = oss2.Auth(oss_args['access_key_id'], oss_args['access_key_secret'])
bucket = oss2.Bucket(auth, oss_args['endpoint'], 'suncash-prd')

device_key = pd.read_sql("select apply_code, file_key \
                         from suncash_data.td_task_info \
                         where type='1' and created_time between '2019-12-01' and '2020-01-03'",cnx)
file_dir = r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\extData\device" 
#已下载
address_detect_pre = pd.read_csv(file_dir+'\\'+'address_detect.csv')
address_detect_pre['download'] = 'Y'
device_info_pre = pd.read_csv(file_dir+'\\'+'device_info.csv')
antifraud_pre = pd.read_csv(file_dir+'\\'+'antifraud.csv',low_memory=False)
#未下载
device_key = pd.merge(device_key,address_detect_pre,how='left',on='apply_code')
device_key = device_key.loc[device_key.download!='Y',:]

for i in range(len(device_key)):
    oss_key = device_key.iloc[i,1]
#    local_file = file_dir + '\\' + device_key.iloc[i,0]
    local_file = file_dir + '\\td_result.json'
    bucket.get_object_to_file(oss_key,local_file)
    with open(local_file,'rb') as temp:
        temp_json = json.load(temp)
    try:
        temp_address_detect = pd.DataFrame([temp_json['result_desc']['INFOANALYSIS']['address_detect']])
        temp_address_detect['apply_code'] = device_key.iloc[i,0]
        temp_device_info = pd.DataFrame([temp_json['result_desc']['INFOANALYSIS']['device_info']])
        temp_device_info['apply_code'] = device_key.iloc[i,0]
        
        temp_antifraud = pd.DataFrame([temp_json['result_desc']['ANTIFRAUD']])
        temp_antifraud['apply_code'] = device_key.iloc[i,0]
        if i == 0:
            address_detect = temp_address_detect
            device_info = temp_device_info
            antifraud = temp_antifraud
        else:
            address_detect = pd.concat([address_detect,temp_address_detect])
            device_info = pd.concat([device_info,temp_device_info])
            antifraud = pd.concat([antifraud,temp_antifraud])
    except KeyError as e:
        print("except:",e)
    except Exception as e:
        print("Error:",e)


address_detect['download'] = 'Y'
address_detect = pd.concat([address_detect_pre,address_detect])
device_info = pd.concat([device_info_pre,device_info])
antifraud = pd.concat([antifraud_pre,antifraud])

address_detect.to_csv(file_dir+'\\'+'address_detect.csv',index=False)
device_info.to_csv(file_dir+'\\'+'device_info.csv',index=False)
antifraud.to_csv(file_dir+'\\'+'antifraud.csv',index=False)

address_detect = pd.read_csv(file_dir+'\\'+'address_detect.csv')
device_info = pd.read_csv(file_dir+'\\'+'device_info.csv')

drop_cols = ['androidId','apkMD5','apkVersion','appOs','basebandVersion',
             'batteryTemp','blueMac','brightness','cpuFrequency','cpuHardware',
             'cpuType','display','dnsAddress','fontHash','fpVersion','gateway',
             'hardware','host','kernelVersion','os','packageName','releaseVersion',
             'screenRes','sdkMD5','sdkVersion','serialNo','signMD5','tags',
             'tokenId','vpnNetmask','wifiIp','wifiNetmask','deviceSVN','error']
device_info.drop(columns=drop_cols,inplace=True)

address_detect.drop_duplicates('apply_code',inplace=True)
device_info.drop_duplicates('apply_code',inplace=True)
device_info['batteryLevel_grp'] = pd.cut(device_info.batteryLevel,
                                         bins=[-1,10,20,30,50,80,np.inf],
                                         labels=['0-10%','11%-20%','21%-30%','31%-50%','51%-80%','81%-'])
device_info['memory_avail_rate'] = device_info['availableMemory'] / device_info['totalMemory']
device_info['storage_avail_rate'] = device_info['availableStorage'] / device_info['totalStorage']
device_info['memory_avail_rate_grp'] = pd.cut(device_info.memory_avail_rate,
                                               bins=[0,0.7,0.8,0.9,np.inf],
                                               labels=['0-70%','71%-80%','81%-90%','91%-'])
device_info['storage_avail_rate_grp'] = pd.cut(device_info.storage_avail_rate,
                                               bins=[0,0.7,0.8,0.9,np.inf],
                                               labels=['0-70%','71%-80%','81%-90%','91%-'])
device_info['boot_time'] = pd.to_datetime(device_info['bootTime'],unit='ms')
device_info['boot_hour'] = device_info['boot_time'].dt.strftime('%H')

repay_mart = pd.merge(repay_mart,address_detect,how='left',on='apply_code')
repay_mart = pd.merge(repay_mart,device_info,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,address_detect,how='left',on='apply_code')
appr_mart = pd.merge(appr_mart,device_info,how='left',on='apply_code')
        
        
#%%---GPS数据-----------------------------------------------------------------------
#gps = pd.read_csv(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\extData\gps.csv")
#gps_max = gps.groupby('applyCode',as_index=False).agg(np.max)
#gps_max.drop(columns=['eventName','Unnamed: 4'],inplace=True)
#gps_max.rename(columns={'homeDistance':'homeDistanceMax', 'companyDistance':'companyDistanceMax'},inplace=True)
#gps_min = gps.groupby('applyCode',as_index=False).agg(np.min)
#gps_min.drop(columns=['eventName','Unnamed: 4'],inplace=True)
#gps_min.rename(columns={'homeDistance':'homeDistanceMin', 'companyDistance':'companyDistanceMin'},inplace=True)
#gps_stat = pd.merge(gps_max,gps_min,how='left',on='applyCode')
#gps_stat['homeMove'] = gps_stat['homeDistanceMax'] - gps_stat['homeDistanceMin']
#gps_stat['companyMove'] = gps_stat['companyDistanceMax'] - gps_stat['companyDistanceMin']
#gps_stat.rename(columns={'applyCode':'apply_code'},inplace=True)
#gps_mart = pd.merge(account_info,gps_stat,how='left',on='apply_code')
#
#def distance_grp(x):
#    if math.isnan(x):
#        return ''
#    elif x<100:
#        return '[0-100)'
#    elif x<200:
#        return '[100-200)'
#    elif x<500:
#        return '[200-500)'
#    elif x<1000:
#        return '[500,1000)'
#    elif x<2000:
#        return '[1000,2000)'
#    elif x<3000:
#        return '[2000,3000)'
#    elif x<5000:
#        return '[3000,5000)'
#    elif x<10000:
#        return '[5km-10km)'
#    elif x<20000:
#        return '[10km-20km)'
#    elif x<50000:
#        return '[20km-50km)'
#    else:
#        return '[50km- )'
#    
#gps_mart['homeDisMinGrp'] = gps_mart['homeDistanceMin'].apply(distance_grp)
#gps_mart['comDisMinGrp'] = gps_mart['companyDistanceMin'].apply(distance_grp)
#
##申请时间
#apply_time = pd.read_sql("select user_code, apply_code, apply_time, \
#                         date_format(apply_time,'%H') as apply_hour \
#                         from approval.apply_info",cnx)
#apply_time['apply_hour_grp'] = pd.cut(apply_time.apply_hour.astype(int),
#                                     bins=[-1,5,8,17,22,100],
#                                     labels=['0-5','6-8','9-17','18-22','23-0'])
#
##各信息项填写时间
#contact_time = pd.read_sql("select user_code, created_time \
#                                from suncash_lend.contact_info",cnx)
#employ_time = pd.read_sql("select user_code, created_time \
#                              from suncash_lend.employment_info",cnx)
#identification_time = pd.read_sql("select user_code, created_time \
#                                from suncash_lend.identification",cnx)
#personal_time = pd.read_sql("select user_code, created_time \
#                                from suncash_lend.personal_info",cnx)
#verification_time = pd.read_sql("select user_code, created_time \
#                                from suncash_lend.user_verification_info",cnx)
#info_time = pd.concat([contact_time,employ_time,identification_time,
#                       personal_time,verification_time],axis=0)
#info_time.sort_values(by=['user_code','created_time'],inplace=True)
#info_time.drop_duplicates(subset='user_code',inplace=True)
#
#time_info = pd.merge(apply_time,info_time,how='left',on='user_code')
#time_info['interval'] = (time_info['apply_time'] - time_info['created_time']).dt.seconds/60
#time_info.drop(['user_code','apply_time','created_time'],axis=1,inplace=True)
#
#gps_mart = pd.merge(gps_mart,time_info,how='left',on='apply_code')
#gps_mart.to_excel(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\extData\gps_mart.xlsx",index=False)
#    
##KPI数据-----------------------------------------------------------------------
##服务费
#service_fee = pd.read_sql("select apply_code, contract_no, \
#                          service_fee_amount, payment_service_fee \
#                          from approval.contract_info", cnx)
#acc = pd.read_sql("select contract_no, \
#                  date_format(loan_date,'%Y%m%d') as loan_day, \
#                  date_format(loan_date,'%Y%m') as loan_month \
#                  from account.account_info \
#                  where loan_date>='2019-09-01 00:00:00'",cnx)
#service_fee = pd.merge(acc,service_fee,how='left',on='contract_no')
##日息、罚息、违约服务费
#fee_sql = """select sum(case when fee_code='INTEREST' then curr_receipt_amt else 0 end) as interest,
#                    sum(case when fee_code='OVERDUE_PENALTY' then curr_receipt_amt else 0 end) as penalty,
#                    sum(case when fee_code='OVERDUE_SERVICE_FEE' then curr_receipt_amt else 0 end) as od_service_fee
#            from account.bill_fee_dtl
#            where offset_date between '2019-09-01' and '2019-09-22'"""
#other_fee = pd.read_sql(fee_sql,cnx)
#red_sql = """select sum(case when fee_type='OVERDUE_PENALTY' then reduction_amount else 0 end) as penalty_red,
#                    sum(case when fee_type='OVERDUE_SERVICE_FEE' then reduction_amount else 0 end) as od_service_fee_red
#            from account.fee_reduction
#            where reduction_date between '2019-09-01' and '2019-09-22'"""
#fee_reduction = pd.read_sql(red_sql,cnx)
#
#
##每个客户放出去的金额
#issue_amt = pd.read_sql("select user_code, sum(loan_amount) as issue_amt \
#                        from account.account_info \
#                        group by user_code",cnx)
#fee_dtl = pd.read_sql("select a.contract_no, a.curr_receipt_amt, a.fee_code, \
#                      b.user_code \
#                      from account.bill_fee_dtl as a join \
#                      account.account_info as b \
#                      on a.contract_no = b.contract_no",cnx)
#receipt_amt = fee_dtl.groupby(['user_code','fee_code'],as_index=False).agg(np.sum)
#receipt_amt = fee_dtl.groupby('user_code',as_index=False).agg(np.sum)
#issue_receipt_amt = pd.merge(issue_amt,receipt_amt,how='left',on='user_code')
#number_of_loans = account_info.loc[:,['user_code','number_of_loans','account_status']]
#number_of_loans.sort_values(['user_code','number_of_loans'],inplace=True)
#number_of_loans.drop_duplicates('user_code',keep='last',inplace=True)
#issue_receipt_amt = pd.merge(issue_receipt_amt,number_of_loans,how='left',on='user_code')
#issue_receipt_amt['profit'] = issue_receipt_amt['curr_receipt_amt'] - issue_receipt_amt['issue_amt']
#issue_receipt_amt.to_excel(r"C:\Users\lenovo\Desktop\a.xlsx",index=False)
#

#%%----推广地的申请-------------------------------------------------------------------
home_district = pd.read_sql("select apply_code, \
                            province_addr_text as home_province, \
                            city_addr_text as home_city, \
                            district_addr_text as home_district\
                            from approval.borrower_info \
                            where created_time>='2019-10-20' and \
                            district_addr_text in ('Sampaloc','Santa Cruz')",cnx)
job_district = pd.read_sql("select apply_code, \
                           province_addr_text as job_province, \
                           city_addr_text as job_city, \
                           district_addr_text as job_district\
                           from approval.employment_info \
                           where created_time>='2019-10-20' and \
                           district_addr_text in ('Sampaloc','Santa Cruz')",cnx)
promote = pd.merge(home_district,job_district,how='outer',on='apply_code')
apply_result = apply_appr_loan_info.loc[:,['apply_code','apply_day','approval_status']]
promote = pd.merge(promote,apply_result,how='left',on='apply_code')
promote.to_excel(r"C:\Users\lenovo\Desktop\promote.xlsx",index=False)

#%%---相同证件因用新手机号被作为新贷申请的订单--------------------------------------------
#有放款的客户
acc = pd.read_sql("select id_number as id_no, loan_date \
                  from account.account_info",cnx)
acc.sort_values(by=['id_no','loan_date'],inplace=True)
acc.drop_duplicates(subset=['id_no'],inplace=True)
#新贷订单
first_apply = pd.read_sql("select apply_code, apply_time \
                          from approval.apply_info \
                          where loan_type='first_apply'",cnx)
id_phone_no = pd.read_sql("select apply_code, id_no, phone_no \
                          from approval.borrower_info",cnx)
first_apply = pd.merge(first_apply,id_phone_no,how='left',on='apply_code')
first_re_apply = pd.merge(first_apply,acc,how='left',on='id_no')
re_apply = first_re_apply.loc[(first_re_apply.loan_date.isnull()==False) &
                              (first_re_apply.loan_date<first_re_apply.apply_time),:]
re_apply.sort_values(by=['id_no','apply_time'],inplace=True)

#%%--TS1001规则命中情况-------------------------------------------------------------
ts1001 = pd.read_sql("select apply_code, warning_type \
                     from approval.warning_information \
                     where warning_type='TS1001'",cnx)
reapply = pd.read_sql("select apply_code, \
                      date_format(apply_time,'%Y%m%d') as apply_date \
                      from approval.apply_info \
                      where loan_type='re_apply' and \
                            created_time>='2019-10-26'",cnx)
reapply = pd.merge(reapply,ts1001,how='left',on='apply_code')
ir_cell_x_id_cnt = pd.read_sql("select apply_code, ir_cell_x_id_cnt, \
                               ir_m1_cell_x_id_cnt, ir_m3_cell_x_id_cnt \
                               from approval.base_derived_variable \
                               where created_time>='2019-10-26'",cnx)
reapply = pd.merge(reapply,ir_cell_x_id_cnt,how='left',on='apply_code')
eqc_cell_gid_num = pd.read_sql("select apply_code, eqc_cell_gid_num \
                               from approval.device_info_derived_variable \
                               where created_time>='2019-10-26'",cnx)
reapply = pd.merge(reapply,eqc_cell_gid_num,how='left',on='apply_code')

#%%--复贷前后信息变更---------------------------------------------------------------
#单项信息
check_info_list = ['id_no','education','gender','children_text','marriage',
                   'home_province','home_city','home_district',
                   'length_of_residence','device_id','job_type',
                   'monthly_net_income_text','on_the_job_time',
                   'job_province','job_city','job_district']
def check_consistent(a,b):
    if a == b:
        return 'Y'
    else:
        return 'N'

repay_mart.sort_values(by=['user_code','number_of_loans'],inplace=True)
for check_info in check_info_list:
    repay_mart['last_'+check_info] = repay_mart[check_info].shift(1)
    repay_mart.loc[repay_mart.number_of_loans==1,'last_'+check_info] = ''

for check_info in check_info_list:
    repay_mart['consistent_'+check_info] = repay_mart.apply(lambda x:check_consistent(x[check_info],x['last_'+check_info]),axis=1)

appr_mart.sort_values(by=['user_code','number_of_apply'],inplace=True)
for check_info in check_info_list:
    appr_mart['last_'+check_info] = appr_mart[check_info].shift(1)
    appr_mart.loc[appr_mart.number_of_apply==1,'last_'+check_info] = ''
    appr_mart['consistent_'+check_info] = appr_mart.apply(lambda x:check_consistent(x[check_info],x['last_'+check_info]),axis=1)

#-------多项信息——联系人信息
linkman = pd.read_sql("select apply_code, \
                      concat(contacts_name,'_',contacts_relation,'_', \
                             replace(contacts_tel,' ','')) as linkman \
                      from approval.contact_info order by apply_code",cnx)
#                      where apply_code in ('PL201906290904550000014','PL201908310838287150036','PL201906101447484080081')" ,cnx)    
last_apply_code = ''
linkman_df = pd.DataFrame()
for i in range(len(linkman)):
    apply_code = linkman.loc[i,'apply_code']
    linkman_info_tmp = [linkman.loc[i,'linkman']]
    if apply_code == last_apply_code:
        linkman_info = linkman_info + linkman_info_tmp
    else:
        if last_apply_code != '':
            tmp = pd.DataFrame({'apply_code':last_apply_code,'linkman_info':[linkman_info]},columns=['apply_code','linkman_info'])
            linkman_df = pd.concat([linkman_df,tmp])
        linkman_info = linkman_info_tmp
    last_apply_code = apply_code
tmp = pd.DataFrame({'apply_code':last_apply_code,'linkman_info':[linkman_info]},columns=['apply_code','linkman_info'])
linkman_df = pd.concat([linkman_df,tmp])

def check_linkman_consistent(a,b):
    if set(a)^set(b) == set():
        return 'Y'
    else:
        return 'N'

repay_mart = pd.merge(repay_mart,linkman_df,how='left',on='apply_code')
repay_mart.sort_values(by=['user_code','number_of_loans'],inplace=True)
repay_mart['last_linkman_info'] = repay_mart['linkman_info'].shift(1)
repay_mart.loc[repay_mart.number_of_loans==1,'last_linkman_info'] = ''
repay_mart['consistent_linkman_info'] = repay_mart.apply(lambda x:check_linkman_consistent(x['linkman_info'],x['last_linkman_info']),axis=1)

appr_mart = pd.merge(appr_mart,linkman_df,how='left',on='apply_code')
appr_mart._values(by=['user_code','number_of_apply'],inplace=True)
appr_mart['last_linkman_info'] = appr_mart['linkman_info'].shift(1)
appr_mart.loc[appr_mart.number_of_apply==1,'last_linkman_info'] = ''
appr_mart['consistent_linkman_info']sort = appr_mart.apply(lambda x:check_linkman_consistent(x['linkman_info'],x['last_linkman_info']),axis=1)
        

#contact = pd.read_sql("select apply_code, contacts_name, contacts_relation, contacts_tel \
#                      from approval.contact_info \
#                      where apply_code in ('PL201906290904550000014','PL201908310838287150036','PL201906101447484080081')" ,cnx)    
#last_apply_code = ''
#df = pd.DataFrame()
#for i in range(len(contact)):
#    apply_code = contact.loc[i,'apply_code']
#    contact_info_tmp = [{contact.loc[i,'contacts_name']:contact.loc[i,'contacts_tel']}]
#    if apply_code == last_apply_code:
#        contact_info = contact_info + contact_info_tmp
#    else:
#        if last_apply_code != '':
#            tmp = pd.DataFrame({'apply_code':last_apply_code,'contact_info':[contact_info]},columns=['apply_code','contact_info'])
#            df = pd.concat([df,tmp])
#        contact_info = contact_info_tmp
#    last_apply_code = apply_code
#tmp = pd.DataFrame({'apply_code':last_apply_code,'contact_info':[contact_info]},columns=['apply_code','contact_info'])
#df = pd.concat([df,tmp])
#
#
#last_apply_code = ''
#df = pd.DataFrame()
#for i in range(len(contact)):
#    apply_code = contact.loc[i,'apply_code']
#    contact_info_tmp = {contact.loc[i,'contacts_name']:contact.loc[i,'contacts_tel']}
#    if apply_code == last_apply_code:
#        contact_info = dict(contact_info,**contact_info_tmp)
#    else:
#        if last_apply_code != '':
#            tmp = pd.DataFrame({'apply_code':last_apply_code,'contact_info':[contact_info]},columns=['apply_code','contact_info'])
#            df = pd.concat([df,tmp])
#        contact_info = contact_info_tmp
#    last_apply_code = apply_code
#tmp = pd.DataFrame({'apply_code':last_apply_code,'contact_info':[contact_info]},columns=['apply_code','contact_info'])
#df = pd.concat([df,tmp])


#%%--首贷模型-----------------------------------------------------------------------
from sklearn.feature_extraction import DictVectorizer
from sklearn.model_selection import train_test_split
from sklearn.metrics import classification_report

#----筛选出首贷合同里的好、坏样本
data = repay_mart.loc[(repay_mart.number_of_loans==1) &
                      ((repay_mart.od_days_ever>7) |
                       ((repay_mart.od_days_ever<=3) & 
                        (repay_mart.account_status=='ACCOUNT_SETTLE')
                       )
                      ),:]
data = data.loc[data.monthly_net_income_text!='Sibling',:]
data = data.loc[data.education!='None',:]
#-----样本特征及好坏标签
feature_names = ['id_type','gender','age',
                 'education','marriage','children_text','length_of_residence',
                 'job_type','on_the_job_time','monthly_net_income_text',
                 'loan_channel',
                 'eqc_direc_cell_num',
                 'allowMockLocation','batteryLevel','batteryStatus','brand',
                 'networkType','root','memory_avail_rate','storage_avail_rate',
                 'boot_hour',
#                 'home_distance_min','work_distance_min',
#                 'home_distance_move','work_distance_move',
#                 'distance_move_interval',
                 'apply_start_hour','apply_interval',
                 'od_days_ever'
                 ]
modeldata = data.loc[:,feature_names].dropna(how='any')
X = modeldata.drop(columns=['od_days_ever'],axis=1)
y = modeldata.loc[:,'od_days_ever'].apply(lambda x: 1 if x>7 else 0)
#-----划分训练集、测试集
X_train, X_test, y_train, y_test = train_test_split(X,y,test_size=0.3,random_state=33)
#-----特征转换
vec = DictVectorizer(sparse=False)
X_train = vec.fit_transform(X_train.to_dict(orient='records'))
X_test = vec.transform(X_test.to_dict(orient='records'))

#%%--决策树-------------------------------------------------------------------------
from sklearn import tree
from sklearn.tree import DecisionTreeClassifier
from sklearn.tree.export import export_text
import graphviz

#-----分类器训练及预测
dtc = DecisionTreeClassifier(max_depth=4,min_samples_leaf=50,random_state=0)
dtc.fit(X_train,y_train)
y_predict = dtc.predict(X_test)
y_train_predict = dtc.predict(X_train)
#-----分类器性能评价
#训练集上
print(dtc.score(X_train,y_train))
print(classification_report(y_train,y_train_predict,target_names=['good','bad']))
#测试集上
print(dtc.score(X_test,y_test))
print(classification_report(y_test,y_predict,target_names=['good','bad']))
#------可视化树图 exported in Graphviz format 
dot_data = tree.export_graphviz(dtc,
                                feature_names=vec.feature_names_,
                                class_names=['good','bad'],
                                filled=True,rounded=True,
                                special_characters=True)
graph = graphviz.Source(dot_data)
#graph
#saved in an output file
graph.render(r"C:\Users\lenovo\Desktop\tree")
#exported in textual format
#r = export_text(dtc,feature_names=vec.feature_names_)
#print(r)

#%%--随机森林----------------------------------------------------------------------
from sklearn.ensemble import RandomForestClassifier

#----分类器训练及预测
rfc = RandomForestClassifier(max_depth=4)
rfc.fit(X_train,y_train)
rfc_y_predict = rfc.predict(X_test)
rfc_y_train_predict = rfc.predict(X_train)
#-----分类器性能评价
print(rfc.score(X_train,y_train))
print(classification_report(y_train,rfc_y_train_predict,target_names=['good','bad']))

print(rfc.score(X_test,y_test))
print(classification_report(y_test,rfc_y_predict,target_names=['good','bad']))

#%%--梯度提升决策树-----------------------------------------------------------------
from sklearn.ensemble import GradientBoostingClassifier

#-----分类器训练及预测
gbc = GradientBoostingClassifier()
gbc.fit(X_train,y_train)
gbc_y_predict = gbc.predict(X_test)
gbc_y_train_predict = gbc.predict(X_train)
#------分类器性能评价
print(gbc.score(X_train,y_train))
print(classification_report(y_train,gbc_y_train_predict,target_names=['good','bad']))
print(gbc.score(X_test,y_test))
print(classification_report(y_test,gbc_y_predict,target_names=['good','bad']))

#------------------------------------------------------------------------------
dtc = DecisionTreeClassifier(max_depth=4,min_samples_leaf=100,random_state=0)
feature_names = ['gender','age','marriage','education','children_text','id_type',
                 'job_type','on_the_job_time','monthly_net_income_text','loan_channel']  

feature_names = ['home_distance_min']  
X = data.loc[:,feature_names]
y = data.loc[:,'od_days_ever'].apply(lambda x: 1 if x>10 else 0)
feature_names = ['home_distance_min','od_days_ever']  
modeldata = data.loc[:,feature_names].dropna(how='any')
X = modeldata.drop(columns=['od_days_ever'],axis=1)
y = modeldata.loc[:,'od_days_ever'].apply(lambda x: 1 if x>10 else 0)

vec = DictVectorizer(sparse=False)
X = vec.fit_transform(X.to_dict(orient='records'))

dtc.fit(X,y)
dot_data = tree.export_graphviz(dtc,
                                feature_names=vec.feature_names_,
                                class_names=['good','bad'],
                                filled=True,rounded=True,
                                special_characters=True)
graph = graphviz.Source(dot_data)
graph.render(r"F:\菲律宾日报\tree")

#%%--新贷自动通过客群---------------------------------------------------------------
first_manual = appr_mart.query("loan_type=='first_apply' & \
                               approval_status in ('MANUAL_APPROVED','MANUAL_REJECTED')")
feature_names = ['gender','age','marriage','education','children_text','id_type',
                 'length_of_residence','job_type','on_the_job_time',
                 'monthly_net_income_text']
feature_names = ['gender','age','marriage','education','children_text',
                 'length_of_residence','on_the_job_time',
                 'monthly_net_income_text','payment_type']

X = first_manual.loc[first_manual.apply_month=='201911',feature_names]
y = first_manual.loc[first_manual.apply_month=='201911','approval_status'].apply(lambda x: 1 if x=='MANUAL_APPROVED' else 0)

vec = DictVectorizer(sparse=False)
X = vec.fit_transform(X.to_dict(orient='records'))
dtc = DecisionTreeClassifier(max_depth=4,min_samples_leaf=100,random_state=0)
dtc.fit(X,y)
dot_data = tree.export_graphviz(dtc,
                                feature_names=vec.feature_names_,
                                class_names=['reject','approval'],
                                filled=True,rounded=True,
                                special_characters=True)
graph = graphviz.Source(dot_data)
graph.render(r"F:\菲律宾日报\approval_tree")
#%%--style样式-----------------------------------------------------------------------
appl.style.bar(subset= ['处理','通过'], align='mid', color=['#d65f5f', '#5fba7d']).format(precision=0, na_rep='MISSING', thousands=" ",formatter={'处理': "{:.2f}" })
appl.style.apply(highlight_max, props='color:white;background-color:purple', axis=0).format(formatter={'处理': "{:,}" })


#%%--提供工作证明的客户-----------------------------------------------------------------------
os.chdir(r"F:\菲律宾日报")
Proof_Employment=pd.read_sql("select apply_code ,1 as poe  from approval.image_info where file_desc='Proof of Employment'   ",cnx)
sas.df2sd(Proof_Employment,'Proof_Employment',"MY",encode_errors='replace')#csv那里有了

#%%--新贷预授信-----------------------------------------------------------------------

import json
import numpy as np
import pandas as pd
import sys
sys.path.append(r"C:\Users\lenovo\Anaconda3\Lib\site-packages")
import pymysql
#import math
import datetime
#import gc
#gc.collect() 垃圾回收，返回处理这些循环引用一共释放掉的对象个数
import os
import oss2
import zipfile
from itertools import islice
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = [u'SimHei']
plt.rcParams['axes.unicode_minus'] = False
import saspy
os.chdir(r"F:\菲律宾日报")



# sas.df2sd(data1,'data_behavior',"MY",encode_errors='replace')
#device-----
# os.chdir(r"H:\搬家\F\TS\PreWork\device")
# data=pd.read_csv('device.csv')
# data1=data.copy()
# data1.drop_duplicates(subset='applyCode',keep='last',inplace=True)
# data1.drop(columns=['limitCount7D','limitCount14D'],inplace=True)
# a=data.head(100)
#device-----

os.chdir(r"F:\菲律宾日报")

data=pd.read_csv('data.csv')
data1=data.copy()
data1.drop_duplicates(subset='applyCode',keep='last',inplace=True)
data1.drop(columns=['limitCount7D','limitCount14D'],inplace=True)
col=['age','gender','customer_source_sys','id_type','pay_type_name','job_type','education','白名单1','apply_month',
     'apply_code','取消','处理','通过','放款','人工通过','人工处理','锁定期',
     'apply_time','apply_loan_amount',
     'loan_type']
appr_data=appr_mart.loc[:,col]
appr_data=pd.merge(appr_data,data1,how='left',left_on='apply_code',right_on='applyCode')
appr_data['apply_time']=pd.to_datetime(appr_data['apply_time']).dt.date

ateste=appr_mart.loc[:,['apply_month','case_state','loan_type','apply_code','user_code','approval_status']]
ateste=pd.merge(ateste,data,how='left',left_on='apply_code',right_on='applyCode')

ateste1=ateste.loc[(ateste.apply_month=='202104') & (ateste.loan_type=='first_apply') & (ateste.case_state=='endevent') ,:]
ateste1.applyCount1M.value_counts()
ateste1.limitCount1M.value_counts()
ateste1.firstLimitIntvl.value_counts()


    

atest11=appr_data.loc[(appr_data.apply_month=='202111') & (appr_data.loan_type=='first_apply') ,['nowHour','applyCount1M','limitCount1M','firstLimitIntvl','apply_code']]


sas.df2sd(tcredit.astype(str),'tcredit1',"MY",encode_errors='replace')
tcredit=sas.sd2df('tcredit1','MY')
urule_09upiso_p1=pd.read_csv('20220406_urule_09upiso.csv')




# upiso的评分卡,20220316发现有39000个
with open("oss_access_key_txl.json") as oss_access_key:
    oss_args = json.load(oss_access_key)
with open(r"db_config.json") as db_config:
    cnx_args = json.load(db_config)
auth = oss2.Auth(oss_args['access_key_id'], oss_args['access_key_secret'])
bucket = oss2.Bucket(auth, oss_args['endpoint'], 'suncash-prd')
cnx = pymysql.connect(**cnx_args)

# sas=saspy.SASsession()
# sas.saslib('MY "F:\菲律宾日报"')
# urule_09upiso_p=sas.sd2df('urule_09upiso','MY')
a=urule_upiso.copy(deep=True)




try:
    os.chdir(r"F:\菲律宾日报")
except:
    os.chdir(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart")
# urule
urule_09upiso_p=pd.read_csv('urule_09upiso.csv')
urule_09upiso_p['updatetime']=datetime.date(2022,5,1)
# urule_09upiso_p['len']=urule_09upiso_p.usercode.apply(lambda x: len(x))#之前的纰漏导致user_code字段存了id内容
# urule_09upiso_p=urule_09upiso_p[urule_09upiso_p.len!=22]#之前的纰漏导致user_code字段存了id内容

uruleone=pd.read_sql("select * from approval.urule_flow_context   ",cnx)
urule_upiso = uruleone.loc[uruleone.package_id=='SunCash/009',:]
urule_upiso.sort_values(by=['user_code','created_time'],inplace=True)
urule_upiso.drop_duplicates(subset='user_code',keep='last',inplace=True)#后面20220411给了名单让it冲洗，用last来
urule_upiso['created_time']=pd.to_datetime(urule_upiso.created_time).dt.date


apply_info=pd.read_sql("select apply_time,loan_type,user_code from approval.apply_info   ",cnx)
apply_info['apply_timeaaaa']=pd.to_datetime(apply_info['apply_time']).dt.date
data=apply_info.loc[(apply_info.apply_timeaaaa>=datetime.date(2022,1,1)) 
                    & (apply_info.loan_type=='first_apply'),['user_code','apply_timeaaaa']]#随便定一个日期

# a=data_it.head(20)
# data=pd.merge(data,urule_09upiso_p,how='left',left_on='user_code',right_on='usercode')
# data=data[data.usercode.isna()  ]
data=pd.merge(data,urule_upiso,how='left',on='user_code')
data=data.loc[:,['user_code','flow_context_key','created_time']]
data['test']=data.user_code.apply(lambda x:x.rfind('FlashLoan')) #0表示找到的位置，-1是没找到

# data=data[data.created_time==datetime.date(2022,4,14)   ]#0用这天创建的200个客户检查了一下，大致应该没问题
data=data[data.flow_context_key.notna()  ]
data=data[data.test==0 ]

os.chdir(r"F:\菲律宾日报\json")
for i in range(len(data)):
# for i in range(0,200):
    oss_key = data.iloc[i,1]
    
    temp_file=r"F:\菲律宾日报\json\temp.json"
    bucket.get_object_to_file(oss_key,temp_file)
    with open('temp.json','r',encoding='UTF-8') as temp:
        temp_json = json.load(temp)
    try:
        print(i)
        # temp_objects1 = pd.DataFrame(temp_json['objects'][0],index=[0])
        temp_objects2 = pd.DataFrame(temp_json['objects'],index=[0])
        temp_objects3 = pd.DataFrame(temp_json['result'],index=[0])
        temp_objects2['usercode']=data.iloc[i,0]
        temp_objects3['usercode']=data.iloc[i,0]
        # temp_objects=pd.merge(temp_objects1,temp_objects2,how='left',on='applyCode')
        temp_objects=pd.merge(temp_objects2,temp_objects3,how='left',on='usercode')
    except:
            continue
    if i == 0:
        urule_v = temp_objects
    else:
        urule_v = pd.concat([urule_v,temp_objects],axis=0,sort=True)  
urule_v.drop_duplicates(subset='usercode',keep='first',inplace=True)
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
sas.df2sd(urule_v.astype(str),'urule_v_F0103',"MY",encode_errors='replace')

b=datetime.date.today().strftime("%Y%m%d")
data_download=pd.concat([urule_09upiso_p,urule_v],axis=0,sort=True)  
data_download.to_csv(r'F:\菲律宾日报\\' +b+ '_urule_09upiso.csv')
data_download.to_csv(r'F:\菲律宾日报\urule_09upiso.csv')

sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
sas.df2sd(urule_v.astype(str),'urule_v200',"MY",encode_errors='replace')

data_download.drop(columns=['userCode'],inplace=True)
sas.df2sd(data_download.astype(str),'urule_09upiso',"MY",encode_errors='replace')


aadata=data.loc[(data.apply_timeaaaa==datetime.date(2022,4,4)) 
                    ,['user_code','apply_timeaaaa']]#随便定一个日期


te1=uruleone.loc[uruleone.user_code=='FlashLoan_202203120612466895990',:]

os.chdir(r"F:\菲律宾日报\json")
oss_key=r'urule/2022/05/19/03/43/7f29bafd-a27a-44f7-b61d-50f43304b471fb0bb0dd1846cec4b8ecf6df7777df38'
temp_file=r"F:\菲律宾日报\json\temp.json"
bucket.get_object_to_file(oss_key,temp_file)
with open('temp.json','r',encoding='UTF-8') as temp:
    temp_json = json.load(temp)

try:
    b=datetime.date.today().strftime("%Y%m%d")
    data_download=pd.concat([data_it,urule_v],axis=0,sort=True)  
    data_download.to_csv(r'F:\菲律宾日报\\' +b+ '_urule.csv')
    data_download.to_csv(r'F:\菲律宾日报\data_urule.csv')
    data_it=data_download.loc[:,['applyCode','totalNumberOfLoans','nowHour','eqcDirecValidCellNum','gender','idType','isWhiteList','jobType','paymentType',
                       'applyCount1M', 'limitCount1M', 'firstLimitIntvl','age','education','eqcDirecCellNum']]
    data_it.rename(columns={'gender':'gender_u','idType':'idType_u','isWhiteList':'isWhiteList_u','jobType':'jobType_u','paymentType':'paymentType_u',
                            'age':'age_u','education':'education_u'},inplace=True)
    data_it.drop_duplicates(subset='applyCode',keep='last',inplace=True)
except:

    data_it=data_it.loc[:,['applyCode','totalNumberOfLoans','nowHour','eqcDirecValidCellNum','gender','idType','isWhiteList','jobType','paymentType',
                           'applyCount1M', 'limitCount1M', 'firstLimitIntvl','age','education','eqcDirecCellNum']]
    data_it.rename(columns={'gender':'gender_u','idType':'idType_u','isWhiteList':'isWhiteList_u','jobType':'jobType_u','paymentType':'paymentType_u',
                            'age':'age_u','education':'education_u'},inplace=True)
    data_it.drop_duplicates(subset='applyCode',keep='last',inplace=True)


data_it.eqcDirecValidCellNum.fillna(0,inplace=True)
data_it['eqcDirecValidCellNum'] = data_it.eqcDirecValidCellNum.map(lambda x:int(x))
data_it['vc_num'] = pd.cut(data_it.eqcDirecValidCellNum,
                                 bins=[-np.inf,9,19,29,39,49,np.inf],
                                 labels=['A1','A2','A3','A4','A5','A6'])


#P包：利用申请日期为1.25后的usercode来urule_flow_context寻找第一条009
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
appr_mart=sas.sd2df('appr_mart','MY')
appr_mart['apply_time'] = pd.to_datetime(appr_mart['apply_time'] ).dt.date
appr_list=appr_mart.loc[(appr_mart.apply_time>=datetime.date(2022,1,25)) & \
                        (appr_mart.customer_source_sys=='SuncashPautang') & \
                        (appr_mart.loan_type=="first_apply"),:]
uruleone=pd.read_sql("select * from approval.urule_flow_context   ",cnx)
urule_kan = uruleone.loc[uruleone.package_id=='SunCash/009',:]
temp_list_p=appr_list.user_code

alist_dict={}
alist_dict = alist_dict.fromkeys(temp_list_p)
alist_t = list(alist_dict.keys())
urule_list=uruleone[uruleone.apply(lambda x: x.user_code  in alist_t,axis=1)]
urule_list = urule_list.loc[urule_list.package_id=='SunCash/009',:]
urule_list = urule_list.sort_values(by=['user_code','created_time'])
urule_list.drop_duplicates(subset='user_code',keep='first',inplace=True)
urule_kan = uruleone.loc[uruleone.package_id=='SunCash/009',:]

# s包
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
appr_mart=sas.sd2df('appr_mart','MY')
appr_mart['apply_time'] = pd.to_datetime(appr_mart['apply_time'] ).dt.date
appr_list_s=appr_mart.loc[(appr_mart.apply_time>=datetime.date(2022,2,25)) & \
                        (appr_mart.customer_source_sys=='SunCash') & \
                        (appr_mart.loan_type=="first_apply"),:]
appr_list_s=appr_list_s.loc[:,['apply_code','user_code','approval_status','人工通过','人工处理']]  
uruleone_s=uruleone.loc[ (uruleone.package_id=='SunCash/001'),:]
uruleone_s = uruleone_s.sort_values(by=['apply_code','created_time'])
uruleone_s.drop_duplicates(subset='apply_code',keep='first',inplace=True)
appr_list_s=pd.merge(appr_list_s,uruleone_s,how='left',on='apply_code')

lista=appr_list_s.loc[:,['apply_code','flow_context_key']]

os.chdir(r"F:\菲律宾日报\json")
for i in range(len(lista)):
    oss_key = lista.iloc[i,1]
    
    temp_file=r"F:\菲律宾日报\json\temp.json"
    bucket.get_object_to_file(oss_key,temp_file)
    with open('temp.json','r',encoding='UTF-8') as temp:
        temp_json = json.load(temp)
    try:
        print(i)
        # temp_objects1 = pd.DataFrame(temp_json['objects'][0],index=[0])
        temp_objects2 = pd.DataFrame(temp_json['objects'][1],index=[0])
        # temp_objects3 = pd.DataFrame(temp_json['objects'][2],index=[0])
        temp_objects2['applyCode']=lista.iloc[i,0]
        # temp_objects=pd.merge(temp_objects1,temp_objects2,how='left',on='applyCode')
        # temp_objects=pd.merge(temp_objects,temp_objects3,how='left',on='applyCode')
    except:
            continue
    if i == 0:
        urule_v = temp_objects2
    else:
        urule_v = pd.concat([urule_v,temp_objects2],axis=0,sort=True)    
sas.saslib('MY "F:\菲律宾日报"')
sas.df2sd(urule_v.astype(str),'urule_v_s',"MY",encode_errors='replace')       
# 检视草稿        
oss_key =r'urule/2022/02/16/15/50/32c1f6cc-b1f8-4b90-ba06-164f22fb870503c69597488c73a2797eae0ee034356f'
temp_file=r"F:\菲律宾日报\kanjson\kan8.json"
# temp_file=r"H:\搬家\F\TS\PreWork\rule\data\kan8.json"
bucket.get_object_to_file(oss_key,temp_file)
#18:484,19:453,21:452,22:459
# os.chdir(r"H:\搬家\F\TS\PreWork\rule\data")
os.chdir(r"F:\菲律宾日报\kanjson")
with open('kan8.json','r',encoding='UTF-8') as temp:
        apemp_json8 = json.load(temp)
apemp_json88 = pd.DataFrame(apemp_json8['objects'][1],index=[0]) 




kan=uruleone.loc[ (uruleone.package_id=='SunCash/009'),:]#看P包的评分卡策略

uruleone=uruleone.loc[(uruleone.apply_code.notna()) & (uruleone.package_id=='SunCash/001'),:]#005预授信还没有订单号，001有，盲猜是先走预授信再走自动拒绝,没有重复
uruleone['created_time']=pd.to_datetime(uruleone['created_time'] ).dt.date
test =uruleone.loc[(uruleone.created_time>=datetime.date(2022,2,24)) & (uruleone.apply_code.notna()) & (uruleone.package_id=='SunCash/001'),:]
sas=saspy.SASsession()
sas.saslib('MY "H:\搬家\F\TS\PreWork\\rule"')



uruleone_before=sas.sd2df('uruleone_before','MY')

uruleone_now=pd.merge(uruleone.loc[:,['apply_code','flow_context_key']],uruleone_before.loc[:,['apply_code','id']],how='left',on='apply_code')


sas.df2sd(uruleone,'uruleone_before',"MY",encode_errors='replace')

lista=uruleone_now[uruleone_now.id.isna()]
lista=uruleone.loc[:,['apply_code','flow_context_key']]

os.chdir(r"H:\搬家\F\TS\PreWork\rule\json")
for i in range(len(lista)):
    oss_key = lista.iloc[i,1]
    
    temp_file=r"H:\搬家\F\TS\PreWork\rule\json\temp.json"
    bucket.get_object_to_file(oss_key,temp_file)
    with open('temp.json','r',encoding='UTF-8') as temp:
        temp_json = json.load(temp)
    try:
        print(i)
        temp_objects1 = pd.DataFrame(temp_json['objects'][0],index=[0])
        temp_objects2 = pd.DataFrame(temp_json['objects'][1],index=[0])
        temp_objects3 = pd.DataFrame(temp_json['objects'][2],index=[0])
        temp_objects3['applyCode']=lista.iloc[i,0]
        temp_objects=pd.merge(temp_objects1,temp_objects2,how='left',on='applyCode')
        temp_objects=pd.merge(temp_objects,temp_objects3,how='left',on='applyCode')
    except:
            continue
    if i == 0:
        urule_v = temp_objects
    else:
        urule_v = pd.concat([urule_v,temp_objects],axis=0,sort=True)    


urule_v=pd.read_excel(r'H:\搬家\F\TS\PreWork\rule\urule_total.xlsx')
u_test=urule_v.loc[:,['applyCode','limitCount1M','firstLimitIntvl','applyCount1M','nowHour']]
test=lista.head(50)
test1=pd.merge(test,urule_v,how='left',left_on='apply_code',right_on='applyCode')
range(len(lista))

        
for i in range(len(alist)):
    with open(alist[i],'r',encoding='UTF-8') as temp:
        temp_json = json.load(temp)
    temp_df = pd.DataFrame(temp_json)
    try:
        temp_df=temp_df.loc[temp_df.phoneList.notna(),:]
        

        s=pd.DataFrame({'phone':np.concatenate(temp_df.phoneList.values),'name':temp_df.name.repeat(temp_df.phoneList.str.len())})
        s['user_code'] = alist[i]
    except:
            continue
    if i == 0:
        contact_book = s
    else:
        contact_book = pd.concat([contact_book,s],axis=0,sort=True)    
    
   kan8是 01自动审批
   kan9是 05新贷授信
PL202201081731568030033     urule/2021/12/06/05/54/cc543123-bc39-4c54-b338-1531bd63e63f8646da55f1f3d2569193d1bc19e9eb4d
PL202201101809418100033     urule/2022/01/10/10/15/bd4b8606-278a-4570-9216-1445a5100b3b870a9c79a21f38102fccd47f1dbe44f4

test_code =uruleone.loc[(uruleone.apply_code=='PL202201300746114100017') ,:]
test_code1 =uruleone.loc[(uruleone.user_code=='suncash-pautang_202202181844508456319') ,:]
test_code2 =appr_mart.loc[(appr_mart.user_code=='suncash-pautang_202202181844508456319') ,:] 
# uruleone['created_time']=pd.to_datetime(uruleone.created_time).dt.date
# test =uruleone.loc[(uruleone.created_time>=datetime.date(2022,1,1))  & (uruleone.package_id=='SunCash/005'),:]
oss_key =r'urule/2022/02/16/00/45/8383a3d9-c5e8-4ffa-b05e-a34e53b7e30f57819c7a6189891f6648421d357dcbc0'
temp_file=r"F:\菲律宾日报\kanjson\kan8.json"
# temp_file=r"H:\搬家\F\TS\PreWork\rule\data\kan8.json"
bucket.get_object_to_file(oss_key,temp_file)
#18:484,19:453,21:452,22:459
# os.chdir(r"H:\搬家\F\TS\PreWork\rule\data")
os.chdir(r"F:\菲律宾日报\kanjson")
with open('kan8.json','r',encoding='UTF-8') as temp:
        temp_json8 = json.load(temp)
temp_objects8 = pd.DataFrame(temp_json8['objects'][2],index=[0])       
        
acheck=data.loc[data.applyCode=="PL202111191653178610087",:]
temp_result = pd.DataFrame(temp_json)


a=uruleone.groupby('apply_code').agg({'id':"count"})
b=uruleone.loc[uruleone.apply_code=='C201912071736309830058',:]
b1=appr_mart.loc[appr_mart.apply_code=='C201912071736309830058',:]









uruleoneaaa=pd.read_sql("select * from approval.urule_flow_context  where apply_code='PL202111241541003920027' ",cnx)
boroone=pd.read_sql("select * from approval.borrower_info  where apply_code='PL202111081302057960008' ",cnx)
borotwo=pd.read_sql("select * from approval.borrower_info  where phone_no='09359549183' ",cnx)
borothree=pd.read_sql("select APPLY_CODE,CREATED_TIME,ID_NO,ID_TYPE from approval.borrower_info  where id_no='011302508652' ",cnx)
userone=pd.read_sql("select * from approval.apply_info  where user_code='suncash-lend_202111081420443109732' ",cnx)
phoneone=pd.read_sql("select * from approval.apply_info  where user_code='suncash-lend_202111081420443109732' ",cnx)
applyone=pd.read_sql("select * from approval.apply_info  where apply_code='PL202111081302057960008' ",cnx)


oss_key =r'urule/2022/02/27/09/55/bd4df201-2601-4c22-b8b5-d4eb4e7104ad86ec88bb6696a83c6d4fca93b24db5a7'
temp_file=r"H:\搬家\F\TS\PreWork\rule\data\kan10.json"
bucket.get_object_to_file(oss_key,temp_file)


temp_param = pd.DataFrame(temp_json['param'])



temp_objects1 = pd.DataFrame(temp_json['objects'][0],index=[0])
temp_objects2 = pd.DataFrame(temp_json['objects'][1],index=[0])
temp_objects3 = pd.DataFrame(temp_json['objects'][2],index=[0])
temp_objects3['applyCode']='PL202111060659237950098'
temp_objects3['userCode']='suncash-lend_202101051948343926189'
temp_objects=pd.merge(temp_objects1,temp_objects2,how='left',on=['applyCode','userCode'])
temp_objects=pd.merge(temp_objects,temp_objects3,how='left',on=['applyCode','userCode'])

temp_objects.to_excel(r'F:\菲律宾日报\数据清洗PL202111060659237950098.xlsx')

a=appr_mart.loc[appr_mart.user_code=='suncash-lend_202111131505408546647',:]

a=list(temp_objects.columns)




b=pd.Series()
for col in a:
    b[col]=','.join(temp_objects[col].unique())


def risk_group(x):
    if x.age>=43 and x.gender=='Female' and x.loan_channel=='DRAGON_PAY':
        return '客群1'
    elif x.age>=43 and x.gender=='Female' and x.loan_channel=='SKY_PAY':
        return '客群2'
    elif x.age>=35 and x.gender=='Female' and x.education in(['College Graduate','Master/PHD']):
        return '客群3'
    elif x.age>=35 and x.gender=='Female':
        return '客群4'
    elif x.age>=43 and x.gender=='Male' and x.loan_channel=='DRAGON_PAY':
        return '客群5'
    elif x.age>=43 and x.gender=='Male':
        return '客群6'
    elif x.age>=35 and x.gender=='Male':
        return '客群7'
    elif x.age<35 and x.gender=='Female' and x.education in(['College Graduate','Master/PHD']) and x.job_type!='Private Company Employee':
        return '客群8'
    else:
        return '客群9'
cnx = pymysql.connect(**cnx_args)
#需要原生appr_mart 跟 repay_mart
base_derived_variable = pd.read_sql("select * from approval.base_derived_variable",cnx)
device_info_derived_variable= pd.read_sql("select * from approval.device_info_derived_variable",cnx)
appr=pd.merge(appr_mart,base_derived_variable,on='apply_code',how='left')
appr=pd.merge(appr,device_info_derived_variable,on='apply_code',how='left')
repay=repay_mart.loc[:,['od_days_ever','apply_code']].copy()
appr=pd.merge(appr,repay,on='apply_code',how='left')
appr.sort_values(by=['phone_no','apply'])
appr['lag_oddays']=appr['contract_no'].shift(1)
case_info['lag_contract_no'] = case_info['contract_no'].shift(1)
case_flow_st = pd.merge(case_flow_st,case_info,how='left',on='contract_no')
case_flow_st['clear_date']=pd.to_datetime(case_flow_st['clear_date']).dt.date
case_flow_st['user_id']=case_flow_st['user_id'].fillna('NOONE')
case_flow_st['user_id']=case_flow_st['user_id'].apply(lambda x:x.lower())


#----近一个月申请次数
apply_info['created_time']=pd.to_datetime(apply_info.created_time)

#这里不能用断点思路，因为有些apply_code会因为首次授信即借款导致缺失，断点意味着无限重复
# Tcredit_p=sas.sd2df('Tcredit_t','MY')
# apply_info=pd.merge(apply_info,Tcredit_p,how='left',left_on='apply_code',right_on='code')
# lista=apply_info.loc[apply_info.code.isna(),['user_code','apply_code','created_time']].apply(tuple,axis=1).tolist()

lista=apply_info.loc[:,['user_code','apply_code','created_time']].apply(tuple,axis=1).tolist()
# b=a.apply(tuple,axis=1)
# c=b.tolist()[0:5].drop(columns=[''])

tapply = pd.DataFrame()
for i,x in enumerate(lista):
    sdate=x[2]-datetime.timedelta(days=30)
    temp=apply_info.loc[(apply_info.user_code==x[0]) & (apply_info.created_time <=x[2]) & (apply_info.created_time >sdate),
                        ['user_code','apply_code','created_time']]
    apply_count=temp.groupby('user_code').agg({'apply_code':np.size}).rename(columns={'apply_code':'apply_count'})
    apply_count['code']=x[1]
    if i==0:
        tapply=apply_count
    else:
        tapply = pd.concat([tapply,apply_count])
        
        
#----近一个月授信次数&距首次授信时间间隔
Scredit_limit_log = pd.read_sql("select * from suncash_lend.credit_limit_log ",cnx)
Pcredit_limit_log = pd.read_sql("select * from suncash_pautang.credit_limit_log ",cnx)

#检查后发现用log表的第一条可以实现
# Scredit_limit = pd.read_sql("select * from suncash_lend.credit_limit",cnx)
# Pcredit_limit = pd.read_sql("select * from suncash_pautang.credit_limit",cnx)


#这里不能用断点思路，因为有些apply_code会因为首次授信即借款导致缺失，断点意味着无限重复

col=['created_time','status','user_code']
S_log=Scredit_limit_log.loc[Scredit_limit_log.status==1,col]
P_log=Pcredit_limit_log.loc[Pcredit_limit_log.status==1,col]
log=pd.concat([S_log,P_log])
logf = log.sort_values(by=['user_code','created_time']).rename(columns={'created_time':'created_time_log'})
logf.drop_duplicates(subset='user_code',keep='first',inplace=True)
# 距首次授信时间间隔
tcredit_inter = pd.merge(apply_info.loc[:,['user_code','apply_code','created_time']],logf,how='left',on='user_code')

#近一个月授信次数
tcredit = pd.DataFrame()
for i,x in enumerate(lista):
    sdate=x[2]-datetime.timedelta(days=30)
    sdate2=x[2]-datetime.timedelta(days=14)
    temp=log.loc[(log.user_code==x[0]) & (log.created_time <=sdate2) & (log.created_time >sdate),:]
    tcredit_count=temp.groupby('user_code').agg({'user_code':np.size}).rename(columns={'user_code':'credit_count'})
    tcredit_count['code']=x[1]
    print(i)
    if i==0:
        tcredit=tcredit_count
    else:
        tcredit = pd.concat([tcredit,tcredit_count])


sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
sas.df2sd(tapply.astype(str),'tapply',"MY",encode_errors='replace')
sas.df2sd(tcredit.astype(str),'tcredit1',"MY",encode_errors='replace')
sas.df2sd(log.astype(str),'log',"MY",encode_errors='replace')
sas.df2sd(tcredit_inter.astype(str),'tcredit_inter',"MY",encode_errors='replace')

  
akan=apply_info.loc[apply_info.user_code=='suncash-lend_201906031129393541109',:]      
akan.sort_values(by=['user_code','created_time'],inplace=True)   
#%%--业务检视-----------------------------------------------------------------------

kan=apply_info.loc[apply_info.apply_code=='PL202202241235121215756',:]
acc_info=pd.read_sql("select * from account.account_info   ",cnx)
kan=acc_info.loc[acc_info.CONTRACT_NO=='C202201021454181720092',:]

sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')

sas.df2sd(acc_info.astype(str),'acc_info',"MY",encode_errors='replace')


#------系统自动拒绝 只看number_of_apply=1的申请
system = appr_mart.loc[(appr_mart.number_of_apply==1) & (appr_mart.apply_month=='201911'),
                            ['apply_code','apply_day','auto_refuse_reason']]
system.to_excel(r"C:\Users\lenovo\Desktop\temp.xlsx",index=False)
#------人工通过率及拒绝原因
manual = appr_mart.loc[appr_mart.number_of_apply==1,
                       ['apply_code','approval_status','apply_month','manual_refuse_reason_1','manual_refuse_reason_2']]
manual.to_excel(r"C:\Users\lenovo\Desktop\temp.xlsx",index=False)
#------人群分布
demo_feature = ['apply_code','apply_month','media_source','payment_type','id_type',
                'age','gender','education','marriage','children_text',
                'length_of_residence','job_type','monthly_net_income_text',
                'on_the_job_time']
demo_feature = ['apply_code','apply_month','contact_num_grp']

demo = appr_mart.loc[appr_mart.number_of_apply==1,demo_feature]
demo.to_excel(r"C:\Users\lenovo\Desktop\temp.xlsx",index=False)


demo_feature = ['apply_code','loan_month','media_source','payment_type','id_type',
                'age','gender','education','marriage','children_text',
                'length_of_residence','job_type','monthly_net_income_text',
                'on_the_job_time','loan_channel','od_days_ever','last_repay_date']
demo = repay_mart.loc[repay_mart.number_of_loans==1,demo_feature]
demo['自然逾期'] = demo.od_days_ever.apply(lambda x: 1 if x>0 else 0)
demo['到期'] = demo.payment_type.apply(lambda x: 0 if x=='ACCOUNT_NORMAL' else 1)
demo['曾经逾期7天以上'] = demo.od_days_ever.apply(lambda x: 1 if x>7 else 0)
demo['到期7天以上'] = demo.last_repay_date.apply(lambda x: 1 if (datetime.date.today()-x).days>7 else 0)
demo['age_grp'] = pd.cut(demo.age,
                        bins=[-np.inf,21,28,35,49,np.inf],
                        labels=['  - 21','22 - 28','29 - 35','36 - 49','50 -  '])
demo.to_excel(r"C:\Users\lenovo\Desktop\temp.xlsx",index=False)


from openpyxl import load_workbook
writer = pd.ExcelWriter(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\repay_mart.xlsx",engine='openpyxl')
writer.book = load_workbook(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\repay_mart.xlsx")
idx = writer.book.sheetnames.index('source')
writer.book.remove(writer.book.worksheets[idx])
#writer.book.create_sheet('source',idx)
repay_mart.to_excel(writer,sheet_name='source',index=False)
writer.save()

#%%--埋点数据----------------------------------------------------------------------
with open(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\data_collect.json") as f:
    db_config = json.load(f)
con = pymysql.connect(**db_config)
save_dir = r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\extData\behavioral"

#behavioral_data = pd.read_sql("select phone_no, event_data, event_date, event_name, \
#                                      ip_address, mobile_brand \
#                               from data_collect_prd.behavioral_data \
#                               where phone_no in ('09773830271','09473109954')",con)
#------增量更新数据start
#behavioral_data = pd.read_sql("select phone_no, event_data, event_date, event_name, \
#                                      ip_address, mobile_brand \
#                               from data_collect_prd.behavioral_data \
#                               where event_date<'2019-11-23'",con)
#behavioral_data.to_csv(save_dir+'\\'+'behavioral_data.csv',index=False)
#
#b_d = pd.read_csv(save_dir+'\\'+'behavioral_data.csv')
#behavioral_data = pd.read_sql("select phone_no, event_data, event_date, event_name, \
#                                      ip_address, mobile_brand \
#                               from data_collect_prd.behavioral_data \
#                               where event_date between '2019-11-23' and '2019-11-24'",con)
#b_d = pd.concat([b_d,behavioral_data])
#b_d.to_csv(save_dir+'\\'+'behavioral_data.csv',index=False)
#
#b_d = pd.read_csv(save_dir+'\\'+'behavioral_data.csv')
#behavioral_data = pd.read_sql("select phone_no, event_data, event_date, event_name, \
#                                      ip_address, mobile_brand \
#                               from data_collect_prd.behavioral_data \
#                               where event_date between '2019-11-24' and '2019-11-27'",con)
#b_d = pd.concat([b_d,behavioral_data])
#b_d.to_csv(save_dir+'\\'+'behavioral_data.csv',index=False)
#
#b_d = pd.read_csv(save_dir+'\\'+'behavioral_data.csv')
#behavioral_data = pd.read_sql("select phone_no, event_data, event_date, event_name, \
#                                      ip_address, mobile_brand \
#                               from data_collect_prd.behavioral_data \
#                               where event_date between '2019-11-27' and '2019-12-03'",con)
#b_d = pd.concat([b_d,behavioral_data])
#b_d.to_csv(save_dir+'\\'+'behavioral_data.csv',index=False)
#
#b_d = pd.read_csv(save_dir+'\\'+'behavioral_data.csv',dtype={'phone_no':str})
#behavioral_data = pd.read_sql("select phone_no, event_data, event_date, event_name, \
#                                      ip_address, mobile_brand \
#                               from data_collect_prd.behavioral_data \
#                               where event_date between '2019-12-03' and '2019-12-12'",con)
#b_d = pd.concat([b_d,behavioral_data])
#b_d.to_csv(save_dir+'\\'+'behavioral_data.csv',index=False)
#
#b_d = pd.read_csv(save_dir+'\\'+'behavioral_data.csv',dtype={'phone_no':str})
#behavioral_data = pd.read_sql("select phone_no, event_data, event_date, event_name, \
#                                      ip_address, mobile_brand \
#                               from data_collect_prd.behavioral_data \
#                               where event_date between '2019-12-12' and '2019-12-17'",con)
#b_d = pd.concat([b_d,behavioral_data])
#b_d.to_csv(save_dir+'\\'+'behavioral_data.csv',index=False)

b_d = pd.read_csv(save_dir+'\\'+'behavioral_data.csv',dtype={'phone_no':str})
behavioral_data = pd.read_sql("select phone_no, event_data, event_date, event_name, \
                                      ip_address, mobile_brand \
                               from data_collect_prd.behavioral_data \
                               where event_date between '2019-12-17' and '2020-01-01'",con)
b_d = pd.concat([b_d,behavioral_data])
b_d.to_csv(save_dir+'\\'+'behavioral_data.csv',index=False)

#%%--------增量更新数据end

b_d = pd.read_csv(save_dir+'\\'+'behavioral_data.csv',dtype={'phone_no':str})
first_apply_time = appr_mart.loc[appr_mart.number_of_apply==1,['phone_no','apply_time']]
b_d = pd.merge(b_d,first_apply_time,how='left',on='phone_no')
b_d = b_d.loc[pd.to_datetime(b_d.event_date)<=b_d.apply_time,:]

#open_personal_information的次数
#open_personal_info_cnt = behavioral_data.loc[behavioral_data.event_name=='open_personal_information',:].groupby('phone_no',as_index=False)['event_name'].agg({'open_personal_info_cnt':np.size})
#%%--各信息模块（项）操作次数
for cate in ['click_identification','open_identification',
             'click_upload_id_photo','select_id_photo',
             'click_identification_back_button','submit_identification',
             'click_personal_information','open_personal_information',
             'click_the_personal_info_back_button','submit_personal_information',
             'click_contact_information','open_contact_information',
             'click_the_contact_info_back_ button','submit_contact_information',
             'click_employment_information','open_employment_information',
             'click_upload_proof_of_employment01','select_the_proof_of_employment01',
             'click_upload_proof_of_employment02','select_the_proof_of_employment02',
             'click_the_employment_info_back_button','submit_employment_information']:
    locals()[cate] = b_d.loc[b_d.event_name==cate,:].groupby('phone_no',as_index=False)['event_name'].agg({cate:np.size})

#%%--各信息项不同取值数、是否有修改
#select_education_cnt = behavioral_data.loc[behavioral_data.event_name=='select_education',:].groupby('phone_no',as_index=False)['event_data'].agg({'select_education_cnt':pd.Series.nunique})
item = ['select_id_type','leave_id_type_input_box',
        'leave_first_name_ input_box','leave_middle_name_ input_box',
        'leave_last_name_ input_box','select_gender','select_birthday',
        'select_education','select_marital_status',
        'select_number_of_children','select_current_address',
        'select_length_of_residence',
        'leave_the_facebook_account_input_box',
        'leave_e-mail_input_box',
        'select_relationship01','select_contact_people01',
        'select_relationship02','select_contact_people02',
        'select_job_type','select_on-the-job_time',
        'select_ monthly_net_Income','leave_the_ company_name',
        'select_work_address','leave_the_detail_work_address_input_box',
        'leave_office_tel_input_box']
item_change = ['change_id_type','change_id',
               'change_first_name','change_middle_name',
               'change_last_name','change_gender','change_birthday',
               'change_education','change_marital_status',
               'change_number_of_children','change_current_address',
               'change_length_of_residence',
               'change_the_facebook_account',
               'change_e-mail',
               'change_relationship01','change_contact_people01',
               'change_relationship02','change_contact_people02',
               'change_job_type','change_on-the-job_time',
               'change_ monthly_net_Income','change_the_ company_name',
               'change_work_address','change_the_detail_work_address',
               'change_office_tel']
for i,j in zip(item,item_change):
    locals()[i] = b_d.loc[b_d.event_name==i,:].groupby('phone_no',as_index=False)['event_data'].agg({i:pd.Series.nunique})
    locals()[i][j] = locals()[i][i].apply(lambda x: 1 if x>1 else 0)
 
#ip_address和mobile_brand不同取值数、是否有修改
item = ['ip_address','mobile_brand']
item_change = ['change_ip_address','change_mobile_brand']
for i,j in zip(item,item_change):
    locals()[i] = b_d.groupby('phone_no',as_index=False)[i].agg({i:pd.Series.nunique})
    locals()[i][j] = locals()[i][i].apply(lambda x: 1 if x>1 else 0)
    
#拼接到appr_mart和repay_mart
for cate in ['click_identification','open_identification',
             'click_upload_id_photo','select_id_photo',
             'click_identification_back_button','submit_identification',
             'click_personal_information','open_personal_information',
             'click_the_personal_info_back_button','submit_personal_information',
             'click_contact_information','open_contact_information',
             'click_the_contact_info_back_ button','submit_contact_information',
             'click_employment_information','open_employment_information',
             'click_the_employment_info_back_button','submit_employment_information',
             'click_upload_proof_of_employment01','select_the_proof_of_employment01',
             'click_upload_proof_of_employment02','select_the_proof_of_employment02']:
    appr_mart = pd.merge(appr_mart,locals()[cate],how='left',on='phone_no')
    repay_mart = pd.merge(repay_mart,locals()[cate],how='left',on='phone_no')

for item in ['select_id_type','leave_id_type_input_box',
             'leave_first_name_ input_box','leave_middle_name_ input_box',
             'leave_last_name_ input_box','select_gender','select_birthday',
             'select_education','select_marital_status',
             'select_number_of_children','select_current_address',
             'select_length_of_residence',
             'leave_the_facebook_account_input_box',
             'leave_e-mail_input_box',
             'select_relationship01','select_contact_people01',
             'select_relationship02','select_contact_people02',
             'select_job_type','select_on-the-job_time',
             'select_ monthly_net_Income','select_work_address',
             'leave_the_detail_work_address_input_box',
             'leave_the_ company_name','leave_office_tel_input_box']:
    appr_mart = pd.merge(appr_mart,locals()[item],how='left',on='phone_no')
    repay_mart = pd.merge(repay_mart,locals()[item],how='left',on='phone_no')

for j in ['ip_address','mobile_brand']:
    appr_mart = pd.merge(appr_mart,locals()[j],how='left',on='phone_no')
    repay_mart = pd.merge(repay_mart,locals()[j],how='left',on='phone_no')


appr_mart['change_item_counts'] = 0
repay_mart['change_item_counts'] = 0
for item in ['change_id_type','change_id',
             'change_first_name','change_middle_name',
             'change_last_name','change_gender','change_birthday',
             'change_education','change_marital_status',
             'change_number_of_children','change_current_address',
             'change_length_of_residence',
             'change_the_facebook_account',
             'change_e-mail',
             'change_relationship01','change_contact_people01',
             'change_relationship02','change_contact_people02',
             'change_job_type','change_on-the-job_time',
             'change_ monthly_net_Income','change_the_ company_name',
             'change_work_address','change_the_detail_work_address',
             'change_office_tel',
             'change_ip_address','change_mobile_brand']:
    appr_mart['change_item_counts'] = appr_mart['change_item_counts'] + appr_mart[item]
    repay_mart['change_item_counts'] = repay_mart['change_item_counts'] + repay_mart[item]

appr_mart['change_identification_item_counts'] = 0
repay_mart['change_identification_item_counts'] = 0
for item in ['change_id_type','change_id',
             'change_first_name','change_middle_name',
             'change_last_name','change_gender','change_birthday']:
    appr_mart['change_identification_item_counts'] = appr_mart['change_identification_item_counts'] + appr_mart[item]
    repay_mart['change_identification_item_counts'] = repay_mart['change_identification_item_counts']+ repay_mart[item]

appr_mart['change_personal_item_counts'] = 0
repay_mart['change_personal_item_counts'] = 0
for item in ['change_education','change_marital_status',
             'change_number_of_children','change_current_address',
             'change_length_of_residence',
             'change_the_facebook_account',
             'change_e-mail']:
    appr_mart['change_personal_item_counts'] = appr_mart['change_personal_item_counts'] + appr_mart[item]
    repay_mart['change_personal_item_counts'] = repay_mart['change_personal_item_counts']+ repay_mart[item]

appr_mart['change_contact_item_counts'] = 0
repay_mart['change_contact_item_counts'] = 0
for item in ['change_relationship01','change_contact_people01',
             'change_relationship02','change_contact_people02']:
    appr_mart['change_contact_item_counts'] = appr_mart['change_contact_item_counts'] + appr_mart[item]
    repay_mart['change_contact_item_counts'] = repay_mart['change_contact_item_counts']+ repay_mart[item]

appr_mart['change_employment_item_counts'] = 0
repay_mart['change_employment_item_counts'] = 0
for item in ['change_job_type','change_on-the-job_time',
             'change_ monthly_net_Income','change_the_ company_name',
             'change_work_address','change_the_detail_work_address',
             'change_office_tel']:
    appr_mart['change_employment_item_counts'] = appr_mart['change_employment_item_counts'] + appr_mart[item]
    repay_mart['change_employment_item_counts'] = repay_mart['change_employment_item_counts']+ repay_mart[item]

#开始时间与用时
event_time = b_d.groupby('phone_no',as_index=False)['event_date'].agg({'event_start_time':min,'event_end_time':max})
event_time['event_end_time'] = pd.to_datetime(event_time['event_end_time'])
event_time['event_start_time'] = pd.to_datetime(event_time['event_start_time'])
event_time['event_start_hour'] = event_time.event_start_time.dt.strftime('%H')
event_time['event_use_times'] = (event_time.event_end_time - event_time.event_start_time).dt.seconds
event_time['event_start_hour_grp'] = pd.cut(event_time.event_start_hour.astype(int),
                                     bins=[-1,5,8,17,22,100],
                                     labels=['0-5','6-8','9-17','18-22','23-0'])
event_time['event_use_times_grp'] = pd.cut(event_time.event_use_times,
                                      bins=[0,300,600,1200,1800,3600,np.inf],
                                      labels=['0-300','301-600','601-1200','1201-1800','1801-3600','3601-'])

appr_mart = pd.merge(appr_mart,event_time,how='left',on='phone_no')
repay_mart = pd.merge(repay_mart,event_time,how='left',on='phone_no')

#总事件数
event_cnt = b_d.groupby('phone_no',as_index=False)['event_name'].agg({'event_cnt':np.size})
appr_mart = pd.merge(appr_mart,event_cnt,how='left',on='phone_no')
repay_mart = pd.merge(repay_mart,event_cnt,how='left',on='phone_no')
#------------------
#%%--未提交申请客户的行为数据
apply_start_time = pd.read_sql("select apply_code, phone_no, \
                                       created_time as apply_start_time, \
                                       date_format(created_time,'%H') as apply_start_hour \
                               from suncash_lend.apply_info",cnx)
apply_now = apply_start_time.query("apply_start_time>'2019-11-21 00:00:00'")
apply_submit = apply_info.loc[:,['apply_code','apply_time','apply_month','apply_day']]
apply_now = pd.merge(apply_now,apply_submit,how='left',on='apply_code')
b_d = pd.merge(apply_now,b_d,how='inner',on='phone_no')
b_d = b_d.loc[pd.to_datetime(b_d.event_date)>b_d.apply_start_time,:]
b_d = b_d.loc[(pd.to_datetime(b_d.event_date)<=b_d.apply_time) | (b_d.apply_time.isnull()),:]

#b_d.sort_values(['apply_code','event_date'],inplace=True)
#last_event = b_d.drop_duplicates(subset='apply_code',keep='last')
#last_event['apply_start_month'] = last_event['apply_start_time'].dt.strftime('%Y%m')
#last_event['apply_start_day'] = last_event['apply_start_time'].dt.strftime('%Y%m%d')

#%%--看客户申请走到哪一步
event_order = pd.read_csv(save_dir+'\\'+'event_order.csv')
b_d = pd.merge(b_d,event_order,how='left',on='event_name')
b_d.sort_values(['apply_code','event_order'],inplace=True)
fast_event = b_d.drop_duplicates(subset='apply_code',keep='last')
fast_event['apply_start_month'] = fast_event['apply_start_time'].dt.strftime('%Y%m')
fast_event['apply_start_day'] = fast_event['apply_start_time'].dt.strftime('%Y%m%d')


#各信息模块（项）操作次数
for cate in ['click_identification','open_identification',
             'click_upload_id_photo','select_id_photo',
             'click_identification_back_button','submit_identification',
             'click_personal_information','open_personal_information',
             'click_the_personal_info_back_button','submit_personal_information',
             'click_contact_information','open_contact_information',
             'click_the_contact_info_back_ button','submit_contact_information',
             'click_employment_information','open_employment_information',
             'click_upload_proof_of_employment01','select_the_proof_of_employment01',
             'click_upload_proof_of_employment02','select_the_proof_of_employment02',
             'click_the_employment_info_back_button','submit_employment_information']:
    locals()[cate] = b_d.loc[b_d.event_name==cate,:].groupby('apply_code',as_index=False)['event_name'].agg({cate:np.size})

#各信息项不同取值数、是否有修改
item = ['select_id_type','leave_id_type_input_box',
        'leave_first_name_ input_box','leave_middle_name_ input_box',
        'leave_last_name_ input_box','select_gender','select_birthday',
        'select_education','select_marital_status',
        'select_number_of_children','select_current_address',
        'select_length_of_residence',
        'leave_the_facebook_account_input_box',
        'leave_e-mail_input_box',
        'select_relationship01','select_contact_people01',
        'select_relationship02','select_contact_people02',
        'select_job_type','select_on-the-job_time',
        'select_ monthly_net_Income','leave_the_ company_name',
        'select_work_address','leave_the_detail_work_address_input_box',
        'leave_office_tel_input_box']
item_change = ['change_id_type','change_id',
               'change_first_name','change_middle_name',
               'change_last_name','change_gender','change_birthday',
               'change_education','change_marital_status',
               'change_number_of_children','change_current_address',
               'change_length_of_residence',
               'change_the_facebook_account',
               'change_e-mail',
               'change_relationship01','change_contact_people01',
               'change_relationship02','change_contact_people02',
               'change_job_type','change_on-the-job_time',
               'change_ monthly_net_Income','change_the_ company_name',
               'change_work_address','change_the_detail_work_address',
               'change_office_tel']
for i,j in zip(item,item_change):
    locals()[i] = b_d.loc[b_d.event_name==i,:].groupby('apply_code',as_index=False)['event_data'].agg({i:pd.Series.nunique})
    locals()[i][j] = locals()[i][i].apply(lambda x: 1 if x>1 else 0)
 
#ip_address和mobile_brand不同取值数、是否有修改
item = ['ip_address','mobile_brand']
item_change = ['change_ip_address','change_mobile_brand']
for i,j in zip(item,item_change):
    locals()[i] = b_d.groupby('apply_code',as_index=False)[i].agg({i:pd.Series.nunique})
    locals()[i][j] = locals()[i][i].apply(lambda x: 1 if x>1 else 0)

#拼接到fast_event
for cate in ['click_identification','open_identification',
             'click_upload_id_photo','select_id_photo',
             'click_identification_back_button','submit_identification',
             'click_personal_information','open_personal_information',
             'click_the_personal_info_back_button','submit_personal_information',
             'click_contact_information','open_contact_information',
             'click_the_contact_info_back_ button','submit_contact_information',
             'click_employment_information','open_employment_information',
             'click_the_employment_info_back_button','submit_employment_information',
             'click_upload_proof_of_employment01','select_the_proof_of_employment01',
             'click_upload_proof_of_employment02','select_the_proof_of_employment02']:
    fast_event = pd.merge(fast_event,locals()[cate],how='left',on='apply_code')

for item in ['select_id_type','leave_id_type_input_box',
             'leave_first_name_ input_box','leave_middle_name_ input_box',
             'leave_last_name_ input_box','select_gender','select_birthday',
             'select_education','select_marital_status',
             'select_number_of_children','select_current_address',
             'select_length_of_residence',
             'leave_the_facebook_account_input_box',
             'leave_e-mail_input_box',
             'select_relationship01','select_contact_people01',
             'select_relationship02','select_contact_people02',
             'select_job_type','select_on-the-job_time',
             'select_ monthly_net_Income','select_work_address',
             'leave_the_detail_work_address_input_box',
             'leave_the_ company_name','leave_office_tel_input_box']:
    fast_event = pd.merge(fast_event,locals()[item],how='left',on='apply_code')

for j in ['ip_address','mobile_brand']:
    fast_event = pd.merge(fast_event,locals()[j],how='left',on='apply_code')

fast_event.to_excel(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\b_d.xlsx",index=False)
#%%--老客户推荐---------------------------------------------------------------------
recommend_cnt = pd.read_sql("select bind_reco_code, count(*) as recommend_cnt \
                            from suncash_lend.user_bind_recommend_info \
                            group by bind_reco_code",cnx)
recommend = pd.read_sql("select phone_no \
                        from suncash_lend.user_bind_recommend_info \
                        where bind_reco_code='9M99HH'",cnx)
recommend = pd.read_sql("select phone_no, bind_reco_code \
                        from suncash_lend.user_bind_recommend_info",cnx)

recommend = pd.merge(recommend,appr_mart,how='left',on='phone_no')
recommend = pd.merge(recommend,linkman_df,how='left',on='apply_code')

from openpyxl import load_workbook
writer = pd.ExcelWriter(r"C:\Users\lenovo\Desktop\recommend.xlsx",engine='openpyxl')
writer.book = load_workbook(r"C:\Users\lenovo\Desktop\recommend.xlsx")
idx = writer.book.sheetnames.index('Sheet1')
writer.book.remove(writer.book.worksheets[idx])
#writer.book.create_sheet('source',idx)
recommend.to_excel(writer,sheet_name='Sheet1',index=False)
writer.save()

#%%--免核变量
mh_contact = pd.read_sql("select apply_code, contacts_relation, \
                                 phone_check_result_code \
                        from approval.contact_info",cnx)
mh_client = pd.read_sql("select apply_code, phone_check_result_code \
                        from approval.borrower_info",cnx)
mh_office = pd.read_sql("select apply_code, phone_check_result_code \
                        from approval.employment_info",cnx)

mh_contact['mh_contact'] = np.where(mh_contact['phone_check_result_code'].isin(['14100','14080','14090','14110']),1,0)
mh_contact['mh_contact_gps_company'] = np.where(mh_contact['phone_check_result_code']=='14100',1,0)
mh_contact['mh_contact_gps_home'] = np.where(mh_contact['phone_check_result_code']=='14110',1,0)
mh_contact['mh_contact_gps_consistent'] = np.where(mh_contact['phone_check_result_code']=='14080',1,0)
mh_contact['mh_contact_good_quality'] = np.where(mh_contact['phone_check_result_code']=='14090',1,0)
mh_contact_r = mh_contact.groupby('apply_code',as_index=False)['mh_contact',
                                                               'mh_contact_gps_company',
                                                               'mh_contact_gps_home',
                                                               'mh_contact_gps_consistent',
                                                               'mh_contact_good_quality'].sum()

mh_client['mh_client'] = np.where(mh_client['phone_check_result_code'].isin(['14100','14080','14090','14110']),1,0)
mh_client['mh_client_gps_company'] = np.where(mh_client['phone_check_result_code']=='14100',1,0)
mh_client['mh_client_gps_home'] = np.where(mh_client['phone_check_result_code']=='14110',1,0)
mh_client['mh_client_gps_consistent'] = np.where(mh_client['phone_check_result_code']=='14080',1,0)
mh_client['mh_client_good_quality'] = np.where(mh_client['phone_check_result_code']=='14090',1,0)
mh_client_r = mh_client.groupby('apply_code',as_index=False)['mh_client',
                                                             'mh_client_gps_company',
                                                             'mh_client_gps_home',
                                                             'mh_client_gps_consistent',
                                                             'mh_client_good_quality'].sum()

mh_office['mh_office'] = np.where(mh_office['phone_check_result_code'].isin(['14100','14080','14090','14110']),1,0)
mh_office['mh_office_gps_company'] = np.where(mh_office['phone_check_result_code']=='14100',1,0)
mh_office['mh_office_gps_home'] = np.where(mh_office['phone_check_result_code']=='14110',1,0)
mh_office['mh_office_gps_consistent'] = np.where(mh_office['phone_check_result_code']=='14080',1,0)
mh_office['mh_office_good_quality'] = np.where(mh_office['phone_check_result_code']=='14090',1,0)
mh_office_r = mh_office.groupby('apply_code',as_index=False)['mh_office',
                                                             'mh_office_gps_company',
                                                             'mh_office_gps_home',
                                                             'mh_office_gps_consistent',
                                                             'mh_office_good_quality'].sum()

mh = pd.merge(mh_client_r,mh_contact_r,how='left',on='apply_code')
mh = pd.merge(mh,mh_office_r,how='left',on='apply_code')
mh['mh_cnt'] = mh['mh_contact'] + mh['mh_client'] + mh['mh_office']

repay_mart = pd.merge(repay_mart,mh,how='left',on='apply_code')

##PH_DEMO-----------------------------------------------------------------------
#from openpyxl import load_workbook
#
##repay
#repay_1st_apply = repay_mart.query("loan_type=='first_apply'")
#
#columns = repay_1st_apply.columns
#ignore_feature = ['contract_no', 'account_status', 'loan_date', 'loan_month', 'loan_day',
#        'clear_date', 'loan_term', 'last_repay_date', 'contract_amount',
#        'loan_amount', 'borrower_tel_one', 'apply_code',
#        'user_code', 'loan_type', 'number_of_loans', 'service_fee_rate',
#        'od_days_ever', 'od_days', 'first_loan_month', 'last_loan', '自然逾期',
#        '到期', '曾经逾期7天以上', '到期7天以上', '逾期还款次数', '曾经最大逾期天数', 'id_no',
#        'home_city', 'home_district',
#        'device_id', 'phone_no', 
#        'job_city', 'job_district', 'number_of_apply'
#        ]
#feature = [f for f in columns if f not in ignore_feature]
#
#writer = pd.ExcelWriter(r"C:\Users\lenovo\Document\TS\08产品\菲律宾\dataMart\PH_Demo.xlsx",engine='openpyxl')
#
#month = ['total','201912','201911','201910']
#for m in month:
#    if m == 'total':
#        stat_df = repay_1st_apply
#    else:
#        stat_df = repay_1st_apply.loc[repay_1st_apply.loan_month==m,:]
#    for i, f in enumerate(feature):
#        f_stat = stat_df.groupby(f,as_index=False)['到期','自然逾期','到期7天以上','曾经逾期7天以上'].sum()
#        f_stat.rename(columns={f:'group'},inplace=True)
#        f_stat.insert(0,'feature', f)
#        if i == 0:
#            feature_stat = f_stat
#        else:
#            feature_stat = pd.concat([feature_stat,f_stat],axis=0)
#    od_rate = (feature_stat['自然逾期']/feature_stat['到期']).apply(lambda x:format(x,'.2%'))
#    feature_stat.insert(4,'自然逾期率',od_rate)
#    feature_stat['曾经7天以上逾期率'] = (feature_stat['曾经逾期7天以上']/feature_stat['到期7天以上']).apply(lambda x:format(x,'.2%'))
#    if m == 'total':
#        feature_stat.to_excel(writer,sheet_name='repay_total',index=False)
#    else:
#        feature_stat.to_excel(writer,sheet_name='repay_'+m,index=False)
#        
#writer.save()

#%%--PH_DEMO-----------------------------------------------------------------------
from openpyxl import load_workbook

feature = pd.read_csv(r"PH_Demo\feature_list.csv",encoding='gbk')

repay_1st_apply = repay_mart.query("loan_type=='first_apply'")
appr_1st_apply = appr_mart.query("number_of_apply==1")
appr_1st_apply.rename(columns={'payment_type':'loan_channel'},inplace=True)


strdate = datetime.date.today().strftime('%Y%m%d')
fn = r"PH_Demo\PH_Demo_%s.xlsx" % strdate
#writer = pd.ExcelWriter(fn,engine='openpyxl')
writer = pd.ExcelWriter(fn)
workbook = writer.book
font_fmt = workbook.add_format({"font_name":"Arial Unicode MS","font_size":9})
bg_fmt = workbook.add_format({'bg_color': '#9FC3D1',"font_name":"Arial Unicode MS","font_size":9})

month = ['total','202002','202001','201912']
#month = ['total']
#各个月份单独一个sheet
#for m in month:
#    repay_stat = pd.DataFrame()
#    appr_stat = pd.DataFrame()
#
#    if m == 'total':
#        repay_df = repay_1st_apply
#        appr_df = appr_1st_apply
#    else:
#        repay_df = repay_1st_apply.loc[repay_1st_apply.loan_month==m,:]
#        appr_df = appr_1st_apply.loc[appr_1st_apply.apply_month==m,:]
#    for i in range(feature.shape[0]):
#        try:
#            f = feature.iloc[i,0]
#            repay_stat_f = repay_df.groupby(f,as_index=False)['到期','自然逾期','到期7天以上','曾经逾期7天以上'].sum()
#            repay_stat_f.rename(columns={f:'group'},inplace=True)
#            repay_stat_f.insert(0,'feature', f)
#            appr_stat_f = appr_df.groupby(f,as_index=False)['处理','通过','人工处理','人工通过'].sum()
#            appr_stat_f.rename(columns={f:'group'},inplace=True)
#            appr_stat_f.insert(0,'feature', f)        
#            if i == 0:
#                repay_stat = repay_stat_f
#                appr_stat = appr_stat_f
#            else:
#                repay_stat = pd.concat([repay_stat,repay_stat_f],axis=0)
#                appr_stat = pd.concat([appr_stat,appr_stat_f],axis=0)
#        except Exception as e:
#            print("Error:",e)
#    rate = (repay_stat['到期']/len(repay_df.query("到期==1"))).apply(lambda x:format(x,'.2%')) 
#    repay_stat.insert(3,'到期占比',rate)
#    od_rate = (repay_stat['自然逾期']/repay_stat['到期']).apply(lambda x:format(x,'.2%'))
#    repay_stat.insert(5,'自然逾期率',od_rate)
#    repay_stat['曾经7天以上逾期率'] = (repay_stat['曾经逾期7天以上']/repay_stat['到期7天以上']).apply(lambda x:format(x,'.2%'))
#    appr_rate = (appr_stat['通过']/appr_stat['处理']).apply(lambda x:format(x,'.2%'))
#    appr_stat.insert(4,'通过率',appr_rate)
#    appr_stat['人工通过率'] = (appr_stat['人工通过']/appr_stat['人工处理']).apply(lambda x:format(x,'.2%'))
#    if m == 'total':
#        repay_stat.to_excel(writer,sheet_name='repay_total',index=False)
#        appr_stat.to_excel(writer,sheet_name='appr_total',index=False)
#    else:
#        repay_stat.to_excel(writer,sheet_name='repay_'+m,index=False)
#        appr_stat.to_excel(writer,sheet_name='appr_'+m,index=False)
#writer.save()

#各个月份拼到一个sheet
repay_stat_all = pd.DataFrame()
appr_stat_all = pd.DataFrame()

for m in month:
    repay_stat = pd.DataFrame()
    appr_stat = pd.DataFrame()

    if m == 'total':
        repay_df = repay_1st_apply
        appr_df = appr_1st_apply
    else:
        repay_df = repay_1st_apply.loc[repay_1st_apply.loan_month.astype(str)==m,:]
        appr_df = appr_1st_apply.loc[appr_1st_apply.apply_month.astype(str)==m,:]
    for i in range(feature.shape[0]):
        try:
            f = feature.iloc[i,0]
            repay_stat_f = repay_df.groupby(f,as_index=False)['到期','自然逾期','到期7天以上','曾经逾期7天以上'].sum()
            repay_stat_f.rename(columns={f:'group'},inplace=True)
            repay_stat_f.insert(0,'feature', f)
            appr_stat_f = appr_df.groupby(f,as_index=False)['处理','通过','人工处理','人工通过'].sum()
            appr_stat_f.rename(columns={f:'group'},inplace=True)
            appr_stat_f.insert(0,'feature', f)        
            if i == 0:
                repay_stat = repay_stat_f
                appr_stat = appr_stat_f
            else:
                repay_stat = pd.concat([repay_stat,repay_stat_f],axis=0)
                appr_stat = pd.concat([appr_stat,appr_stat_f],axis=0)
        except Exception as e:
            print("Error:",e)
    rate = (repay_stat['到期']/len(repay_df.query("到期==1"))).apply(lambda x:format(x,'.2%')) 
    repay_stat.insert(3,'到期占比',rate)
    od_rate = (repay_stat['自然逾期']/repay_stat['到期']).apply(lambda x:format(x,'.2%'))
    repay_stat.insert(5,'自然逾期率',od_rate)
    repay_stat['曾经7天以上逾期率'] = (repay_stat['曾经逾期7天以上']/repay_stat['到期7天以上']).apply(lambda x:format(x,'.2%'))
    appr_rate = (appr_stat['通过']/appr_stat['处理']).apply(lambda x:format(x,'.2%'))
    appr_stat.insert(4,'通过率',appr_rate)
    appr_stat['人工通过率'] = (appr_stat['人工通过']/appr_stat['人工处理']).apply(lambda x:format(x,'.2%'))
    
    repay_stat.rename(columns={'到期':'到期_'+m,'到期占比':'到期占比_'+m,
                               '自然逾期':'自然逾期_'+m,
                               '自然逾期率':'自然逾期率_'+m,
                               '到期7天以上':'到期7天以上_'+m,
                               '曾经逾期7天以上':'曾经逾期7天以上_'+m,
                               '曾经7天以上逾期率':'曾经7天以上逾期率_'+m},
                        inplace=True)
    repay_stat['id'] = repay_stat.apply(lambda x: x.feature+'_'+str(x.group),axis=1)
    k = [c for c in appr_stat.columns if c not in ['feature','group']]
    v = [i+'_'+m for i in k]
    appr_stat.rename(columns=dict(zip(k,v)),inplace=True)
    appr_stat['id'] = appr_stat.apply(lambda x: x.feature+'_'+str(x.group),axis=1)

    if m == 'total':
        repay_stat_all = repay_stat.copy()
        appr_stat_all = appr_stat.copy()
    else:
        repay_stat.drop(columns={'feature','group'},inplace=True)
        appr_stat.drop(columns={'feature','group'},inplace=True)
        repay_stat_all = pd.merge(repay_stat_all,repay_stat,how='left',on='id')
        appr_stat_all = pd.merge(appr_stat_all,appr_stat,how='left',on='id')

repay_stat_all.drop(columns={'id'},inplace=True)
appr_stat_all.drop(columns={'id'},inplace=True)

kp1 = ['到期占比','自然逾期率','曾经7天以上逾期率','到期','自然逾期','到期7天以上','曾经逾期7天以上']
col_seq = ['feature','group']
for i in kp1:
    for j in month:
        col_seq.append(i+'_'+j)
repay_stat_all = repay_stat_all[col_seq]
kp2 = ['通过率','人工通过率','处理','通过','人工处理','人工通过']
col_seq = ['feature','group']
for i in kp2:
    for j in month:
        col_seq.append(i+'_'+j)
appr_stat_all = appr_stat_all[col_seq]

repay_stat_all.to_excel(writer,sheet_name='repay',index=False)
appr_stat_all.to_excel(writer,sheet_name='appr',index=False)
worksheet1 = writer.sheets['repay']
worksheet1.set_column('A:AD',15,font_fmt)
worksheet1.set_column('C:F',15,bg_fmt)
worksheet1.set_column('K:N',15,bg_fmt)
worksheet1.set_column('S:V',15,bg_fmt)
worksheet1.set_column('AA:AD',15,bg_fmt)
worksheet2 = writer.sheets['appr']
worksheet2.set_column('A:AD',15,font_fmt)
worksheet2.set_column('C:F',15,bg_fmt)
worksheet2.set_column('K:N',15,bg_fmt)
worksheet2.set_column('S:V',15,bg_fmt)

writer.save()

#%%--客群分类-------------------------------------------------
def risk_group(x):
    if x.age>=43 and x.gender=='Female' and x.loan_channel=='DRAGON_PAY':
        return '客群1'
    elif x.age>=43 and x.gender=='Female' and x.loan_channel=='SKY_PAY':
        return '客群2'
    elif x.age>=35 and x.gender=='Female' and x.education in(['College Graduate','Master/PHD']):
        return '客群3'
    elif x.age>=35 and x.gender=='Female':
        return '客群4'
    elif x.age>=43 and x.gender=='Male' and x.loan_channel=='DRAGON_PAY':
        return '客群5'
    elif x.age>=43 and x.gender=='Male':
        return '客群6'
    elif x.age>=35 and x.gender=='Male':
        return '客群7'
    elif x.age<35 and x.gender=='Female' and x.education in(['College Graduate','Master/PHD']) and x.job_type!='Private Company Employee':
        return '客群8'
    else:
        return '客群9'

repay_mart['风险定价客群'] = repay_mart.apply(risk_group,axis=1)
repay_mart.to_excel(r"repay_mart.xlsx",index=False)

appr_mart.rename(columns={'payment_type':'loan_channel'},inplace=True)
appr_mart['风险定价客群'] = appr_mart.apply(risk_group,axis=1)

#%%--手机号所属运营商------------------------------------------------
def phone_operator(x):
    if x[:4] in (['0922','0923','0925','0931','0932','0933','0934','0942',
                '0943','0944']):
        return 'SUN'
    elif x[:4] in (['0817','0905','0906','0915','0916','0917','0926','0927',
                  '0935','0936','0937','0955','0956','0975','0977','0994',
                  '0995','0996']):
        return 'Globe'
    else:
        return 'Smart'

repay_mart['phone_operator'] = repay_mart['phone_no'].apply(phone_operator)
appr_mart['phone_operator'] = appr_mart['phone_no'].apply(phone_operator)

#%%--模型计算-------------------------------------
def lr_pred(x):
    z = 0.3651 \
        - 0.5726*int(x.gender=='Female') \
        + 0.3692*int(x.age<35) \
        - 0.3611*int(x.age>42) \
        - 0.4313*int(x.education in (['College Graduate','Master/PHD'])) \
        + 0.1515*int(x.marriage=='Single') \
        - 0.3516*int(x.id_type in (['PRC ID','SSS CARD'])) \
        + 0.3731*int(x.job_type in (['Private Company Employee',
                                     'Self-employed household',
                                     'BPO Professionals'])) \
        - 0.4325*int(x.loan_channel in (['DRAGON_PAY','SKY_PAY_BANK']))
    return math.exp(z)/(1+math.exp(z))
        
repay_mart['lr_pred'] = repay_mart.apply(lr_pred,axis=1)
repay_mart['lr_pred_grp'] = pd.cut(repay_mart.lr_pred,
                                  bins=[0,0.25,0.3,0.4,0.5,0.6,np.inf],
                                  labels=['A','B','C','D','E','F'])

churn_mart.info()
#%%--客户流失桑基图----------------------------------
from pyecharts.charts import Sankey
from pyecharts import options as opts
import saspy 
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
repay_mart=sas.sd2df('repay_mart','MY')
repay_mart["number_of_loans"] = repay_mart["number_of_loans"].astype(int)
repay_mart["到期"] = repay_mart["到期"].astype(int)
repay_mart["自然逾期"] = repay_mart["自然逾期"].astype(int)
repay_mart["催回"] = repay_mart["催回"].astype(int)


# repay_mart = pd.read_excel("repay_mart.xlsx")
def settle_status(x):
    if x.到期==0:
        return '未到期'
    elif x.自然逾期==0:
        return '正常结清'
    elif x.催回==1:
        return '逾期催回'
    else:
        return '逾期中'
        
repay_mart['结清状态'] = repay_mart.apply(settle_status,axis=1)
churn_mart = repay_mart.loc[:,['borrower_tel_one','number_of_loans','结清状态']]
init_nodes = []
links = []
def churn(times):
    for i in range(times):
        if i==0:
            s_t = churn_mart.loc[churn_mart.number_of_loans==1,:]
            stv = s_t.groupby('结清状态').size()
            init_nodes.append({"name":"begin"})
            for j in range(len(stv)):
                init_nodes.append({"name":str(i+1)+stv.index[j]})
                dic = {}
                dic['source'] = 'begin'
                dic['target'] = str(i+1)+stv.index[j]
                dic['value'] = int(stv.values[j])
                links.append(dic)
        else:
            source = churn_mart.loc[churn_mart.number_of_loans==i,:]
            source.rename(columns={'结清状态':'结清状态source'},inplace=True)
            target = churn_mart.loc[churn_mart.number_of_loans==i+1,:]
            target.rename(columns={'结清状态':'结清状态target'},inplace=True)
            s_t = pd.merge(source,target,how='left',on='borrower_tel_one')
            s_t.fillna('流失',inplace=True)
            stv = pd.pivot_table(s_t,
                                 index=['结清状态source','结清状态target'],
                                 aggfunc=np.size,
                                 fill_value='未放款')
            for j in range(len(stv)):
                init_nodes.append({"name":str(i+1)+stv.index[j][1]})
                dic = {}
                dic['source'] = str(i)+stv.index[j][0]
                dic['target'] = str(i+1)+stv.index[j][1]
                dic['value'] = int(stv.iloc[j,0])
                links.append(dic)

churn(6)
nodes = []
for n in init_nodes:
    if n not in nodes:
        nodes.append(n)
nodes
links

pic = (
    Sankey()
    .add('', #图例名称
         nodes,    #传入节点数据
         links,   #传入边和流量数据
         #设置透明度、弯曲度、颜色
         linestyle_opt=opts.LineStyleOpts(opacity = 0.3, curve = 0.5, color = "source"),
         #标签显示位置
         label_opts=opts.LabelOpts(position="right"),
         #节点之前的距离
         node_gap = 30,
    )
    .set_global_opts(title_opts=opts.TitleOpts(title = '客户流失桑基图'))
)
pic.render('客户流失桑基图.html')

#每轮客户漏斗图
from pyecharts.charts import Funnel
funnel = repay_mart.groupby('number_of_loans').size().reset_index()
c = (
    Funnel()
    .add("贷款次数", 
         [list(z) for z in zip(funnel.iloc[:,0], funnel.iloc[:,1])],
         label_opts=opts.LabelOpts(position="inside"),)
    .set_global_opts(title_opts=opts.TitleOpts(title="每轮客户漏斗图"))
    .render("每轮客户漏斗图.html")
)            
    











#%%--词云图
from pyecharts.charts import WordCloud

words = [
('数据分析',50),
('零售信贷',50),
('信用风险',40),
('信用评分模型',35),
('风控策略',45),
('风控模型',45),
('反欺诈',40),
('征信',35),
('大数据',35),
('规则引擎',35),
('决策引擎',35),
('Python',50),
('SAS',50),
('Mysql',35),
('MongoDB',35),
('Amazon S3',35),
('阿里云oss',35),
('Scikit-learn',40),
('DecisionTree',40),
('LogisticRegression',40),
('RandomForest',35),
('GradientBoosting',35),
('SVM',35),
('NEO4j',35)
]
WordCloud().add("",words).render(r"C:\Users\tsjr\Desktop\a.html")

#疫情以来的还款客户
f = 'account.trans_journal_detail.csv'
locals()[f.split('.')[1]] = pd.read_csv(f,dtype=dtype)

trans_journal_detail['created_time'] = pd.to_datetime(trans_journal_detail['created_time'])
recentRepayDtl = trans_journal_detail.query("created_time>='2020-03-16' and \
                                             trans_status=='DEAL_SUCCESS'")
recentRepayDtl['repay_date'] = recentRepayDtl['created_time'].dt.date
piv = pd.pivot_table(recentRepayDtl,
                     values='trans_amount',
                     index='repay_date',
                     columns='repayment_store_name',
                     aggfunc='sum',
                     margins=True,
                     fill_value=0)
recentRepayAmt = recentRepayDtl.groupby('contract_no',as_index=False).agg({'id':np.size,'trans_amount':np.sum,'repay_date':max})
recentRepayAmt.rename(columns={'id':'还款次数','repay_date':'最后还款日'},inplace=True)
recentRepay = pd.merge(recentRepayAmt,repay_mart,on='contract_no',how='left')
recentRepay.to_excel(r"C:\Users\tsjr\Desktop\recentRepay.xlsx",index=False)

recentRepayDtl.sort_values(by=['contract_no','created_time'],inplace=True)
recentRepayDtl['还款次序'] = recentRepayDtl['created_time'].groupby(recentRepayDtl['contract_no']).rank(method='dense')
recentRepayDtl['last_repay_date'] = recentRepayDtl['repay_date'].shift(1)
recentRepayDtl['还款间隔'] = (recentRepayDtl['repay_date'] - recentRepayDtl['last_repay_date']).dt.days
recentRepayDtl['还款间隔'] = recentRepayDtl.apply(lambda x: np.nan if x.还款次序==1 else x.还款间隔,axis=1)
recentRepayDtl.groupby('还款间隔').size()
#%%--客户实际还款金额
account_info['interest'] = account_info.apply(lambda x: x.contract_amount*x.loan_term*0.01, axis=1)
actualRepayAmt = trans_journal_detail.groupby('contract_no',as_index=False)['trans_amount'].agg({'actualRepayAmt':sum})
repayAmt = pd.merge(account_info,actualRepayAmt,on='contract_no',how='left')
repayAmt['已还够到手'] = repayAmt.apply(lambda x: 1 if x.actualRepayAmt>=x.loan_amount else 0, axis=1)
repayAmt['已还够合同'] = repayAmt.apply(lambda x: 1 if x.actualRepayAmt>=x.contract_amount else 0, axis=1)
repayAmt['已还够合同与利息'] = repayAmt.apply(lambda x: 1 if x.actualRepayAmt>=x.contract_amount+x.interest else 0, axis=1)
repayAmt['分母'] = 1
repayAmt['逾期'] = repayAmt['account_status'].apply(lambda x: 0 if x=='ACCOUNT_SETTLE' else 1)
repayAmt.to_excel(r"C:\Users\tsjr\Desktop\repayAmt.xlsx",index=False)


#%%--认证后未申请客户分析
#注册
reg = user.loc[:,['created_time','channel_id','phone_no','user_code']]
reg['reg_time'] = pd.to_datetime(reg['created_time'])
reg.drop(columns=['created_time'],inplace=True)
reg['reg_day'] = reg['reg_time'].dt.strftime('%Y-%m-%d')
reg['reg_month'] = reg['reg_time'].dt.strftime('%Y-%m')
reg = reg.query("reg_time>='2020-09-01'")

#认证
IDENTIFICATION = user_verification_info.loc[user_verification_info.verify_type==1,['user_code']] #身份证信息
IDENTIFICATION['完成身份认证'] = 1
PERSONAL = user_verification_info.loc[user_verification_info.verify_type==2,['user_code']] #个人信息
PERSONAL['完成个人认证'] = 1
CONTACT = user_verification_info.loc[user_verification_info.verify_type==3,['user_code']] #联系人信息
CONTACT['完成联系人认证'] = 1
EMPLOYMENT = user_verification_info.loc[user_verification_info.verify_type==4,['user_code']] #工作信息
EMPLOYMENT['完成工作认证'] = 1
PAYMENT = user_verification_info.loc[user_verification_info.verify_type==5,['user_code','created_time']] #支付信息
PAYMENT['payment_time'] = pd.to_datetime(PAYMENT['created_time'])
PAYMENT.sort_values(by=['user_code','payment_time'],inplace=True)
PAYMENT.drop_duplicates(subset=['user_code'],inplace=True)
PAYMENT['完成认证'] = 1
PAYMENT.drop(columns=['created_time'],inplace=True)

payment_reg = pd.merge(reg,PAYMENT,on='user_code',how='inner')

apply = apply_info.loc[:,['user_code','face_photo_status','face_photo_key','created_time']]
apply['apply_time'] = pd.to_datetime(apply['created_time'])
apply.sort_values(by=['user_code','apply_time'],inplace=True)
apply.drop_duplicates(subset=['user_code'],keep='first',inplace=True)
apply['申请'] = 1
apply.drop(columns=['created_time'],inplace=True)

payment_reg_apply = pd.merge(payment_reg,apply,how='left',on='user_code')
payment_reg_apply['认证后当天申请'] = payment_reg_apply.apply(lambda x: 1 if (x.apply_time-x.payment_time).days==0 else 0,axis=1)
payment_reg_apply['认证后2天内申请'] = payment_reg_apply.apply(lambda x: 1 if (x.apply_time-x.payment_time).days<2 else 0,axis=1)
payment_reg_apply['认证后3天内申请'] = payment_reg_apply.apply(lambda x: 1 if (x.apply_time-x.payment_time).days<3 else 0,axis=1)

identification_info = pd.read_sql("select user_code, id_type_en_name, \
                                          gender_en_name, birth_date \
                                   from suncash_lend.identification",cnx)
personal_info = pd.read_sql("select user_code, childern_num_en_name, \
                                    education_en_name, marital_status_en_name, \
                                    residence_city_en_name, residence_provience_en_name, \
                                    residence_time_en_name \
                            from suncash_lend.personal_info",cnx)
employment_info = pd.read_sql("select user_code, job_income_en_name, \
                                      job_time_en_name, job_type_en_name, \
                                      work_city_en_name, work_provience_en_name \
                              from suncash_lend.employment_info",cnx)
payment_bind_info = pd.read_sql("select user_code, pay_type_name \
                                from suncash_lend.payment_bind_info",cnx)

aeid = af_event_info_detail.loc[:,['apps_flyer_id','media_source','channel','campaign']]
ub = user_bind_recommend_info.loc[:,['phone_no','bind_reco_code']]

identification_info.drop_duplicates('user_code',inplace=True)
personal_info.drop_duplicates('user_code',inplace=True)
employment_info.drop_duplicates('user_code',inplace=True)
payment_bind_info.drop_duplicates('user_code',inplace=True)
aeid.drop_duplicates('apps_flyer_id',inplace=True)

user_conv = pd.merge(payment_reg_apply,identification_info,how='left',on='user_code')
user_conv = pd.merge(user_conv,personal_info,how='left',on='user_code')
user_conv = pd.merge(user_conv,employment_info,how='left',on='user_code')
user_conv = pd.merge(user_conv,payment_bind_info,how='left',on='user_code')
user_conv = pd.merge(user_conv,aeid,how='left',left_on='channel_id',right_on='apps_flyer_id')
user_conv = pd.merge(user_conv,ub,how='left',on='phone_no')

def channel(x):
    if pd.isnull(x.bind_reco_code)==False:
        return '推荐码'
    elif pd.isnull(x.media_source):
        return '自然流量'
    elif x.channel=='Facebook' and 'inmobiagency' in x.campaign:
        return 'inmobiagency'
    elif x.media_source=='restricted':
        return 'Facebook Ads'
    else:
        return x.media_source

user_conv['渠道'] = user_conv.apply(channel,axis=1)
user_conv['age'] = user_conv.apply(lambda x: datetime.date.today().year - int(x.birth_date[:4]),axis=1)

credit_limit = pd.read_sql("select user_code, contract_amount_max, \
                                   contract_amount_min, service_fee_rate \
                            from suncash_lend.credit_limit",cnx)
credit_limit.drop_duplicates('user_code',inplace=True)
user_conv = pd.merge(user_conv,credit_limit,how='left',on='user_code')
user_conv['调整人群'] = user_conv.apply(lambda x: 1 if x.id_type_en_name=='PRC ID' or x.education_en_name=='Master/PHD' or x.job_type_en_name=='Teacher' else 0,axis=1)
user_conv.to_excel(r"C:\Users\tsjr\Desktop\user_conv.xlsx",index=False)
    
#推荐码客户自动拒绝分析
reco_code = user_conv.loc[user_conv.渠道=='推荐码',['user_code']]
reco_code = pd.merge(reco_code,appr_mart,how='inner',on='user_code')

#%%--白名单
whitelist = pd.read_sql("select * from approval.white_list_info",cnx)
white = whitelist.loc[:,['value','created_time']].rename(columns={'value':'phone_no'})
white['白名单日期'] = white['created_time'].dt.strftime('%Y-%m-%d')
white.drop(columns=['created_time'],inplace=True)
first_apply_month = (
                    appr_mart.loc[:,['phone_no','apply_month','apply_time']]
                    .sort_values(by=['phone_no','apply_time'])
                    .drop_duplicates(subset='phone_no')
                    .drop(columns='apply_time')
                    .rename(columns={'apply_month':'first_apply_month'})
                    )
first_loan_month = (
                    repay_mart.loc[repay_mart.loan_type=='first_apply',['phone_no','loan_month']]
                    .rename(columns={'loan_month':'first_loan_month'})
                    )
white = pd.merge(white,first_apply_month,how='left',on='phone_no')
white = pd.merge(white,first_loan_month,how='left',on='phone_no')
#未在我司申请过的白名单
white1 = white.loc[(white.first_apply_month=='202006') | (pd.isnull(white.first_apply_month)),:]
white1.shape  #不重叠白名单客户数
white1.info() #申请的客户数
white1_apply = pd.merge(white1,appr_mart,how='inner',on='phone_no')
white1_apply.groupby(['approval_status','loan_type']).size().unstack() #申请的订单批核情况
white1_apply.loc[white1_apply.approval_status=='MANUAL_APPROVED',:].phone_no.nunique() #批核的客户数
#未在我司放款过的白名单
white2 = white.loc[(white.first_loan_month=='202006') | (pd.isnull(white.first_loan_month)),:]
white2.shape  #不重叠白名单客户数
white2_apply = pd.merge(white2,appr_mart,how='inner',on='phone_no') 
white2_apply.loc[white2_apply.apply_month=='202006',:].phone_no.nunique() #申请的客户数
white2_apply.loc[white2_apply.apply_month=='202006',:].groupby(['approval_status','loan_type']).size().unstack() #申请的订单批核情况
white2_apply.loc[(white2_apply.apply_month=='202006') & (white2_apply.approval_status=='MANUAL_APPROVED'),:].phone_no.nunique() #批核的客户数


white_apply = pd.merge(white,appr_mart,how='left',on='phone_no')
white_apply.to_excel(r"C:\Users\tsjr\Desktop\white_apply.xlsx",index=False)

white['白名单'] = 1
repay_mart = pd.merge(repay_mart,white,how='left',on='phone_no')

#白名单注册情况
reg['注册'] = 1
white_reg = pd.merge(white,reg,on='phone_no')
white_reg.groupby(['reg_month','白名单日期']).size().unstack().fillna("")
white_reg.query("reg_month=='2020-06'").groupby(['白名单日期']).size()
#白名单认证情况
white_payment = pd.merge(white_reg,PAYMENT,on='user_code')
white_payment.groupby(['reg_month','白名单日期']).size().unstack().fillna("")
white_payment.query("reg_month=='2020-06'").groupby(['白名单日期']).size()
#白名单申请情况
apply = apply_info.loc[:,['user_code','face_photo_status','face_photo_key','created_time']]
apply['apply_time'] = pd.to_datetime(apply['created_time'])
apply.sort_values(by=['user_code','apply_time'],inplace=True)
apply.drop_duplicates(subset=['user_code'],keep='last',inplace=True)
apply['申请'] = 1
apply.drop(columns=['created_time'],inplace=True)
white_apply = pd.merge(white_reg,apply,on='user_code')
white_apply.query("apply_time>='2020-05-28'").groupby(['白名单日期']).size()
white_apply.query("reg_month=='2020-06'").groupby(['白名单日期']).size()
#白名单审批情况
appr = appr_mart.loc[:,['user_code','apply_time','apply_month','loan_type',
                        'approval_status','auto_refuse_reason','refuse_info_1',
                        'cancle_info_1','apply_code']]
white_appr = pd.merge(white_reg,appr,on='user_code')
white_appr.query("apply_time>='2020-05-28'").groupby(['白名单日期','approval_status','loan_type']).size().unstack().fillna("")
white_appr.query("reg_month=='2020-06'").groupby(['白名单日期','approval_status','loan_type']).size().unstack().fillna("")
#白名单还款情况
repay = repay_mart.loc[:,['apply_code','放款','到期','自然逾期','催回','当前逾期']]
white_repay = pd.merge(white_appr,repay,on='apply_code')
white_repay.to_excel(r"C:\Users\tsjr\Desktop\white_repay.xlsx",index=False)

#------单激活情况
appr_s1 = appr_mart.query("apply_time<='2020-06-15' and apply_time>='2020-05-10'")
appr_s2 = appr_mart.query("apply_time<='2020-06-30' and apply_time>='2020-06-16'")
appr_st = appr_mart.query("apply_time>='2020-05-10'")


invite1 = pd.read_clipboard(dtype={'phone_no':str})

invite1_appr_st = pd.merge(invite1,appr_st,on='phone_no')
invite1_appr_st.groupby('listCategory').phone_no.nunique()
invite1_appr_st.groupby('listCategory').size()

invite1_appr_s1 = pd.merge(invite1,appr_s1,on='phone_no')
invite1_appr_s1.groupby('listCategory').phone_no.nunique()
phone1 = invite1_appr_s1.loc[:,['phone_no','listCategory']]
phone1.drop_duplicates(subset='phone_no',inplace=True)
phone1.shape
phone1_apply = pd.merge(phone1,appr_st,on='phone_no')
phone1_apply.groupby('listCategory').phone_no.nunique()
phone1_apply.groupby('listCategory').size()


whitelist = pd.read_sql("select * from approval.white_list_info",cnx)
white = whitelist.loc[:,['value','created_time']].rename(columns={'value':'phone_no'})
white['白名单日期'] = white['created_time'].dt.strftime('%Y-%m-%d')
white.drop(columns=['created_time'],inplace=True)
invite2 = white.loc[white.白名单日期=='2020-05-28',:]

invite2_appr_st = pd.merge(invite2,appr_st,on='phone_no')
invite2_appr_st.phone_no.nunique()
invite2_appr_st.shape

invite2_appr_s1 = pd.merge(invite2,appr_s1,on='phone_no')
invite2_appr_s1.phone_no.nunique()
phone1 = invite2_appr_s1.loc[:,['phone_no']]
phone1.drop_duplicates(subset='phone_no',inplace=True)
phone1.shape
phone1_apply = pd.merge(phone1,appr_st,on='phone_no')
phone1_apply.phone_no.nunique()
phone1_apply.shape


white1 = pd.read_clipboard(dtype={'phone_no':str})

white1_appr_st = pd.merge(white1,appr_st,on='phone_no')
white1_appr_st.phone_no.nunique()
white1_appr_st.shape

white1_appr_s1 = pd.merge(white1,appr_s1,on='phone_no')
white1_appr_s1.phone_no.nunique()
phone1 = white1_appr_s1.loc[:,['phone_no']]
phone1.drop_duplicates(subset='phone_no',inplace=True)
phone1.shape
phone1_apply = pd.merge(phone1,appr_st,on='phone_no')
phone1_apply.phone_no.nunique()
phone1_apply.shape


white2 = pd.read_clipboard(dtype={'phone_no':str})


cd = pd.merge(invite1,invite2,on='phone_no')
cd.shape
cd_appr = pd.merge(cd,appr_st,on='phone_no')
cd_appr.phone_no.nunique()
cd_appr.shape
cd_appr = pd.merge(cd,appr_s1,on='phone_no')
cd_appr.phone_no.nunique()
phone1 = cd_appr.loc[:,['phone_no']]
phone1.drop_duplicates(subset='phone_no',inplace=True)
pd.merge(phone1,appr_st,on='phone_no').shape

invite = pd.concat([invite1,invite2])
invite.drop_duplicates(subset='phone_no',inplace=True)

cd = pd.merge(invite,white1,on='phone_no')
cd = pd.merge(invite,white2,on='phone_no')

invite2['listCategory'] = '免息邀请名单'
invite2.drop(columns='白名单日期',inplace=True)
white1['listCategory'] = '白名单1'
white2['listCategory'] = '白名单2'
listall = pd.concat([invite1,invite2,white1,white2])
listall.shape
listall.groupby('listCategory').size()
repay_mart = pd.read_excel(r"repay_mart.xlsx",dtype={'phone_no':str})
listall_repay = pd.merge(listall,repay_mart,on='phone_no')
listall_repay.to_excel(r"C:\Users\tsjr\Desktop\listall_repay.xlsx",index=False)
listall_nodup = listall.drop_duplicates(subset='phone_no')
listall_nodup_repay = pd.merge(listall_nodup,repay_mart,on='phone_no')
listall_nodup_repay.to_excel(r"C:\Users\tsjr\Desktop\listall_nodup_repay.xlsx",index=False)

credit_limit_log = pd.read_sql("select user_code, contract_amount_max, \
                                    duration_longest, service_fee_rate, \
                                    created_time as limit_time, \
                                    date_format(created_time,'%Y%m%d') as limit_date \
                                from suncash_lend.credit_limit_log \
                                where created_time>='2020-06-01'",cnx)
credit_limit_log.sort_values(by=['user_code','limit_date','limit_time'],inplace=True)
limit_log = credit_limit_log.drop_duplicates(subset=['user_code','limit_date'],keep='last')
user_phone = user.loc[:,['phone_no','user_code']]
limit_log = pd.merge(limit_log,user_phone,on='user_code')


#%%--每日激活名单
white1_invite = pd.read_clipboard(dtype={'phone_no':str}) #待激活白名单
invite = pd.read_clipboard(dtype={'phone_no':str}) #待激活复贷客户

appr_st = appr_mart.query("apply_time>='2020-05-12'")
appr_1st = appr_st.sort_values(by=['phone_no','apply_time'])
appr_1st.drop_duplicates(subset='phone_no',inplace=True)
white1_invite_appr = pd.merge(white1_invite,appr_1st.loc[appr_1st.apply_day=='20200805',:],on='phone_no')
white1_invite_appr.loc[:,['phone_no']].to_clipboard(index=False)

appr_july = appr_mart.query("apply_time>='2020-07-01'")
appr_july_1st = appr_july.sort_values(by=['phone_no','apply_time'])
appr_july_1st.drop_duplicates(subset='phone_no',inplace=True)
invite_appr = pd.merge(invite,appr_july_1st.loc[appr_july_1st.apply_day=='20200805',:],on='phone_no')
invite_appr.loc[:,['phone_no']].to_clipboard(index=False)

white3 = pd.read_clipboard(dtype={'phone_no':str}) #第三批白名单
#recent_apply = appr_mart.query("apply_time>='2020-07-01' and apply_time<='2020-07-28'").loc[:,['phone_no']]
#recent_apply.drop_duplicates(subset='phone_no',inplace=True)
#recent_apply['近期已激活'] = 1
#white3 = pd.merge(white3,recent_apply,how='left',on='phone_no')
white3_appr = pd.merge(white3,appr_july_1st.loc[appr_july_1st.apply_day=='20200805',:],on='phone_no')
white3_appr.loc[:,['phone_no']].to_clipboard(index=False)

white4 = pd.read_clipboard(dtype={'phone_no':str}) #第三批白名单
white4_appr = pd.merge(white4,appr_july_1st.loc[appr_july_1st.apply_day=='20200805',:],on='phone_no')
white4_appr.loc[:,['phone_no']].to_clipboard(index=False)


repay_last = repay_mart.loc[repay_mart.last_loan=='Y',['user_code','od_days','account_status','clear_date']]
def curr_state(x):
    if x.od_days>10:
        return '逾期10天以上'
    elif x.account_status=='ACCOUNT_OVERDUE':
        return '逾期10天以内'
    elif (datetime.date.today()-x.clear_date).days>10:
        return '结清10天以上'
    elif x.account_status=='ACCOUNT_SETTLE':
        return '结清10天以内'
    else:
        return '未到期'

repay_last['当前状态'] = repay_last.apply(curr_state,axis=1)
repay_last['当前状态'].value_counts()
repay_last['当前状态'].value_counts().plot(kind='pie')
repay_last.drop(columns=['od_days','account_status','clear_date'],inplace=True)

#根据每个客户累计放出与回收金额来定义好坏
revenue_expenditure = repay_mart.loc[repay_mart.loan_month!='202003',['user_code','contract_amount','service_fee_rate',
                               'number_of_loans','od_days_ever','actualRepayAmt']]
revenue_expenditure['acutalLoanAmt'] = revenue_expenditure.apply(lambda x: x.contract_amount*(1-x.service_fee_rate),axis=1)
revenue_expenditure.actualRepayAmt.fillna(0,inplace=True)
rev_exp = revenue_expenditure.groupby('user_code',as_index=False).agg({'acutalLoanAmt':sum,'actualRepayAmt':sum,'number_of_loans':max,'od_days_ever':max})
rev_exp['break_even'] = rev_exp['actualRepayAmt'] - rev_exp['acutalLoanAmt']

def gb_define(x):
    if x.break_even>=1000:
        return 'good'
    elif x.od_days_ever>30 and x.break_even<1000:
        return 'bad'
    else:
        return 'indet'

rev_exp['gb'] = rev_exp.apply(gb_define,axis=1)

rev_exp['gb'].value_counts()

feature_names = ['gender','age','marriage','education','children_text','id_type',
                 'job_type','on_the_job_time','monthly_net_income_text',
                 'loan_channel','user_code','first_service_fee_rate','first_loan_month']
user_feature = repay_mart.loc[repay_mart.number_of_loans==1,feature_names]
rev_exp = pd.merge(rev_exp,user_feature,on='user_code')
rev_exp = pd.merge(rev_exp,repay_last,on='user_code')
rev_exp.to_excel(r"C:\Users\tsjr\Desktop\rev_exp.xlsx",index=False)
modeldata = rev_exp.loc[rev_exp.gb!='indet',:]
y = modeldata.gb.apply(lambda x: 1 if x=='bad' else 0)
X = modeldata.loc[:,feature_names].drop(columns='user_code')


#%%--业务恢复有来申请的客户
active_user = repay_mart.query("loan_date>='2020-05-12'").loc[:,['user_code']]
active_user.drop_duplicates(subset='user_code',inplace=True)
rev_exp = pd.merge(rev_exp,active_user,on='user_code')
rev_exp['gb'].value_counts()


##
#%%-- facebook and email account
facebook_account = pd.read_sql("select user_code, facebook_account, email \
                        from suncash_lend.personal_info",cnx)
user_phone = user.loc[:,['user_code','phone_no']]
facebook = pd.merge(facebook_account,user_phone,on='user_code')
facebook.drop_duplicates(subset='user_code',inplace=True)
#最后一次申请自动拒绝客户
apply_last = appr_mart.sort_values(by=['user_code','number_of_apply'])
apply_last.drop_duplicates(subset='user_code',keep='last',inplace=True)
auto_refuse = apply_last.loc[apply_last.自动拒绝==1,['user_code','自动拒绝']]
#当前未结清客户
last_loan = repay_mart.loc[repay_mart.last_loan=='Y',['user_code','last_loan','account_status']]
unsettle = last_loan.loc[last_loan.account_status!='ACCOUNT_SETTLE',:]
unsettle['未结清'] = 1
unsettle.drop(columns=['last_loan','account_status'],inplace=True)

facebook = pd.merge(facebook,auto_refuse,how='left',on='user_code')
facebook = pd.merge(facebook,unsettle,how='left',on='user_code')
facebook_active = facebook.query("自动拒绝!=1 and 未结清!=1")
facebook_active.drop(columns=['user_code','自动拒绝','未结清'],inplace=True)
facebook_active.to_csv(r"C:\Users\tsjr\Desktop\facebook营销名单.csv",index=False)

#facebook激活名单
facebook_invite = pd.read_csv(r"C:\Users\tsjr\Desktop\facebook营销名单.csv",dtype={'phone_no':str}) #待激名单
appr_facebook = appr_mart.query("apply_time>='2020-07-17'")
appr_facebook_1st = appr_facebook.sort_values(by=['phone_no','apply_time'])
appr_facebook_1st.drop_duplicates(subset='phone_no',inplace=True)
invite_appr = pd.merge(invite,appr_facebook_1st.loc[appr_facebook_1st.apply_day=='20200706',:],on='phone_no')
#%%---Individual Collection(U志培版)
import saspy
import pandas as pd
import datetime
import numpy as np

sas=saspy.SASsession()
sas.saslib('MY "F:\TS\\Nwork\\rawdata"')
# 数据量小，直接用etl-out的
# case_flow_info=sas.sd2df('case_flow_info','MY')
# my_case_info=sas.sd2df('my_case_info','MY')
# account_info=sas.sd2df('account_info','MY')
# trans_journal_detail=sas.sd2df('trans_journal_detail','MY')
# realtime_draw=sas.sd2df('realtime_draw','MY')
# extension_info=sas.sd2df('extension_info','MY')

case_flow_info.columns = map(str.lower, case_flow_info.columns)
my_case_info.columns = map(str.lower, my_case_info.columns)
# a2=case_flow_all.loc[case_flow_all.contract_no=='C201906211949040110036',:]
case_flow_cols = ['created_time','case_no','contract_no','user_id']
case_flow = case_flow_info.loc[:,case_flow_cols].copy()
my_case = my_case_info.loc[:,case_flow_cols]
case_flow_all = pd.concat([my_case,case_flow])
case_flow_all['流转日期'] = pd.to_datetime(case_flow_all['created_time']).dt.date
# case_flow_st= case_flow_all.loc[case_flow_all.流转日期>=datetime.date(2020,3,1),:]
case_flow_st= case_flow_all.copy()
case_info = account_info.loc[:,['contract_no','account_status','loan_date','clear_date','extend_date','last_repay_date','user_code','contract_amount']]

case_info.sort_values(by=['user_code','loan_date'],inplace=True)
case_info['lag_contract_no'] = case_info['contract_no'].shift(1)
case_flow_st = pd.merge(case_flow_st,case_info,how='left',on='contract_no')
case_flow_st['clear_date']=pd.to_datetime(case_flow_st['clear_date']).dt.date
case_flow_st['user_id']=case_flow_st['user_id'].fillna('NOONE')
case_flow_st['user_id']=case_flow_st['user_id'].apply(lambda x:x.lower())
#查询个人
#很严重的错误，不能在源头进行筛选，应该是总体跑出来后再筛选;
# importname=['QSSI-1','QSSI-2','QSSI-3','QSSI-4','QSSI-5','QSSI-6','QSSI-7','QSSI-8','QSSI-9','QSSI-10','QSSI-11','QSSI-12']
# case_flow_st=case_flow_st.loc[((case_flow_st.user_id=='qssi-1')  | 
#                                (case_flow_st.user_id=='qssi-2')  |
#                                (case_flow_st.user_id=='qssi-3')  |
#                                (case_flow_st.user_id=='qssi-4')  |
#                                (case_flow_st.user_id=='qssi-5')  |
#                                (case_flow_st.user_id=='qssi-6')  |
#                                (case_flow_st.user_id=='qssi-7')  |
#                                (case_flow_st.user_id=='qssi-8')  |
#                                (case_flow_st.user_id=='qssi-9')  |
#                                (case_flow_st.user_id=='qssi-10')  |
#                                (case_flow_st.user_id=='qssi-11')  |
#                                (case_flow_st.user_id=='qssi-12')  
#                                ),:]

assignment = pd.DataFrame()

for (i,cut_date) in enumerate(pd.date_range(datetime.date(2022,9,1),
                                            datetime.date(2022,10,1),
                                            closed='left')):
    assign = case_flow_st.loc[case_flow_st.流转日期<=cut_date,:].copy()
    assign.sort_values(['contract_no','created_time'],inplace=True)
    assign.drop_duplicates('contract_no',keep='last',inplace=True)
    # assign['cut_date'] = cut_date.date()
    assign['cut_date'] = cut_date
    assign['od_days']=assign.apply(lambda x: (pd.to_datetime(x.cut_date)-pd.to_datetime(x.last_repay_date)).days if pd.isnull(x.clear_date) or x.clear_date>=x.cut_date else 0,axis=1)
    assign['逾期天数']=assign.apply(lambda x: (pd.to_datetime(x.clear_date)-pd.to_datetime(x.last_repay_date)).days if pd.isnull(x.clear_date)  else 0,axis=1)
    if i==0:
        assignment = assign
    else:
        assignment = pd.concat([assignment,assign])
    # sas.df2sd(assignment.astype(str),'assignment2111',"MY",encode_errors='replace')
    
pay=repay_detail.rename(columns={'settle_date':'cut_date'}).groupby(['contract_no','cut_date']).agg({'trans_amount':np.sum}).reset_index()
pay['cut_date'] = pd.to_datetime(pay['cut_date'])
assignment_t = pd.merge(assignment,pay,how='left',on=['contract_no','cut_date'])
out=assignment_t[assignment_t.trans_amount.notna()  ]
os.chdir(r"F:\TS\Nwork")

out.to_excel(r"U0901_0930回收金额明细.xlsx",index=False)

sas.df2sd(assignment.astype(str),'assignment2109',"MY",encode_errors='replace')


# 还款日期=cut-date



target=repay_mart.loc[repay_mart.last_repay_date>=datetime.date(2021,11,1),:]
target['clear_date']=target.apply(lambda x:x.extend_date if x.account_status=='ACCOUNT_CLOSED' else x.clear_date,axis=1)
assignment_t=assignment_t.loc[:,['contract_no','cut_date','user_id','trans_amount']]
assignment_t['cut_date']=pd.to_datetime(assignment_t['cut_date']).dt.date
target=target[target.clear_date.notna()]
target['clear_date']=pd.to_datetime(target['clear_date']).dt.date
target=target.loc[:,['contract_no','clear_date','last_repay_date','account_status','od_days_ever']]
target=pd.merge(target,assignment_t,how='left',left_on=['contract_no','clear_date'],right_on=['contract_no','cut_date'])

trans_journal_detail.columns = trans_journal_detail.columns.str.lower()
repay_detail_1 = trans_journal_detail.loc[:,['contract_no','trans_amount','settle_date']]
repay_detail_2 = realtime_draw.query("BUSI_TYPE=='PUBLIC_TRANSFER'").loc[:,['CONTRACT_NO','REAL_AMOUNT','CREATED_TIME']].rename(columns={'CONTRACT_NO':'contract_no','REAL_AMOUNT':'trans_amount','CREATED_TIME':'settle_date'})
repay_detail = pd.concat([repay_detail_1,repay_detail_2])
paydate=repay_detail.groupby('contract_no').agg({'settle_date':'max'}).reset_index()

target=pd.merge(target,paydate,how='left',on='contract_no')
target.to_excel(r'近一个月还款名单.xlsx',index=False)



assignment.sort_values(by=['contract_no','cut_date'],inplace=True)
assignment['lag_user_id'] = assignment['user_id'].shift(1)
assignment['lag_contract_no'] = assignment['contract_no'].shift(1)
assignment['当天流入'] = assignment.apply(lambda x: 1 if x.contract_no!=x.lag_contract_no or x.user_id!=x.lag_user_id else 0,axis=1)
assignment['当天流入合同金额'] = assignment.apply(lambda x: x.contract_amount if x.contract_no!=x.lag_contract_no or x.user_id!=x.lag_user_id else 0,axis=1)

assignment['7天组分母']= assignment.apply(lambda x: x.contract_amount if x.当天流入==1 and  x.cut_date>=datetime.date(2021,9,25) and x.cut_date<=datetime.date(2021,11,4)  else 0,axis=1)
assignment['8_15天组分母']= assignment.apply(lambda x: x.contract_amount if x.当天流入==1 and  x.cut_date>=datetime.date(2021,9,24) and x.cut_date<=datetime.date(2021,11,3)  else 0,axis=1)
outfm=assignment[assignment.当天流入==1 ]
outfm.to_excel(r"U1001_1110分母.xlsx",index=False)

#临时需求
pay16=repay_mart.loc[(repay_mart.last_repay_date>=datetime.date(2021,11,29)) & (repay_mart.last_repay_date<=datetime.date(2021,12,5)),
                     ['contract_no','account_status','clear_date','contract_amount','extend_date','last_repay_date','od_days_ever']]
assignment_t['cut_date']=pd.to_datetime(assignment_t['cut_date']).dt.date
ass16=assignment_t.loc[(assignment_t.cut_date>=datetime.date(2021,11,29)) & (assignment_t.cut_date<=datetime.date(2021,12,5)) ,['user_id','cut_date','contract_no']]
pay16m=pd.merge(pay16,ass16,how='left',left_on=['contract_no','last_repay_date'],right_on=['contract_no','cut_date'])


pay16m.to_excel(r"F:\TS\Nwork\temp.xlsx",index=False)



#%%---Individual Collection(志培版)
#----------------------------------------------------------------------------------------------------------------------------------------------------------
import saspy
import pandas as pd
import datetime
import numpy as np
sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
case_flow_info=sas.sd2df('case_flow_info','MY')
my_case_info=sas.sd2df('my_case_info','MY')
account_info=sas.sd2df('account_info','MY')
trans_journal_detail=sas.sd2df('trans_journal_detail','MY')
realtime_draw=sas.sd2df('realtime_draw','MY')
extension_info=sas.sd2df('extension_info','MY')

# a2=case_flow_all.loc[case_flow_all.contract_no=='C201906211949040110036',:]
case_flow_cols = ['created_time','case_no','contract_no','user_id']
case_flow = case_flow_info.loc[:,case_flow_cols].copy()
my_case = my_case_info.loc[:,case_flow_cols]
case_flow_all = pd.concat([my_case,case_flow])
case_flow_all['流转日期'] = pd.to_datetime(case_flow_all['created_time']).dt.date
# case_flow_st= case_flow_all.loc[case_flow_all.流转日期>=datetime.date(2020,3,1),:]
case_flow_st= case_flow_all.copy()
case_info = account_info.loc[:,['contract_no','account_status','loan_date','clear_date','extend_date','last_repay_date','user_code','contract_amount']]
case_info['clear_date']=account_info.apply(lambda x: x.extend_date if x.account_status=='ACCOUNT_CLOSED' else x.clear_date,axis=1)

case_info.sort_values(by=['user_code','loan_date'],inplace=True)

case_info['lag_contract_no'] = case_info['contract_no'].shift(1)
case_flow_st = pd.merge(case_flow_st,case_info,how='left',on='contract_no')
case_flow_st['clear_date']=pd.to_datetime(case_flow_st['clear_date']).dt.date
case_flow_st['user_id']=case_flow_st['user_id'].fillna('NOONE')
case_flow_st['user_id']=case_flow_st['user_id'].apply(lambda x:x.lower())
#查询个人
#很严重的错误，不能在源头进行筛选，应该是总体跑出来后再筛选;
# importname=['QSSI-1','QSSI-2','QSSI-3','QSSI-4','QSSI-5','QSSI-6','QSSI-7','QSSI-8','QSSI-9','QSSI-10','QSSI-11','QSSI-12']
# case_flow_st=case_flow_st.loc[((case_flow_st.user_id=='qssi-1')  | 
#                                (case_flow_st.user_id=='qssi-2')  |
#                                (case_flow_st.user_id=='qssi-3')  |
#                                (case_flow_st.user_id=='qssi-4')  |
#                                (case_flow_st.user_id=='qssi-5')  |
#                                (case_flow_st.user_id=='qssi-6')  |
#                                (case_flow_st.user_id=='qssi-7')  |
#                                (case_flow_st.user_id=='qssi-8')  |
#                                (case_flow_st.user_id=='qssi-9')  |
#                                (case_flow_st.user_id=='qssi-10')  |
#                                (case_flow_st.user_id=='qssi-11')  |
#                                (case_flow_st.user_id=='qssi-12')  
#                                ),:]

assignment = pd.DataFrame()

for (i,cut_date) in enumerate(pd.date_range(datetime.date(2022,11,1),
                                            datetime.date(2022,12,1),
                                            closed='left')):
    assign = case_flow_st.loc[case_flow_st.流转日期<=cut_date,['contract_no','created_time','last_repay_date','clear_date','user_id']].copy()
    assign.sort_values(['contract_no','created_time'],inplace=True)
    assign.drop_duplicates('contract_no',keep='last',inplace=True)
    # assign['cut_date'] = cut_date.date()
    assign['cut_date'] = cut_date
    assign['od_days']=assign.apply(lambda x: (pd.to_datetime(x.cut_date)-pd.to_datetime(x.last_repay_date)).days if pd.isnull(x.clear_date) or x.clear_date>=x.cut_date else 0,axis=1)

    if i==0:
        assignment = assign
    else:
        assignment = pd.concat([assignment,assign])
    sas.df2sd(assignment.astype(str),'assignment2211',"MY",encode_errors='replace')



assignment.sort_values(by=['contract_no','cut_date'],inplace=True)
assignment['lag_user_id'] = assignment['user_id'].shift(1)
assignment['lag_contract_no'] = assignment['contract_no'].shift(1)
assignment['当天流入'] = assignment.apply(lambda x: 1 if x.contract_no!=x.lag_contract_no or x.user_id!=x.lag_user_id else 0,axis=1)
assignment['当天流入合同金额'] = assignment.apply(lambda x: x.contract_amount if x.contract_no!=x.lag_contract_no or x.user_id!=x.lag_user_id else 0,axis=1)
assignment['当天逾期'] = assignment.apply(lambda x: 1 if pd.isnull(x.clear_date) and pd.to_datetime(x.last_repay_date)==x.cut_date else 0,axis=1)
assignment['当天队列'] = assignment.apply(lambda x: 1 if x.contract_no==x.lag_contract_no and x.user_id==x.lag_user_id and (pd.isnull(x.clear_date) or x.clear_date>=x.cut_date) else 0,axis=1)
assignment['当天队列合同金额'] = assignment.apply(lambda x: x.contract_amount if x.contract_no==x.lag_contract_no and x.user_id==x.lag_user_id and (pd.isnull(x.clear_date) or x.clear_date>=x.cut_date) else 0,axis=1)
assignment['当天催回'] = assignment.apply(lambda x: 1 if x.clear_date==x.cut_date else 0,axis=1)
assignment['当天展期'] = assignment.apply(lambda x: 1 if x.account_status=='ACCOUNT_CLOSED' and x.extend_date==x.cut_date else 0,axis=1)

repay_detail_1 = trans_journal_detail.loc[:,['contract_no','trans_amount','settle_date']]
repay_detail_2 = realtime_draw.query("busi_type=='PUBLIC_TRANSFER'").loc[:,['contract_no','real_amount','created_time']].rename(columns={'real_amount':'trans_amount','created_time':'settle_date'})
repay_detail = pd.concat([repay_detail_1,repay_detail_2])
repay_detail['settle_date'] = pd.to_datetime(repay_detail['settle_date']).dt.date
last_repay_date = account_info.loc[:,['contract_no','last_repay_date']]
repay_detail = pd.merge(repay_detail,last_repay_date,on='contract_no',how='left')
repay_detail['last_repay_date']=pd.to_datetime(repay_detail['last_repay_date']).dt.date
repay_detail['到期前还款'] = repay_detail.apply(lambda x: 1 if x.settle_date<=x.last_repay_date else 0,axis=1)
repay_detail['逾期后还款'] = repay_detail.apply(lambda x: 1 if x.settle_date>x.last_repay_date else 0,axis=1)
repay_detail['hk_date'] = pd.to_datetime(repay_detail['settle_date']).dt.date

od_repay = repay_detail.query("逾期后还款==1").groupby(['contract_no','hk_date'],as_index=False)['trans_amount'].agg({'逾期回款金额':sum}).rename(columns={'hk_date':'yq_hk_date'})
od_repay['yq_hk_date']=pd.to_datetime(od_repay['yq_hk_date'])
pron_repay = repay_detail.query("到期前还款==1").groupby(['contract_no','hk_date'],as_index=False)['trans_amount'].agg({'到期前含回款金额':sum}).rename(columns={'hk_date':'pron_hk_date'})
pron_repay['pron_hk_date']=pd.to_datetime(pron_repay['pron_hk_date'])
sas.df2sd(od_repay.astype(str),'od_repay',"MY",encode_errors='replace')
sas.df2sd(pron_repay.astype(str),'pron_repay',"MY",encode_errors='replace')

hm_repay=realtime_draw.query("busi_type in ('CLEAR_INTEREST','OVERPAYMENT')").loc[:,['contract_no','real_amount','created_time']]
hm_repay['created_time']=pd.to_datetime(hm_repay['created_time']).dt.date
hm_repay = hm_repay.groupby(['contract_no','created_time'],as_index=False)['real_amount'].agg({'豁免金额':sum}).rename(columns={'created_time':'hm_created_time'})
hm_repay['hm_created_time']=pd.to_datetime(hm_repay['hm_created_time'])
extension_info=extension_info.loc[:,['overdue_penalty_amt','overdue_service_fee_amt','created_time','act_contract_no']]
extension_info['created_time']=pd.to_datetime(extension_info['created_time']).dt.date
zq_amount = extension_info.groupby(['act_contract_no','created_time'],as_index=False).agg({'overdue_penalty_amt':'sum','overdue_service_fee_amt':'sum'})
hk_amt = repay_detail.groupby(['contract_no','hk_date'])['trans_amount'].sum().reset_index()
hk_amt['hk_date'] = pd.to_datetime(hk_amt['hk_date'])
sas.df2sd(hk_amt.astype(str),'hk_amt',"MY",encode_errors='replace')
# assignment_aa=assignment.copy()
assignment = pd.merge(assignment,hk_amt,how='left',left_on=['contract_no','cut_date'],right_on=['contract_no','hk_date'])
case_infoaa = pd.merge(case_info.loc[:,['contract_no','lag_contract_no']],zq_amount,how='left',left_on=['contract_no'],right_on=['act_contract_no']).rename(columns={'lag_contract_no':'zq_contract_no','created_time':'zq_created_time'})
case_infoaa['zq_created_time']=pd.to_datetime(case_infoaa['zq_created_time'])
assignment = pd.merge(assignment,case_infoaa.loc[:,['zq_contract_no','zq_created_time','overdue_penalty_amt','overdue_service_fee_amt']],how='left',left_on=['contract_no','cut_date'],right_on=['zq_contract_no','zq_created_time'])

assignment = pd.merge(assignment,hm_repay,how='left',left_on=['contract_no','cut_date'],right_on=['contract_no','hm_created_time'])

assignment = pd.merge(assignment,od_repay,how='left',left_on=['contract_no','cut_date'],right_on=['contract_no','yq_hk_date'])
assignment = pd.merge(assignment,pron_repay,how='left',left_on=['contract_no','cut_date'],right_on=['contract_no','pron_hk_date'])

assignment['当天催回金额'] = assignment.apply(lambda x: x.trans_amount if x.当天催回==1 else 0,axis=1)
assignment['当天展期金额'] = assignment.apply(lambda x: x.trans_amount if x.当天展期==1 else 0,axis=1)
assignment['当天回款金额'] = assignment['trans_amount']
assignment['cut_date'] = pd.to_datetime(assignment['cut_date']).dt.date
assignment['od_days_interval']=pd.cut(assignment['od_days'],bins=[-np.inf,-1,0,7,15,np.inf],labels=['less then 0','0','1-7','8-15','15+'])
###test

a=assignment.loc[(assignment.user_id=='cris') & (assignment.cut_date==datetime.date(2021,3,1)) & (assignment.当天队列==1),:]
a['od_days_interval'].value_counts()
a1=a.loc[a.od_days==17,]
a2=case_flow_all.loc[case_flow_all.contract_no=='C201906211949040110036',:]
a3=assignment.loc[assignment.contract_no=='C202101291351386740028',['当天队列','当天队列合同金额','当天流入','当天流入合同金额','当天催回','到期前(含)回款金额','逾期回款金额','cut_date','clear_date','last_repay_date','流转日期','user_id']]
a=assignment.loc[(assignment.user_id=='cris') & (assignment.cut_date<=datetime.date(2021,3,31)) & (assignment.cut_date>=datetime.date(2021,3,1)) & (assignment.当天催回==1),:]
a['od_days_interval'].value_counts()
a1=a.loc[a.od_days==2,]
a['account_status'].value_counts()

#最近6个月；
import calendar
mtnall=pd.DataFrame()
# mtnalla=pd.DataFrame()
for i in range(1,7):
    d = calendar.monthrange(2021,i)
    month_begin_queue = assignment.loc[assignment.cut_date==datetime.date(2021,i,1),:].groupby(['user_id','od_days_interval'])['当天队列','当天队列合同金额'].sum()
    mtd_queue = assignment.loc[(assignment.cut_date<=datetime.date(2021,i,d[1])) &
                           (assignment.cut_date>=datetime.date(2021,i,1)),:].groupby(['user_id','od_days_interval'])['当天流入','当天流入合同金额','当天催回','到期前(含)回款金额','逾期回款金额'].sum()
    mtd_queue.insert(0,'月初队列',month_begin_queue)
    month_begin_queue['cut_date']=datetime.date(2021,i,d[1])
    mtd_queue['cut_date']=datetime.date(2021,i,d[1])
    if i==0:
        mtnall=mtd_queue
        # mtnalla=month_begin_queue
    else:
        mtnall=pd.concat([mtnall,mtd_queue])
        # mtnalla=pd.concat([mtnalla,month_begin_queue])
mtnall.to_excel(r"F:\TS\PreWork\act_task\近6个月催回率.xlsx",sheet_name='月内流入')
# mtnalla.to_excel(r"F:\TS\PreWork\act_task\近6个月催回率a.xlsx",sheet_name='月初队列')

close_date = datetime.date.today() - datetime.timedelta(days=11)
individual_col = assignment.loc[assignment.cut_date==close_date,:].groupby('user_id')['当天队列','当天催回','当天展期','当天催回金额','当天展期金额','当天回款金额'].sum()

print()
month_begin_queue = assignment.loc[assignment.cut_date==datetime.date(2021,5,1),:].groupby('user_id')['当天队列'].sum()

mtd_queue = assignment.loc[(assignment.cut_date.dt.date<=close_date) &
                           (assignment.cut_date.dt.date>=datetime.date(2021,5,1)),:].groupby('user_id')['当天流入','当天催回',
                                                                                                '当天展期','当天催回金额','当天展期金额','当天回款金额',
                                                                                               '逾期回款金额','豁免金额','overdue_penalty_amt','overdue_service_fee_amt'].sum()
Billy = assignment.loc[(pd.to_datetime(assignment.cut_date).dt.date>=datetime.date(2021,5,1)),:].groupby('user_id').agg({'当天流入':'sum','当天逾期':'sum'}).rename(columns={'当天流入':'累计流入','当天逾期':'累计逾期'})

Billy['流入率'] = (Billy['累计逾期']/Billy['累计流入']).apply(lambda x: format(x,'.2%'))
individual_col.insert(3,'月初队列',month_begin_queue)
individual_col.insert(4,'累计流入',mtd_queue['当天流入'])
individual_col.insert(5,'累计催回',mtd_queue['当天催回'])
individual_col.insert(6,'累计展期',mtd_queue['当天展期'])
individual_col.insert(7,'累计催回金额',mtd_queue['当天催回金额'])
individual_col.insert(8,'累计展期金额',mtd_queue['当天展期金额'])
individual_col.insert(9,'累计回款金额',mtd_queue['当天回款金额'])
individual_col.insert(10,'逾期回款金额',mtd_queue['逾期回款金额'])
individual_col.insert(11,'累计豁免金额',mtd_queue['豁免金额'])
individual_col.insert(12,'累计overdue_penalty_amt',mtd_queue['overdue_penalty_amt'])
individual_col.insert(13,'累计overdue_service_fee_amt',mtd_queue['overdue_service_fee_amt'])

individual_col.to_excel(r'F:\TS\PreWork\act_task\11.xlsx')

assignment['ods_interval']=pd.cut(assignment.od_days,bins=[0,30,60,90,180,360,np.inf],labels=['0-30','31-60','61-90','91-180','180-360','360+'])
a=assignment.query("cut_date<=datetime.date(2021,6,30) and  cut_date>=datetime.date(2021,6,1) and user_id in ('QSSI-1','QSSI-2','QSSI-3','QSSI-4', 'QSSI-5','QSSI-6','QSSI-7','QSSI-8','QSSI-9','QSSI-10','QSSI-11','QSSI-12')").drop(columns=['created_time','case_no','user_code','loan_date','lag_user_id','lag_contract_no'])
a.to_csv(r'F:\TS\PreWork\act_task\cs.csv')
#%%--Individual Collection(登锋版)
case_flow_cols = ['created_time','case_no','contract_no','user_id']
case_flow = case_flow_info.loc[:,case_flow_cols]
my_case = my_case_info.loc[:,case_flow_cols]
case_flow_all = pd.concat([case_flow,my_case])
case_flow_all['流转日期'] = pd.to_datetime(case_flow_all['created_time']).dt.date
case_flow_st= case_flow_all.loc[case_flow_all.流转日期>=datetime.date(2020,5,10),:]
case_info = account_info.loc[:,['contract_no','account_status','loan_date','clear_date','extend_date','last_repay_date']]
case_flow_st = pd.merge(case_flow_st,case_info,how='left',on='contract_no')
assignment = pd.DataFrame()
for (i,cut_date) in enumerate(pd.date_range(datetime.date(2020,8,10),
                                            datetime.date.today(),
                                            closed='left')):
    assign = case_flow_st.loc[case_flow_st.流转日期<=cut_date,:].copy()
    assign.sort_values(['contract_no','created_time'],inplace=True)
    assign.drop_duplicates('contract_no',keep='last',inplace=True)
    assign['cut_date'] = cut_date.date()
#    assign['od_days_at_cutdate'] = assign.apply(od_days_at_cutdate,axis=1)
    if i==0:
        assignment = assign
    else:
        assignment = pd.concat([assignment,assign])

assignment.sort_values(by=['contract_no','cut_date'],inplace=True)
assignment['lag_user_id'] = assignment['user_id'].shift(1)
assignment['lag_contract_no'] = assignment['contract_no'].shift(1)
assignment['当天流入'] = assignment.apply(lambda x: 1 if x.contract_no!=x.lag_contract_no or x.user_id!=x.lag_user_id else 0,axis=1)
assignment['当天队列'] = assignment.apply(lambda x: 1 if x.contract_no==x.lag_contract_no and x.user_id==x.lag_user_id else 0,axis=1)
assignment['当天催回'] = assignment.apply(lambda x: 1 if x.clear_date==x.cut_date else 0,axis=1)
assignment['当天展期'] = assignment.apply(lambda x: 1 if x.account_status=='ACCOUNT_CLOSED' and x.extend_date==x.cut_date else 0,axis=1)

repay_detail_1 = trans_journal_detail.loc[:,['contract_no','trans_amount','settle_date']]
repay_detail_2 = realtime_draw.query("busi_type=='PUBLIC_TRANSFER'").loc[:,['contract_no','real_amount','created_time']].rename(columns={'real_amount':'trans_amount','created_time':'settle_date'})
repay_detail = pd.concat([repay_detail_1,repay_detail_2])
repay_detail['hk_date'] = pd.to_datetime(repay_detail['settle_date']).dt.date    
hk_amt = repay_detail.groupby(['contract_no','hk_date'])['trans_amount'].sum().reset_index()

assignment = pd.merge(assignment,hk_amt,how='left',left_on=['contract_no','cut_date'],right_on=['contract_no','hk_date'])
assignment['当天催回金额'] = assignment.apply(lambda x: x.trans_amount if x.当天催回==1 else 0,axis=1)
assignment['当天展期金额'] = assignment.apply(lambda x: x.trans_amount if x.当天展期==1 else 0,axis=1)
assignment['当天回款金额'] = assignment['trans_amount']

close_date = datetime.date.today() - datetime.timedelta(days=1)
individual_col = assignment.loc[assignment.cut_date==close_date,:].groupby('user_id')['当天队列','当天催回','当天展期','当天催回金额','当天展期金额','当天回款金额'].sum()
month_begin_queue = assignment.loc[assignment.cut_date==datetime.date(2020,9,1),:].groupby('user_id')['当天队列'].sum()
mtd_queue = assignment.loc[(assignment.cut_date<=close_date) &
                           (assignment.cut_date>=datetime.date(2020,9,1)),:].groupby('user_id')['当天流入','当天催回','当天展期','当天催回金额','当天展期金额','当天回款金额'].sum()
individual_col.insert(3,'月初队列',month_begin_queue)
individual_col.insert(4,'累计流入',mtd_queue['当天流入'])
individual_col.insert(5,'累计催回',mtd_queue['当天催回'])
individual_col.insert(6,'累计展期',mtd_queue['当天展期'])
individual_col.insert(7,'累计催回金额',mtd_queue['当天催回金额'])
individual_col.insert(8,'累计展期金额',mtd_queue['当天展期金额'])
individual_col.insert(9,'累计回款金额',mtd_queue['当天回款金额'])
#%%-------登锋app数量
os.chdir(r"F:\TS\PreWork\app")
app=pd.read_excel('.\Online Lending apps in the philippines.xlsx',sheet_name='cyt')
app.dropna(subset=['ONLINE LENDING APPLICATIONS'],inplace=True)
ArithmeticErrorapp['ONLINE LENDING APPLICATIONS']=app['ONLINE LENDING APPLICATIONS'].apply(lambda x: x.split(','))
app['counts']=app['ONLINE LENDING APPLICATIONS'].apply(lambda x: len(x))
app=app.set_index('COMPANY NAME')
s=pd.DataFrame({'app':np.concatenate(app['ONLINE LENDING APPLICATIONS'].values)},index=app.index.repeat(app.counts.values))
s.to_excel(r".\cyt_app.xlsx")



#%%----------jupyter 抽验
start_dt = today - datetime.timedelta(days=15)  #按天统计的只显示近半个月的数据
applMart = appr_st.loc[(appr_st.customer_source_sys=='SunCash') & (appr_st.loan_type=='first_apply'),:].copy()
applMart['apply_time'] = applMart['apply_time'].apply(lambda x: x).dt.normalize()
applMart.index = pd.DatetimeIndex(applMart.apply_time)
appl_day = applMart[applMart.index.isin(pd.date_range(start_dt.strftime('%Y-%m-%d'),today.strftime('%Y-%m-%d')))].groupby('apply_day').agg({'user_code':lambda x:x.nunique(),'apply_code':np.size,
                                                                              '取消':'sum','锁定期':'sum','处理':'sum','通过':'sum',
                                                                              '放款':'sum','自动拒绝':'sum','自动通过':'sum',
                                                                              '人工通过':'sum','人工处理':'sum'
                                                                            }).rename(columns={'user_code':'客户','apply_code':'申请'})

appl_day.to_csv(r'C:\Users\Administrator\Desktop\914\apply.csv')


#%%----------逾期率、催回率、展期率
zq_type = repay_st.loc[repay_st.展期==1,['contract_no','逾期后展期']].rename(columns={'contract_no':'source_contract_no','逾期后展期':'原合同逾期'})
repay_st = pd.merge(repay_st,zq_type,how='left',on='source_contract_no')
if datetime.date.today().day==1:
    cols = ['loan_month','od_days_ever','account_status','到期','自然逾期','催回','逾期后展期','loan_type',
            'actualRepayAmt','contract_amount','interest','service_fee_amount']

    for contract_type in ['source','normal_act','od_act','all']:
        if contract_type=='source':
            colMart = repay_st.loc[repay_st.展期后的合同!=1,cols].copy()
            print("初始合同")
        elif contract_type=='normal_act':
            colMart = repay_st.loc[(repay_st.展期后的合同==1) & (repay_st.原合同逾期==0),cols].copy()
            print("正常展期后的合同")
        elif contract_type=='od_act':
            colMart = repay_st.loc[(repay_st.展期后的合同==1) & (repay_st.原合同逾期==1),cols].copy()
            print("逾期展期后的合同")
        else:
            colMart = repay_st.copy()
            print("所有合同")

    #     colMart = colMart.query("loan_month!='202009'")

        for loan_type in ['total','first_apply','re_apply']:
            if loan_type=='total':
                col_mart = colMart.copy()
            else:
                col_mart = colMart.loc[colMart.loan_type==loan_type,:].copy()

            col_mart['正常展期'] = col_mart.apply(lambda x: 1 if x.od_days_ever==0 and x.account_status=='ACCOUNT_CLOSED' else 0, axis=1)
            col_mart['ever_od3p'] = col_mart.od_days_ever.apply(lambda x: 1 if x>3 else 0)
            col_mart['ever_od7p'] = col_mart.od_days_ever.apply(lambda x: 1 if x>7 else 0)
            col_mart['ever_od15p'] = col_mart.od_days_ever.apply(lambda x: 1 if x>15 else 0)
            col_mart['ever_od30p'] = col_mart.od_days_ever.apply(lambda x: 1 if x>30 else 0)
            col_mart['col_1t3'] = col_mart.apply(lambda x: 1 if 3>=x.od_days_ever>0 and x.account_status=='ACCOUNT_SETTLE' else 0, axis=1)
            col_mart['col_4t7'] = col_mart.apply(lambda x: 1 if 7>=x.od_days_ever>3 and x.account_status=='ACCOUNT_SETTLE' else 0, axis=1)
            col_mart['col_8t15'] = col_mart.apply(lambda x: 1 if 15>=x.od_days_ever>7 and x.account_status=='ACCOUNT_SETTLE' else 0, axis=1)
            col_mart['col_16t30'] = col_mart.apply(lambda x: 1 if 30>=x.od_days_ever>15 and x.account_status=='ACCOUNT_SETTLE' else 0, axis=1)
            col_mart['col_30p'] = col_mart.apply(lambda x: 1 if x.od_days_ever>30 and x.account_status=='ACCOUNT_SETTLE' else 0, axis=1)
            col_mart['ext_1t3'] = col_mart.apply(lambda x: 1 if 3>=x.od_days_ever>0 and x.account_status=='ACCOUNT_CLOSED' else 0, axis=1)
            col_mart['ext_4t7'] = col_mart.apply(lambda x: 1 if 7>=x.od_days_ever>3 and x.account_status=='ACCOUNT_CLOSED' else 0, axis=1)
            col_mart['ext_8t15'] = col_mart.apply(lambda x: 1 if 15>=x.od_days_ever>7 and x.account_status=='ACCOUNT_CLOSED' else 0, axis=1)
            col_mart['ext_16t30'] = col_mart.apply(lambda x: 1 if 30>=x.od_days_ever>15 and x.account_status=='ACCOUNT_CLOSED' else 0, axis=1)
            col_mart['ext_30p'] = col_mart.apply(lambda x: 1 if x.od_days_ever>30 and x.account_status=='ACCOUNT_CLOSED' else 0, axis=1)
            col_mart['部分还款'] = col_mart.apply(lambda x: 1 if 0<x.actualRepayAmt<x.contract_amount+x.interest else 0, axis=1)
            col_mart['类展期的部分还款'] = col_mart.apply(lambda x: 1 if x.service_fee_amount+x.interest<=x.actualRepayAmt<x.contract_amount+x.interest else 0, axis=1)

            col = col_mart.groupby("loan_month")['到期','自然逾期','正常展期','催回','逾期后展期','col_1t3','ext_1t3',
                                                 'ever_od3p','col_4t7','ext_4t7','ever_od7p','col_8t15','ext_8t15',
                                                 'ever_od15p','col_16t30','ext_16t30','ever_od30p','col_30p','ext_30p',
                                                 '部分还款','类展期的部分还款'
                                                ].sum()
            col_rate = pd.DataFrame()
            # col_rate['自然逾期率'] = (col['自然逾期']/col['到期']).apply(lambda x: format(x,'.2%'))
            # col_rate['1_3天催回率'] = (col['col_1t3']/col['自然逾期']).apply(lambda x: format(x,'.2%'))
            # col_rate['1_3天展期率'] = (col['ext_1t3']/col['自然逾期']).apply(lambda x: format(x,'.2%'))
            col_rate['自然逾期率'] = (col['自然逾期']/col['到期'])
            col_rate['正常展期率'] = (col['正常展期']/col['到期'])
            col_rate['逾期后催回率'] = col['催回']/col['自然逾期']
            col_rate['逾期后展期率'] = col['逾期后展期']/col['自然逾期']
            col_rate['1_3天催回率'] = (col['col_1t3']/col['自然逾期'])
            col_rate['1_3天展期率'] = (col['ext_1t3']/col['自然逾期'])
            col_rate['3天以上逾期率'] = col_rate['自然逾期率']*(1-col_rate['1_3天催回率']-col_rate['1_3天展期率'])
            col_rate['4_7天催回率'] = (col['col_4t7']/col['ever_od3p'])
            col_rate['4_7天展期率'] = (col['ext_4t7']/col['ever_od3p'])
            col_rate['7天以上逾期率'] = col_rate['3天以上逾期率']*(1-col_rate['4_7天催回率']-col_rate['4_7天展期率'])
            col_rate['8_15天催回率'] = (col['col_8t15']/col['ever_od7p'])
            col_rate['8_15天展期率'] = (col['ext_8t15']/col['ever_od7p'])
            col_rate['15天以上逾期率'] = col_rate['7天以上逾期率']*(1-col_rate['8_15天催回率']-col_rate['8_15天展期率'])
            col_rate['16_30天催回率'] = (col['col_16t30']/col['ever_od15p'])
            col_rate['16_30天展期率'] = (col['ext_16t30']/col['ever_od15p'])
            col_rate['30天以上逾期率'] = col_rate['15天以上逾期率']*(1-col_rate['16_30天催回率']-col_rate['16_30天展期率'])
            col_rate['30天后催回率'] = col['col_30p']/col['ever_od30p']
            col_rate['30天后展期率'] = col['ext_30p']/col['ever_od30p']
            col_rate['部分还款率'] = col['部分还款']/col['到期']
            col_rate['类展期的部分还款率'] = col['类展期的部分还款']/col['到期']
            print(loan_type)
            col_rate.applymap(lambda x: format(x,'.2%'))
            col
else:
    print("月初输出，其他时间不输出")
#%%----------临时分析下载数据
import openpyxl
import json
import numpy as np
import pandas as pd
import sys
sys.path.append(r"C:\Users\lenovo\Anaconda3\Lib\site-packages")
import pymysql
import math
import datetime
import gc
#gc.collect() 垃圾回收，返回处理这些循环引用一共释放掉的对象个数
import os
import oss2
import zipfile
import saspy

from itertools import islice
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = [u'SimHei']
plt.rcParams['axes.unicode_minus'] = False

os.chdir(r"F:\菲律宾日报")
with open(r"db_config.json") as db_config:
    cnx_args = json.load(db_config)

cnx = pymysql.connect(**cnx_args)  
credit_strategy_info=pd.read_sql("select * from approval.credit_strategy_info  ",cnx)
device_info_derived_variable=pd.read_sql("select * from approval.device_info_derived_variable  ",cnx)

sas=saspy.SASsession()

sas.saslib('MY "F:\菲律宾日报"')
sas.df2sd(credit_strategy_info.astype(str),'credit_strategy_info',"MY",encode_errors='replace')#appr_mart特殊，单独搞
sas.df2sd(device_info_derived_variable.astype(str),'device_info_derived_variable',"MY",encode_errors='replace')#appr_mart特殊，单独搞

#%%---------- 制作动图


from PIL import Image

im = Image.open("扎克伯格1.jpg")
images = []
images.append(Image.open('扎克伯格2.jpg'))
images.append(Image.open('扎克伯格3.jpg'))
im.save('gif.gif', save_all=True, append_images=images, loop=1, duration=1, comment=b"aaabb")



#%%----------AI下发名单
import json
import numpy as np
import pandas as pd
import sys
sys.path.append(r"C:\Users\lenovo\Anaconda3\Lib\site-packages")
import pymysql
#import math
import datetime
#import gc
#gc.collect() 垃圾回收，返回处理这些循环引用一共释放掉的对象个数
import os
import oss2
import zipfile
from itertools import islice
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = [u'SimHei']
plt.rcParams['axes.unicode_minus'] = False
import saspy

os.chdir(r"F:\菲律宾日报")
with open(r"db_config.json") as db_config:
    cnx_args = json.load(db_config)

cnx = pymysql.connect(**cnx_args)  


sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
repay_mart=sas.sd2df('repay_mart','MY')

repay_mart['od_days']=repay_mart.od_days.apply(lambda x: int(x))

target=repay_mart.loc[(repay_mart.od_days>=4)  & (repay_mart.od_days<=6),:]

collection_log_info=pd.read_sql("select COLLECTION_DATE,CREATED_TIME,CONTRACT_NO,COLLECTION_RESULT_CODE,RELATIONSHIP,USER_NAME,CONTACT_PHONE,remark from collection.collection_log_info where DATE(CREATED_TIME)>=DATE_SUB(curdate(),INTERVAL 5 DAY) ",cnx)



# ta=group_c.get_group('C202112041557279320024')

# list_no=collection_log_info.CONTRACT_NO.unique()



# 'Promise to pay back'  'promise to pay partial first'
group_c=collection_log_info.groupby('CONTRACT_NO')

target_list = []
for i in list(list_no):
    group_one=group_c.get_group(i)
    if ('Promise to pay back' not in group_one.COLLECTION_RESULT_CODE.values) & ('promise to pay partial first' not in group_one.COLLECTION_RESULT_CODE.values)  :
        target_list.append(i)
df_list=pd.DataFrame({'contract_no':target_list})
df_list['target']=1
target=target.loc[:,['contract_no','phone_no','last_repay_date','customer_source_sys','name','loan_type']]
target=pd.merge(target,df_list,how='left',on='contract_no')
target=target[target.target==1].drop(columns=['target'])
os.chdir(r"F:\菲律宾拨打\ai")
target.to_excel(r'名单1215_逾期4至6天.xlsx',index=False)
#%%----------催记质检名单
import json
import numpy as np
import pandas as pd
import sys
sys.path.append(r"C:\Users\lenovo\Anaconda3\Lib\site-packages")
import pymysql
#import math
import datetime
#import gc
#gc.collect() 垃圾回收，返回处理这些循环引用一共释放掉的对象个数
import os
import oss2
import zipfile
from itertools import islice
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = [u'SimHei']
plt.rcParams['axes.unicode_minus'] = False
import saspy

os.chdir(r"F:\菲律宾日报")
with open(r"db_config.json") as db_config:
    cnx_args = json.load(db_config)

cnx = pymysql.connect(**cnx_args)  


sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
repay_mart=sas.sd2df('repay_mart','MY')

repay_mart['od_days']=repay_mart.od_days.apply(lambda x: int(x))

target=repay_mart.loc[(repay_mart.od_days>=1)  & (repay_mart.od_days<=6),:]

collection_log_info=pd.read_sql("select COLLECTION_DATE,CREATED_TIME,CONTRACT_NO,COLLECTION_RESULT_CODE,RELATIONSHIP,USER_NAME,CONTACT_PHONE,remark from collection.collection_log_info where DATE(CREATED_TIME)>=DATE_SUB(curdate(),INTERVAL 6 DAY) ",cnx)



target=target.loc[:,['contract_no','phone_no','last_repay_date','customer_source_sys','name','loan_type']]
target=pd.merge(target,collection_log_info,how='left',left_on='contract_no',right_on='CONTRACT_NO')


target.to_excel(r'质检.xlsx',index=False)
#%%----------urule_决策树数据源


def amount_base(x):
    if x.totalNumberOfLoans>=1:
        return '内部共债'
    elif x.白名单1==1:
        return '白名单'
    elif x.apply_time<=datetime.date(2021,9,23):
        if x.customer_source_sys=='SunCash' or x.customer_source_sys=='FlashLoan':
            if x.gender=='Female' and x.age>48:
                return '优质客群1级a'
            elif  x.id_type=="PRC ID" or ((x.pay_type_name=="Bank transfer" or x.pay_type_name=="E-wallet") and (x.job_type!="Private Company Employee" and  x.job_type!="BPO Professionals") and x.gender=="Female"):
                if x.nowHour>5 and x.nowHour<23 and x.applyCount1M<1 and x.limitCount1M>0 and x.firstLimitIntvl>1:
                    return '优质客群1级b'
                else:
                    return '优质客群2级a'
            elif x.job_type=="Teacher" or ( x.age>41 and x.gender=="Female" ):
                if x.nowHour>5 and x.nowHour<23 and x.applyCount1M<1 and x.limitCount1M>0 and x.firstLimitIntvl>1:
                    return '优质客群2级b'
                else:
                    return '优质客群3级'
            elif x.pay_type_name=="Cash pickup" and x.age<42 and x.id_type!="PRC ID" and (x.education=="High school" or x.education=="Primary school"):
                return '劣质客群1级'
            elif x.job_type=="BPO Professionals" and  x.gender!="Female":
                return '劣质客群2级'
            elif x.pay_type_name=="Cash pickup" and x.age<42 and x.id_type!="PRC ID":
                return '劣质客群3级'
            elif x.nowHour>5 and x.nowHour<23 and x.applyCount1M<1 and x.limitCount1M>1 and x.firstLimitIntvl>3:
                return '优质客群4级'
            else :
                return '剩余客群'
#             return 'amount_S(x)'
#             amount_S(x)
             
        elif x.customer_source_sys=='SuncashPautang':
            if (x.gender!="Female" and ((x.education!="College Graduate" and x.education!="Master/PHD") or x.age>=45)) or (x.gender=="Female" and x.education!="College Graduate" and x.education!="Master/PHD" and x.age<30):
                return '劣质客群1级'
            else :
                return '剩余客群'
#             return 'amount_P(x)'
#             amount_P(x)
             
        else:
            return '其他包'
    else:
        if x.customer_source_sys=='SunCash' or x.customer_source_sys=='SuncashPautang' or x.customer_source_sys=='FlashLoan':
            if x.gender=='Female' and x.age>48:
                return '优质客群1级a'
            elif  x.id_type=="PRC ID" or ((x.pay_type_name=="Bank transfer" or x.pay_type_name=="E-wallet") and (x.job_type!="Private Company Employee" and  x.job_type!="BPO Professionals") and x.gender=="Female"):
                if x.nowHour>5 and x.nowHour<23 and x.applyCount1M<1 and x.limitCount1M>0 and x.firstLimitIntvl>1:
                    return '优质客群1级b'
                else:
                    return '优质客群2级a'
            elif x.job_type=="Teacher" or ( x.age>41 and x.gender=="Female" ):
                if x.nowHour>5 and x.nowHour<23 and x.applyCount1M<1 and x.limitCount1M>0 and x.firstLimitIntvl>1:
                    return '优质客群2级b'
                else:
                    return '优质客群3级'
            elif x.pay_type_name=="Cash pickup" and x.age<42 and x.id_type!="PRC ID" and (x.education=="High school" or x.education=="Primary school"):
                return '劣质客群1级'
            elif x.job_type=="BPO Professionals" and  x.gender!="Female":
                return '劣质客群2级'
            elif x.pay_type_name=="Cash pickup" and x.age<42 and x.id_type!="PRC ID":
                return '劣质客群3级'
            elif x.nowHour>5 and x.nowHour<23 and x.applyCount1M<1 and x.limitCount1M>1 and x.firstLimitIntvl>3:
                return '优质客群4级'
            else :
                return '剩余客群'
#             return 'amount_S(x)'
#             amount_S(x)
             
        else:
            return '其他包'
def amount_base_e(x):
    if x.totalNumberOfLoans>=1:
        return '内部共债'
    elif x.白名单1==1:
        return '白名单'
    elif x.apply_time<=datetime.date(2021,9,23):
        if x.customer_source_sys=='SunCash' or x.customer_source_sys=='FlashLoan':
            if x.gender=='Female' and x.age>48:
                return '优质客群1级'
            elif  x.id_type=="PRC ID" or ((x.pay_type_name=="Bank transfer" or x.pay_type_name=="E-wallet") and (x.job_type!="Private Company Employee" and  x.job_type!="BPO Professionals") and x.gender=="Female"):
                return '优质客群2级'
            elif x.job_type=="Teacher" or ( x.age>41 and x.gender=="Female" ):
                return '优质客群3级'
            elif x.pay_type_name=="Cash pickup" and x.age<42 and x.id_type!="PRC ID" and (x.education=="High school" or x.education=="Primary school"):
                return '劣质客群1级'
            elif x.job_type=="BPO Professionals" and  x.gender!="Female":
                return '劣质客群2级'
            elif x.pay_type_name=="Cash pickup" and x.age<42 and x.id_type!="PRC ID":
                return '劣质客群3级'
            elif x.nowHour>5 and x.nowHour<23 and x.applyCount1M<1 and x.limitCount1M>1 and x.firstLimitIntvl>3:
                return '优质客群4级'
            else :
                return '剩余客群'
#             return 'amount_S(x)'
#             amount_S(x)
             
        elif x.customer_source_sys=='SuncashPautang':
            if (x.gender!="Female" and ((x.education!="College Graduate" and x.education!="Master/PHD") or x.age>=45)) or (x.gender=="Female" and x.education!="College Graduate" and x.education!="Master/PHD" and x.age<30):
                return '劣质客群1级'
            else :
                return '剩余客群'
#             return 'amount_P(x)'
#             amount_P(x)
             
        else:
            return '其他包'
    else:
        if x.customer_source_sys=='SunCash' or x.customer_source_sys=='SuncashPautang' or x.customer_source_sys=='FlashLoan':
            if x.gender=='Female' and x.age>48:
                return '优质客群1级'
            elif  x.id_type=="PRC ID" or ((x.pay_type_name=="Bank transfer" or x.pay_type_name=="E-wallet") and (x.job_type!="Private Company Employee" and  x.job_type!="BPO Professionals") and x.gender=="Female"):

                return '优质客群2级'
            elif x.job_type=="Teacher" or ( x.age>41 and x.gender=="Female" ):

                return '优质客群3级'
            elif x.pay_type_name=="Cash pickup" and x.age<42 and x.id_type!="PRC ID" and (x.education=="High school" or x.education=="Primary school"):
                return '劣质客群1级'
            elif x.job_type=="BPO Professionals" and  x.gender!="Female":
                return '劣质客群2级'
            elif x.pay_type_name=="Cash pickup" and x.age<42 and x.id_type!="PRC ID":
                return '劣质客群3级'
            elif x.nowHour>5 and x.nowHour<23 and x.applyCount1M<1 and x.limitCount1M>1 and x.firstLimitIntvl>3:
                return '优质客群4级'
            else :
                return '剩余客群'
#             return 'amount_S(x)'
#             amount_S(x)
             
        else:
            return '其他包'
def uc_code(x):
    if 'endevent' in list(x.case_state):
        return '放款'
    elif 'refuseEnd' in list(x.case_state):
        return '拒绝'
    elif 'cancelEnd' in list(x.case_state):
        return '取消'
    else:
        return '其他'    
    
    
import json
import numpy as np
import pandas as pd
import sys
sys.path.append(r"C:\Users\lenovo\Anaconda3\Lib\site-packages")
import pymysql
#import math
import datetime
#import gc
#gc.collect() 垃圾回收，返回处理这些循环引用一共释放掉的对象个数
import os
import oss2
import zipfile
from itertools import islice
import matplotlib.pyplot as plt
plt.rcParams['font.sans-serif'] = [u'SimHei']
plt.rcParams['axes.unicode_minus'] = False
import saspy

os.chdir(r"F:\菲律宾日报")
with open(r"db_config.json") as db_config:
    cnx_args = json.load(db_config)

cnx = pymysql.connect(**cnx_args)  


sas=saspy.SASsession()
sas.saslib('MY "F:\菲律宾日报"')
repay_mart=sas.sd2df('repay_mart','MY')
appr_mart=sas.sd2df('appr_mart','MY')
data=appr_mart.loc[:,['pay_type_name','id_type','children_text','education','gender','marriage','length_of_residence',
                       'age','job_type','on_the_job_time','monthly_net_income_text','apply_code','apply_time']]
data['apply_time']=pd.to_datetime(data['apply_time']).dt.date

data_perform=repay_mart.loc[(repay_mart.loan_type=='first_apply') & (repay_mart.展期!='1'),['apply_code','自然逾期','到期','customer_source_sys','白名单1','曾经逾期7天以上']]

data_behavior=pd.read_csv('data.csv')
data_behavior.drop_duplicates(subset='applyCode',keep='last',inplace=True)

device_var=pd.read_csv('device.csv')
device_var.drop_duplicates(subset='applyCode',keep='last',inplace=True)

data_base=pd.merge(data_perform,data,how='left',on='apply_code')
data_base=pd.merge(data_base,data_behavior,how='left',left_on='apply_code',right_on='applyCode')
data_base=pd.merge(data_base,device_var,how='left',left_on='apply_code',right_on='applyCode')
data_base['age']=data_base.age.apply(lambda x:int(x))
    
        
data_base['amount_tr']=data_base.apply(amount_base_e,axis=1)
# 用jupyter的数据源
# data_base.rename(columns={'自然逾期':'f_od','到期':'ondue','白名单1':'wlist','曾经逾期7天以上':'toseven',
#                           '设备指纹读取时间':'d_cretime','手机品牌':'p_type','系统语言':'s_lang','总内存':'t_memory','剩余内存':'e_memory',
#                           '电池状态':'charged_type','设备名称':'d_name','屏幕分辨率':'c_rate'},inplace=True)
        
        


# sas.saslib('MY "H:\搬家\F\TS\PreWork\\rule\决策树"')
# sas.df2sd(data_base,'data_base',"MY")

test=appr_mart.loc[:,['apply_code','apply_day']]
test1=pd.merge(data_behavior,test,how='left',left_on='applyCode',right_on='apply_code')
a=test1.apply_day

